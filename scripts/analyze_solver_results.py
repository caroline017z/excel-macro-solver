"""scripts.analyze_solver_results — Per-phase calc-time breakdown.

Reads the ``__SolverResults`` sheet from a ``*_SOLVED.xlsx`` produced by the
headless macro and prints a fixed-width per-project table plus a portfolio
summary. The macro accumulates cumulative recalc seconds per phase
(DSCR / NPP / Appraisal / Full) into module-level Doubles and writes them
to columns O-R at end-of-project; this script consumes those values so we
can see where solve time is actually going and decide which phase scopes
are worth tightening.

Usage:
    python -m scripts.analyze_solver_results <path_to_solved_xlsx>
    python scripts/analyze_solver_results.py <path_to_solved_xlsx>

Column layout (from SolveHeadless.bas):
    A Project Offset       J Calc Tier
    B Project Name         K GS Retry Limit
    C DSCR                 L Mode (warm/cold)
    D NPP                  M Solve Seconds (total per project)
    E Dev Fee              N Heartbeat UTC
    F Equity Pct           O Calc Secs DSCR
    G IRR Gap              P Calc Secs NPP
    H Appraisal Gap        Q Calc Secs Appr
    I Converged            R Calc Secs Full
"""
from __future__ import annotations

import argparse
import logging
import statistics
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl

log = logging.getLogger(__name__)

SHEET_NAME = "__SolverResults"

# 1-indexed column positions on __SolverResults
COL_OFFSET = 1     # A
COL_NAME = 2       # B
COL_CONVERGED = 9  # I
COL_SOLVE_SEC = 13  # M
COL_DSCR_SEC = 15  # O
COL_NPP_SEC = 16   # P
COL_APPR_SEC = 17  # Q
COL_FULL_SEC = 18  # R


@dataclass(frozen=True)
class ProjectRow:
    """One row of __SolverResults, narrowed to the fields this script uses."""

    name: str
    converged: bool
    solve_secs: float
    dscr_secs: float
    npp_secs: float
    appr_secs: float
    full_secs: float

    @property
    def total_calc_secs(self) -> float:
        return self.dscr_secs + self.npp_secs + self.appr_secs + self.full_secs


def _coerce_float(value: object) -> float:
    """Coerce a cell value to float; missing/blank/non-numeric -> 0.0."""
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return 0.0


def _coerce_bool(value: object) -> bool:
    """Coerce a Converged cell to bool. Excel writes TRUE/FALSE as bool."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def load_results(xlsx_path: Path) -> list[ProjectRow]:
    """Read __SolverResults rows from a _SOLVED.xlsx file.

    Uses ``data_only=True`` so we get cached values rather than formula text;
    Excel recalculates before save in the headless flow, so cached values are
    fresh. Skips the header row (row 1) and any trailing rows where column A
    is blank.
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    log.debug("Opening %s (data_only=True)", xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{SHEET_NAME}' not found in {xlsx_path.name}; "
                f"available: {wb.sheetnames}"
            )
        ws = wb[SHEET_NAME]

        rows: list[ProjectRow] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Tolerate short or blank trailing rows.
            if row is None or len(row) < COL_NAME:
                continue
            offset_cell = row[COL_OFFSET - 1]
            if offset_cell is None or offset_cell == "":
                continue

            name = row[COL_NAME - 1] if len(row) >= COL_NAME else None
            if name is None or str(name).strip() == "":
                continue

            def _at(idx: int) -> object:
                return row[idx - 1] if len(row) >= idx else None

            rows.append(ProjectRow(
                name=str(name).strip(),
                converged=_coerce_bool(_at(COL_CONVERGED)),
                solve_secs=_coerce_float(_at(COL_SOLVE_SEC)),
                dscr_secs=_coerce_float(_at(COL_DSCR_SEC)),
                npp_secs=_coerce_float(_at(COL_NPP_SEC)),
                appr_secs=_coerce_float(_at(COL_APPR_SEC)),
                full_secs=_coerce_float(_at(COL_FULL_SEC)),
            ))

        return rows
    finally:
        wb.close()


def _pct(part: float, whole: float) -> str:
    """Format part/whole as a 1-decimal percent or '—' when whole is 0."""
    if whole <= 0:
        return "  -  "
    return f"{(part / whole) * 100:5.1f}%"


def _quantile(values: list[float], q: float) -> float:
    """Return the q-quantile of values using linear interpolation.

    ``statistics.quantiles`` requires n >= 2 and returns cut points; this
    helper handles the n in {0, 1} edge cases too.
    """
    if not values:
        return 0.0
    if len(values) == 1:
        return values[0]
    sorted_vals = sorted(values)
    pos = q * (len(sorted_vals) - 1)
    lo = int(pos)
    hi = min(lo + 1, len(sorted_vals) - 1)
    frac = pos - lo
    return sorted_vals[lo] * (1 - frac) + sorted_vals[hi] * frac


def render_table(rows: Iterable[ProjectRow]) -> str:
    """Render the per-project + portfolio table as a single string."""
    rows = list(rows)

    header = (
        f"{'Project':<32} "
        f"{'DSCR':>6} {'NPP':>6} {'Appr':>6} {'Full':>6} "
        f"{'Total':>6} {'Solve':>6} "
        f"{'DSCR%':>6} {'NPP%':>5} {'Appr%':>5} {'Full%':>6}"
    )
    sep = (
        f"{'-'*32} "
        f"{'-'*6} {'-'*6} {'-'*6} {'-'*6} "
        f"{'-'*6} {'-'*6} "
        f"{'-'*6} {'-'*5} {'-'*5} {'-'*6}"
    )

    out: list[str] = [header, sep]

    for r in rows:
        total = r.total_calc_secs
        out.append(
            f"{r.name[:32]:<32} "
            f"{r.dscr_secs:6.1f} {r.npp_secs:6.1f} {r.appr_secs:6.1f} {r.full_secs:6.1f} "
            f"{total:6.1f} {r.solve_secs:6.1f} "
            f"{_pct(r.dscr_secs, total):>6} "
            f"{_pct(r.npp_secs, total):>5} "
            f"{_pct(r.appr_secs, total):>5} "
            f"{_pct(r.full_secs, total):>6}"
        )

    # Portfolio totals row.
    p_dscr = sum(r.dscr_secs for r in rows)
    p_npp = sum(r.npp_secs for r in rows)
    p_appr = sum(r.appr_secs for r in rows)
    p_full = sum(r.full_secs for r in rows)
    p_total = p_dscr + p_npp + p_appr + p_full
    p_solve = sum(r.solve_secs for r in rows)

    out.append(sep)
    out.append(
        f"{'PORTFOLIO':<32} "
        f"{p_dscr:6.1f} {p_npp:6.1f} {p_appr:6.1f} {p_full:6.1f} "
        f"{p_total:6.1f} {p_solve:6.1f} "
        f"{_pct(p_dscr, p_total):>6} "
        f"{_pct(p_npp, p_total):>5} "
        f"{_pct(p_appr, p_total):>5} "
        f"{_pct(p_full, p_total):>6}"
    )

    return "\n".join(out)


def render_summary(rows: list[ProjectRow]) -> str:
    """Render the portfolio summary stats block."""
    n = len(rows)
    if n == 0:
        return "\nNo project rows found in __SolverResults.\n"

    converged_n = sum(1 for r in rows if r.converged)
    conv_rate = (converged_n / n) * 100

    p_dscr = sum(r.dscr_secs for r in rows)
    p_npp = sum(r.npp_secs for r in rows)
    p_appr = sum(r.appr_secs for r in rows)
    p_full = sum(r.full_secs for r in rows)
    p_calc = p_dscr + p_npp + p_appr + p_full
    p_solve = sum(r.solve_secs for r in rows)

    def _stats(vals: list[float]) -> tuple[float, float, float]:
        if not vals:
            return 0.0, 0.0, 0.0
        return (
            statistics.fmean(vals),
            statistics.median(vals),
            _quantile(vals, 0.95),
        )

    dscr_m, dscr_med, dscr_p95 = _stats([r.dscr_secs for r in rows])
    npp_m, npp_med, npp_p95 = _stats([r.npp_secs for r in rows])
    appr_m, appr_med, appr_p95 = _stats([r.appr_secs for r in rows])
    full_m, full_med, full_p95 = _stats([r.full_secs for r in rows])

    solve_mean = statistics.fmean(r.solve_secs for r in rows)

    lines: list[str] = []
    lines.append("")
    lines.append("=" * 88)
    lines.append("  Portfolio Summary")
    lines.append("=" * 88)
    lines.append(f"  Projects                : {n}")
    lines.append(f"  Converged               : {converged_n}/{n} ({conv_rate:.1f}%)")
    lines.append(f"  Mean Solve Seconds      : {solve_mean:.1f}")
    lines.append("")
    lines.append(f"  Total recalc seconds    : {p_calc:.1f}  ({_pct(p_calc, p_solve).strip()} of total solve time)")
    lines.append(f"    DSCR                  : {p_dscr:7.1f}  ({_pct(p_dscr, p_calc).strip()} of recalc)")
    lines.append(f"    NPP                   : {p_npp:7.1f}  ({_pct(p_npp, p_calc).strip()} of recalc)")
    lines.append(f"    Appraisal             : {p_appr:7.1f}  ({_pct(p_appr, p_calc).strip()} of recalc)")
    lines.append(f"    Full                  : {p_full:7.1f}  ({_pct(p_full, p_calc).strip()} of recalc)")
    lines.append("")
    lines.append(f"  {'Per-project per phase':<24} {'mean':>8} {'median':>8} {'p95':>8}")
    lines.append(f"  {'-'*24} {'-'*8} {'-'*8} {'-'*8}")
    lines.append(f"  {'DSCR seconds':<24} {dscr_m:8.1f} {dscr_med:8.1f} {dscr_p95:8.1f}")
    lines.append(f"  {'NPP seconds':<24} {npp_m:8.1f} {npp_med:8.1f} {npp_p95:8.1f}")
    lines.append(f"  {'Appraisal seconds':<24} {appr_m:8.1f} {appr_med:8.1f} {appr_p95:8.1f}")
    lines.append(f"  {'Full seconds':<24} {full_m:8.1f} {full_med:8.1f} {full_p95:8.1f}")
    lines.append("")

    return "\n".join(lines)


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="analyze_solver_results",
        description=(
            "Print per-phase calc-time breakdown from a _SOLVED.xlsx "
            "produced by the headless solver."
        ),
    )
    parser.add_argument(
        "xlsx",
        type=Path,
        help="Path to a *_SOLVED.xlsx file containing __SolverResults.",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Enable DEBUG logging.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = _build_arg_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(message)s",
        stream=sys.stderr,
    )

    rows = load_results(args.xlsx)
    if not rows:
        print(f"No project rows found in {args.xlsx} ({SHEET_NAME}).")
        return 1

    print(render_table(rows))
    print(render_summary(rows))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
