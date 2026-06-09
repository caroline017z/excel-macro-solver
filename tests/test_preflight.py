"""Tests for dn38_solver.shadow.preflight.

Synthetic workbooks (built in-memory with openpyxl) cover each error code
without depending on any specific deal model. Integration tests against
real workbook fixtures are gated on the fixtures' presence — they're
skipped when the fixtures aren't accessible (CI without Box mounted, etc).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook

from dn38_solver.shadow.preflight import (
    DEV_FEE_BOUNDS,
    ITERATE_DELTA_CEILING,
    NPP_BOUNDS,
    REQUIRED_SHEETS,
    REQUIRED_PI_CELLS,
    apply_auto_fixes,
    format_preflight_report,
    run_preflight,
)


# --- Fixture builders ---------------------------------------------------

def _baseline_workbook(tmp_path: Path, name: str = "test.xlsm") -> Path:
    """Build a synthetic workbook with all required sheets and a single
    active project at column H. Calc properties set to the SMP-known-good
    shape (calcMode=manual, iterate=True, iterateDelta=0.0001).
    """
    wb = Workbook()
    # Default sheet -> rename to first required sheet
    ws_default = wb.active
    ws_default.title = REQUIRED_SHEETS[0]  # Project Inputs
    for sheet in REQUIRED_SHEETS[1:]:
        wb.create_sheet(sheet)

    # Populate Project Inputs with the cells the macro reads.
    pi = wb["Project Inputs"]
    pi["F2"] = 1                # active project index
    pi["F30"] = 0.0725          # WACC target
    pi["F31"] = 0.08            # Live Appraisal IRR (placeholder)
    pi["F32"] = 0.20            # Dev Fee (master col)
    pi["F36"] = 0.20            # Equity IRR target
    pi["F37"] = 0.18            # Live Levered Pre-Tax IRR
    # One active project at column H, with values inside macro bounds
    pi["H4"] = "Test Project"
    pi["H7"] = 1                # active flag
    pi["H32"] = 0.30            # Dev Fee within bounds
    pi["H38"] = 0.50            # NPP within bounds

    # Calc properties -- match the working SMP shape
    cp = wb.calculation
    cp.calcMode = "manual"
    cp.iterate = True
    cp.iterateDelta = 0.0001

    # openpyxl's Workbook.save writes .xlsx by default; we want .xlsm-shape
    # but save as .xlsx is sufficient for the tests (preflight reads the
    # same calcPr regardless of macro presence).
    out = tmp_path / name.replace(".xlsm", ".xlsx")
    wb.save(out)
    return out


# --- A-tier (calc properties) --------------------------------------------

class TestCalcProperties:
    def test_baseline_passes(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        result = run_preflight(path)
        assert result.ok, format_preflight_report(result)
        assert not any(f.code.startswith("A") for f in result.findings)

    def test_a1_iterate_delta_missing(self, tmp_path):
        wb = Workbook()
        wb.active.title = "Project Inputs"
        for s in REQUIRED_SHEETS[1:]:
            wb.create_sheet(s)
        wb.calculation.iterate = True
        wb.calculation.iterateDelta = None  # the IL TEST regression
        wb.calculation.calcMode = "manual"
        # Min cells so B-tier passes
        for cell, _ in REQUIRED_PI_CELLS.items():
            wb["Project Inputs"][cell] = 0.1
        wb["Project Inputs"]["H4"] = "P"
        wb["Project Inputs"]["H7"] = 1
        wb["Project Inputs"]["H32"] = 0.20
        wb["Project Inputs"]["H38"] = 0.50
        path = tmp_path / "missing_delta.xlsx"
        wb.save(path)

        result = run_preflight(path)
        codes = {f.code for f in result.findings}
        assert "A1" in codes
        a1 = next(f for f in result.findings if f.code == "A1")
        assert a1.severity == "error"
        assert a1.auto_fixable is True

    def test_a1_iterate_delta_too_loose(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        # Reopen, set iterateDelta above ceiling, save
        wb = openpyxl.load_workbook(path)
        wb.calculation.iterateDelta = 0.001  # 10x looser than ceiling
        wb.save(path)
        result = run_preflight(path)
        assert any(f.code == "A1" for f in result.findings)

    def test_a2_iterate_disabled(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        wb.calculation.iterate = False
        wb.save(path)
        result = run_preflight(path)
        codes = {f.code for f in result.findings}
        assert "A2" in codes


# --- B-tier (structure) --------------------------------------------------

class TestStructure:
    def test_b5_missing_sheet(self, tmp_path):
        wb = Workbook()
        wb.active.title = "Project Inputs"
        # Skip half the required sheets to trigger B5
        for s in REQUIRED_SHEETS[1:5]:
            wb.create_sheet(s)
        wb.calculation.iterate = True
        wb.calculation.iterateDelta = 0.0001
        wb.calculation.calcMode = "manual"
        for cell, _ in REQUIRED_PI_CELLS.items():
            wb["Project Inputs"][cell] = 0.1
        wb["Project Inputs"]["H4"] = "P"
        wb["Project Inputs"]["H7"] = 1
        path = tmp_path / "missing_sheets.xlsx"
        wb.save(path)
        result = run_preflight(path)
        assert any(f.code == "B5" for f in result.findings)

    def test_b7_missing_required_cell(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        wb["Project Inputs"]["F31"] = None  # blank a required cell
        wb.save(path)
        result = run_preflight(path)
        b7 = [f for f in result.findings if f.code == "B7"]
        assert len(b7) == 1
        assert "F31" in b7[0].location

    def test_b9_no_active_projects(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        wb["Project Inputs"]["H7"] = 0  # turn off the only active project
        wb.save(path)
        result = run_preflight(path)
        assert any(f.code == "B9" for f in result.findings)

    def test_b9_scans_full_project_col_range_not_just_h_to_s(self, tmp_path):
        """B9 must scan H..BG (the full PI_PROJECT_COL_RANGE), not just
        H..S. The old implementation hardcoded H..S (cols 8..19) and
        silently missed every active toggle past col S — a 15-project
        portfolio with toggles at V..AJ (cols 22..36) false-positive'd
        as "no active projects." Caught on 2026-05-19 SolarStone re-run.

        Fixture: turn off H..S entirely, set V7 = 1 (a column the old
        range didn't cover). B9 must NOT fire.
        """
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Project Inputs"]
        # Clear every toggle in the old H..S range
        for col_letter in ("H", "I", "J", "K", "L", "M", "N", "O", "P",
                           "Q", "R", "S"):
            ws[f"{col_letter}7"] = 0
        # Set a toggle past the old scan boundary
        ws["V7"] = 1
        # Plant a project name on V so the row reads as populated; B9
        # is permissive about column names, but populated rows make the
        # fixture realistic.
        ws["V4"] = "Mock project past col S"
        wb.save(path)
        result = run_preflight(path)
        b9 = [f for f in result.findings if f.code == "B9"]
        assert b9 == [], (
            f"B9 must NOT fire when V7=1 — preflight scans full "
            f"PI_PROJECT_COL_RANGE (H..BG). Old H..S hardcode regressed. "
            f"Got: {b9}"
        )


# --- E-tier (input bounds) -----------------------------------------------

class TestInputBounds:
    def test_e13_dev_fee_above_max(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        wb["Project Inputs"]["H32"] = 6.00  # above DEV_FEE_MAX=5.5 (raised 2026-05-15)
        wb.save(path)
        result = run_preflight(path)
        e13 = [f for f in result.findings if f.code == "E13"]
        assert len(e13) == 1
        assert "6.00" in e13[0].message
        # E13 is a warning, not a blocking error — SMP empirically converges
        # with E13 firing, so we only flag it as worth investigating.
        assert e13[0].severity == "warning"

    def test_e13_dev_fee_below_min(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        wb["Project Inputs"]["H32"] = 0.01
        wb.save(path)
        result = run_preflight(path)
        assert any(f.code == "E13" for f in result.findings)

    def test_e14_npp_above_max(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        wb["Project Inputs"]["H38"] = 2.50  # > NPP_MAX = 2.0 (raised 2026-05-14)
        wb.save(path)
        result = run_preflight(path)
        e14 = [f for f in result.findings if f.code == "E14"]
        assert len(e14) == 1
        assert e14[0].severity == "warning"

    def test_e15a_custom_with_empty_rate_name(self, tmp_path):
        """E15a fires when RC sub-block Toggle='Custom' but Rate Name is empty.

        Reproduces Queen City 2026-05-15 failure mode where cols O-T had RC5
        Toggle='Custom' with no Rate Name, producing $4-5/W Dev Fees instead
        of the expected $1.50-$2.00/W.
        """
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Project Inputs"]
        # Master Equity toggle ON for RC5 (row 154). Required since
        # the SolarStone 2026-05-18 fix — E15a/b/c skip (col, rc) pairs
        # where the master toggle = 0 (RC inactive for Equity, so the
        # sub-block content is cosmetic).
        ws["H154"] = 1
        # RC5 PI sub-block base = 197 → Rate Name=198, Toggle=199
        ws["H198"] = None         # empty Rate Name
        ws["H199"] = "Custom"     # Custom toggle
        wb.save(path)

        result = run_preflight(path)
        e15a = [f for f in result.findings if f.code == "E15a"]
        assert len(e15a) == 1, f"expected E15a; got: {[f.code for f in result.findings]}"
        assert "RC5" in e15a[0].message
        # E15a is severity=warning (informational). It fires whenever any
        # Custom toggle has an empty Rate Name, but uniform Custom-empty
        # across all projects is harmless (intentional RC disable, SMP
        # 2026-05-18 pattern). The ERROR-severity check for the truly
        # dangerous case is E15c (asymmetric revenue) — see below.
        assert e15a[0].severity == "warning"

    def test_e15b_mixed_modes_across_projects(self, tmp_path):
        """E15b fires when same RC has different Toggle values across active
        projects in the same workbook (Queen City pattern: RC5 = Generic on
        H-N, Custom on O-T)."""
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Project Inputs"]
        # Add a second active project at col I with mismatched RC5 toggle.
        ws["I4"] = "P2"
        ws["I7"] = 1
        ws["I32"] = 0.30
        ws["I38"] = 0.50
        # Master Equity toggle ON for RC5 on both projects (row 154).
        ws["H154"] = 1
        ws["I154"] = 1
        # RC5: H=Generic with name, I=Custom with name (both populated, just mixed)
        ws["H198"] = "Merchant Rate"
        ws["H199"] = "Generic"
        ws["I198"] = "Custom Tariff"
        ws["I199"] = "Custom"
        wb.save(path)

        result = run_preflight(path)
        e15b = [f for f in result.findings if f.code == "E15b"]
        assert len(e15b) == 1, f"expected E15b; got: {[f.code for f in result.findings]}"
        assert "RC5" in e15b[0].message

    def test_e15_clean_workbook(self, tmp_path):
        """Baseline workbook (no RC sub-blocks populated) doesn't trip E15."""
        path = _baseline_workbook(tmp_path)
        result = run_preflight(path)
        codes = {f.code for f in result.findings}
        # All RC toggles are None on baseline → not "Custom" → no E15a.
        # All projects have identical (None) toggle → no E15b.
        assert "E15a" not in codes
        assert "E15b" not in codes
        assert "E15c" not in codes

    def test_e15c_asymmetric_revenue_blocks_solve(self, tmp_path):
        """E15c fires as ERROR on the Queen City 2026-05-15 pattern: some
        projects have Generic + non-zero rate, others have Custom + empty
        Name on the same RC. This is what produced the wrong-but-valid
        $4-5/W Dev Fees that motivated the whole E15 family."""
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Project Inputs"]
        # Master Equity toggle ON for RC5 on both projects (row 154).
        ws["H154"] = 1
        ws["I154"] = 1
        # Project H: RC5 Generic with 5.5% merchant rate (Queen City H-N pattern)
        ws["H198"] = "Merchant Rate"
        ws["H199"] = "Generic"
        ws["H200"] = 0.055
        # Project I: active, RC5 Custom with empty Name (Queen City O-T pattern)
        ws["I4"] = "P2"
        ws["I7"] = 1
        ws["I32"] = 0.30
        ws["I38"] = 0.50
        ws["I198"] = None
        ws["I199"] = "Custom"
        wb.save(path)

        result = run_preflight(path)
        e15c = [f for f in result.findings if f.code == "E15c"]
        assert len(e15c) == 1, f"expected E15c; got: {[f.code for f in result.findings]}"
        assert "RC5" in e15c[0].message
        assert e15c[0].severity == "error"

    def test_e15_master_off_suppresses_all_three_checks(self, tmp_path):
        """SolarStone 2026-05-18 fix: when the Equity master toggle for an
        RC is 0 (row 150-155), E15a/b/c must skip that RC entirely — the
        per-block sub-block content is cosmetic when the master is off.

        Without this, a workbook with the master RC5 turned off on all
        projects but leftover template values in rows 197-215 produces
        a false-positive E15c block on a perfectly valid solve. (The
        SolarStone next-wave run hit exactly this — 25 projects with
        RC5 master-off but rows 197-203 retained template merchant-rate
        config, and the run was blocked until the operator overrode.)
        """
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Project Inputs"]
        # Master Equity toggle OFF for RC5 (row 154) — the value of 0
        # means RC5 contributes zero revenue regardless of the sub-block.
        ws["H154"] = 0
        # Even with a fully-populated Generic + revenue config on H,
        # and an asymmetric Custom-empty on I, the master-off should
        # suppress all three E15 checks for RC5.
        ws["H198"] = "Merchant Rate"
        ws["H199"] = "Generic"
        ws["H200"] = 0.055  # would normally trigger E15c
        ws["I4"] = "P2"
        ws["I7"] = 1
        ws["I32"] = 0.30
        ws["I38"] = 0.50
        ws["I154"] = 0      # master OFF on I too
        ws["I198"] = None
        ws["I199"] = "Custom"  # would normally trigger E15a/E15b
        wb.save(path)

        result = run_preflight(path)
        codes = {f.code for f in result.findings}
        assert "E15a" not in codes, (
            f"E15a fired despite master-off; got: "
            f"{[(f.code, f.message) for f in result.findings if f.code.startswith('E15')]}"
        )
        assert "E15b" not in codes
        assert "E15c" not in codes

    def test_e15c_uniform_custom_empty_does_not_fire(self, tmp_path):
        """E15c must NOT fire when ALL active projects have Custom + empty
        on the same RC (uniform disable, SMP 2026-05-18 pattern). Without
        any Generic-with-revenue project, there's no asymmetry."""
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Project Inputs"]
        # Master Equity toggle ON for RC5 on both projects so the E15
        # checks actually evaluate the sub-blocks (uniform Custom-empty
        # is the intentional-disable case the test asserts is harmless).
        ws["H154"] = 1
        ws["I154"] = 1
        # Project H: RC5 Custom + empty (uniform pattern)
        ws["H198"] = None
        ws["H199"] = "Custom"
        # Project I: also RC5 Custom + empty
        ws["I4"] = "P2"
        ws["I7"] = 1
        ws["I32"] = 0.30
        ws["I38"] = 0.50
        ws["I198"] = None
        ws["I199"] = "Custom"
        wb.save(path)

        result = run_preflight(path)
        e15c = [f for f in result.findings if f.code == "E15c"]
        assert e15c == [], (
            "E15c should NOT fire on uniform Custom-empty (SMP pattern). "
            f"Got: {[(f.code, f.message) for f in result.findings]}"
        )
        # E15a (Custom-with-empty-Name) still fires as a warning for visibility
        e15a = [f for f in result.findings if f.code == "E15a"]
        assert len(e15a) == 1
        assert e15a[0].severity == "warning"

    def test_inactive_projects_ignored(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        # Out-of-range Dev Fee on an INACTIVE project should not flag
        wb["Project Inputs"]["I7"] = 0
        wb["Project Inputs"]["I32"] = 99.0
        wb.save(path)
        result = run_preflight(path)
        e13 = [f for f in result.findings if f.code == "E13"]
        assert e13 == []

    def test_e16_placeholder_project_detected(self, tmp_path):
        """E16 fires when active project has RC1 Generic + Generic Rate=0.

        Reproduces SMP 2026-05-18 failure: cols N..W had Generic toggle with
        zero Generic Energy Rate at COD, crashed both workers at first
        attempt with HRESULT 0x800a9c68."""
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Project Inputs"]
        # Project H baseline: active but RC1 Generic + Rate=0 = placeholder
        ws["H159"] = "Generic"
        ws["H160"] = 0
        wb.save(path)

        result = run_preflight(path)
        e16 = [f for f in result.findings if f.code == "E16"]
        assert len(e16) == 1, f"expected E16; got: {[f.code for f in result.findings]}"
        assert "H=" in e16[0].message
        assert e16[0].severity == "warning"

    def test_e16_does_not_fire_on_real_revenue(self, tmp_path):
        """E16 must NOT fire when Generic toggle has a non-zero Generic Rate.
        Real projects with merchant rate stay quiet."""
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Project Inputs"]
        ws["H159"] = "Generic"
        ws["H160"] = 0.055  # 5.5% merchant rate
        wb.save(path)

        result = run_preflight(path)
        e16 = [f for f in result.findings if f.code == "E16"]
        assert e16 == []

    def test_e16_does_not_fire_on_custom_toggle(self, tmp_path):
        """E16 is RC1-Generic-specific. Custom toggle (revenue from Rate
        Curves tab vector) doesn't trigger even with zero/empty Generic Rate."""
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Project Inputs"]
        ws["H158"] = "GH25 -15%"  # Custom rate name
        ws["H159"] = "Custom"
        ws["H160"] = None  # Custom toggle reads from Rate Curves, not this cell
        wb.save(path)

        result = run_preflight(path)
        e16 = [f for f in result.findings if f.code == "E16"]
        assert e16 == []

    def test_e16_inactive_projects_ignored(self, tmp_path):
        """Placeholder shape on an INACTIVE project (row 7 != 1) doesn't fire."""
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Project Inputs"]
        # Add an inactive column with placeholder shape
        ws["I7"] = 0  # inactive
        ws["I159"] = "Generic"
        ws["I160"] = 0
        wb.save(path)

        result = run_preflight(path)
        e16 = [f for f in result.findings if f.code == "E16"]
        assert e16 == []


# --- D-tier (macro version) ----------------------------------------------

class TestMacroVersion:
    """D15/D16 scan vbaProject.bin in the .xlsm. Synthetic .xlsx test files
    have no VBA project so D15 fires with the 'no VBA project' branch."""

    def test_d_tier_skipped_for_xlsx(self, tmp_path):
        """Pure .xlsx workbooks are macro-free by spec; D-tier must not
        fire on them so the rest of the preflight remains useful for
        non-macro testing fixtures."""
        path = _baseline_workbook(tmp_path)  # .xlsx
        result = run_preflight(path)
        assert not [f for f in result.findings if f.code in ("D15", "D16")]

    def test_d15_no_vba_in_xlsm(self, tmp_path):
        """An .xlsm without xl/vbaProject.bin is malformed; D15 fires."""
        import zipfile as zf
        path = _baseline_workbook(tmp_path)
        xlsm = path.with_suffix(".xlsm")
        with zf.ZipFile(path, "r") as zin, zf.ZipFile(xlsm, "w") as zout:
            for item in zin.infolist():
                zout.writestr(item, zin.read(item.filename))
        result = run_preflight(xlsm)
        d15 = [f for f in result.findings if f.code == "D15"]
        assert len(d15) == 1
        assert "no VBA project" in d15[0].message
        assert d15[0].severity == "error"

    def test_d15_missing_functions_detected(self, tmp_path):
        """Construct a fake .xlsm with a vbaProject.bin containing only
        SOME of the required functions. Confirm D15 lists the missing ones.
        """
        import zipfile as zf
        # Start from the _baseline_workbook (.xlsx) and add a fake vbaProject.bin
        path = _baseline_workbook(tmp_path)
        xlsm = path.with_suffix(".xlsm")
        # Build a minimal vba binary that mentions only 3 of the required
        # functions (and SolveHeadless to avoid the no-VBA branch above).
        partial_vba = (
            b"\x00SolveHeadless\x00 ... \x00InitSolveEnvHL\x00 ... "
            b"\x00SolveOneProjectByColHL\x00"
        )
        # Repackage the .xlsx as .xlsm with the fake VBA
        with zf.ZipFile(path, "r") as zin, zf.ZipFile(xlsm, "w") as zout:
            for item in zin.infolist():
                zout.writestr(item, zin.read(item.filename))
            zout.writestr("xl/vbaProject.bin", partial_vba)
        result = run_preflight(xlsm)
        d15 = [f for f in result.findings if f.code == "D15"]
        assert len(d15) == 1
        # Missing FinalizeSolveEnvHL, CalcSheetsForAppraisal, etc.
        for missing in ("FinalizeSolveEnvHL", "CalcSheetsForAppraisal",
                        "ClassifyConvergenceHL", "StampActiveProjectColumnHL"):
            assert missing in d15[0].message, (
                f"Expected {missing} in D15 message: {d15[0].message}"
            )
        # Tranche 7.13: D15 (missing-functions variant) is auto_fixable
        # — orchestrator handles via reimport_macro_subprocess on the
        # _FIXED.xlsm sibling under --auto-fix.
        assert d15[0].auto_fixable is True, (
            "D15 (missing functions) must be auto_fixable=True per "
            "Tranche 7.13 — orchestrator routes a macro re-import into "
            "the _FIXED.xlsm sibling under --auto-fix"
        )

    def test_d16_stale_modules_detected(self, tmp_path):
        import zipfile as zf
        path = _baseline_workbook(tmp_path)
        xlsm = path.with_suffix(".xlsm")
        # Include every currently-required function (derived from the live
        # REQUIRED_MACRO_FUNCTIONS so this fixture can't go stale when the
        # contract grows — C4) + a stale Module2_Optimized name.
        from dn38_solver.shadow.preflight import REQUIRED_MACRO_FUNCTIONS
        full_vba = b"\x00".join(
            name.encode("ascii")
            for name in (*REQUIRED_MACRO_FUNCTIONS, "Module2_Optimized3")
        )
        with zf.ZipFile(path, "r") as zin, zf.ZipFile(xlsm, "w") as zout:
            for item in zin.infolist():
                zout.writestr(item, zin.read(item.filename))
            zout.writestr("xl/vbaProject.bin", full_vba)
        result = run_preflight(xlsm)
        d16 = [f for f in result.findings if f.code == "D16"]
        assert len(d16) == 1
        assert d16[0].severity == "warning"
        assert "Module2_Optimized" in d16[0].message
        # And no D15 since all required functions present
        assert not [f for f in result.findings if f.code == "D15"]


# --- Auto-fix ------------------------------------------------------------

class TestAutoFix:
    def test_a1_patched_into_fixed_copy(self, tmp_path):
        # Build a workbook with iterateDelta missing
        wb = Workbook()
        wb.active.title = "Project Inputs"
        for s in REQUIRED_SHEETS[1:]:
            wb.create_sheet(s)
        wb.calculation.iterate = True
        wb.calculation.iterateDelta = None
        wb.calculation.calcMode = "manual"
        for cell, _ in REQUIRED_PI_CELLS.items():
            wb["Project Inputs"][cell] = 0.1
        wb["Project Inputs"]["H4"] = "P"
        wb["Project Inputs"]["H7"] = 1
        wb["Project Inputs"]["H32"] = 0.20
        wb["Project Inputs"]["H38"] = 0.50
        src = tmp_path / "broken.xlsx"
        wb.save(src)

        result = run_preflight(src)
        a1 = [f for f in result.findings if f.code == "A1"]
        assert len(a1) == 1
        assert a1[0].auto_fixable

        dst = tmp_path / "fixed.xlsx"
        out_path, applied = apply_auto_fixes(src, dst, result.findings)
        assert out_path == dst
        assert "A1" in applied

        # Source file untouched
        src_check = run_preflight(src)
        assert any(f.code == "A1" for f in src_check.findings)
        # Fixed file passes A1
        fixed_check = run_preflight(dst)
        assert not any(f.code == "A1" for f in fixed_check.findings)

    def test_apply_auto_fixes_skips_unfixable(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        wb["Project Inputs"]["F31"] = None  # B7 - not auto-fixable
        wb.save(path)
        result = run_preflight(path)
        b7 = [f for f in result.findings if f.code == "B7"]
        assert b7 and not b7[0].auto_fixable

        dst = tmp_path / "fixed.xlsx"
        _, applied = apply_auto_fixes(path, dst, result.findings)
        assert "B7" not in applied


# --- Reporting -----------------------------------------------------------

class TestReport:
    def test_report_format_pass(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        result = run_preflight(path)
        text = format_preflight_report(result)
        assert "Pre-flight: PASS" in text

    def test_report_format_fail_includes_codes(self, tmp_path):
        path = _baseline_workbook(tmp_path)
        wb = openpyxl.load_workbook(path)
        wb.calculation.iterateDelta = None
        wb["Project Inputs"]["H32"] = 6.0  # > DEV_FEE_MAX = 5.5 (raised 2026-05-15)
        wb.save(path)
        result = run_preflight(path)
        text = format_preflight_report(result)
        assert "FAIL" in text
        assert "A1" in text
        assert "E13" in text
        assert "remediation" not in text  # internal field name shouldn't leak
        assert "fix:" in text


# --- Integration against real fixtures (skipped if not present) ---------

REAL_IL_TEST = Path(
    r"C:\Users\CarolineZepecki\Desktop"
    r"\38DN-IL_US Solar_PricingModel_PV Only_2026.05.13_TEST.xlsm"
)
REAL_SMP = Path(
    r"C:\Users\CarolineZepecki\Desktop\Archive"
    r"\38DN-SMP_PricingModel_2026.05.12_WalkTEST.xlsm"
)


@pytest.mark.skipif(not REAL_IL_TEST.exists(), reason="IL TEST fixture not available")
def test_real_il_test_fails_preflight():
    """Regression: the IL TEST 2026-05-13 file must FAIL preflight on
    BOTH:
      - D15: the embedded macro is the 627-line outdated version,
        missing 11 of the 12 marker functions (THE actual root cause).
      - A1: the workbook XML is missing iterateDelta=0.0001 (a real
        defect even if not the immediate cause of non-convergence).
    """
    result = run_preflight(REAL_IL_TEST)
    assert not result.ok, format_preflight_report(result)
    codes = {f.code for f in result.errors}
    assert "D15" in codes, (
        f"Expected D15 (outdated macro) in errors, got {codes}\n\n"
        f"{format_preflight_report(result)}"
    )
    assert "A1" in codes, (
        f"Expected A1 (missing iterateDelta) in errors, got {codes}\n\n"
        f"{format_preflight_report(result)}"
    )


@pytest.mark.skipif(not REAL_SMP.exists(), reason="SMP fixture not available")
def test_real_smp_passes_preflight():
    """Regression: the SMP WalkTEST file (which solves cleanly end-to-end)
    must NOT produce any preflight ERROR-severity findings. Warnings are
    allowed (e.g., E13 if Dev Fee just barely exceeds DEV_FEE_MAX in some
    project columns)."""
    result = run_preflight(REAL_SMP)
    if not result.ok:
        # Surface what failed for diagnosis.
        pytest.fail(
            "SMP fixture failed preflight unexpectedly:\n"
            + format_preflight_report(result)
        )


# --- D17 macro hash drift ------------------------------------------------

import zipfile
from dn38_solver.shadow.preflight import (
    BAS_HASH_PROP,
    _current_bas_sha256,
)


def _baseline_xlsm(tmp_path: Path, name: str = "macro.xlsm") -> Path:
    """Create a baseline .xlsm-named file (synthetic; no real vbaProject).

    D15 will fire ("no VBA project") but D17 logic still runs against the
    docProps/custom.xml read path, which is what these tests target. We
    filter by code so the D15 noise doesn't mask the assertion.
    """
    path = _baseline_workbook(tmp_path, name=name.replace(".xlsm", ".xlsx"))
    xlsm_path = path.with_suffix(".xlsm")
    path.rename(xlsm_path)
    return xlsm_path


def _inject_bas_hash_stamp(xlsm_path: Path, hash_value: str) -> None:
    """Add docProps/custom.xml with a DN38_BAS_SHA256 property.

    Mirrors what Excel COM CustomDocumentProperties.Add would write. Done
    by rewriting the zip rather than via openpyxl to avoid the openpyxl-
    xlsm save round-trip pattern (which is exactly what the rule under
    test forbids).
    """
    custom_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Properties xmlns='
        '"http://schemas.openxmlformats.org/officeDocument/2006/custom-properties" '
        'xmlns:vt='
        '"http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        f'<property fmtid="{{D5CDD505-2E9C-101B-9397-08002B2CF9AE}}" pid="2" '
        f'name="{BAS_HASH_PROP}">'
        f'<vt:lpwstr>{hash_value}</vt:lpwstr>'
        '</property>'
        '</Properties>'
    )
    # zipfile doesn't have an in-place edit primitive; rewrite the archive
    # with the new part appended (overwriting any prior copy).
    tmp_out = xlsm_path.with_suffix(".xlsm.tmp")
    with zipfile.ZipFile(xlsm_path, "r") as zin:
        with zipfile.ZipFile(tmp_out, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == "docProps/custom.xml":
                    continue
                zout.writestr(item, zin.read(item.filename))
            zout.writestr("docProps/custom.xml", custom_xml)
    tmp_out.replace(xlsm_path)


class TestMacroHashDrift:
    def test_no_stamp_produces_d17_warning(self, tmp_path):
        # Skip when the repo doesn't have SolveHeadless.bas (defensive —
        # the .bas should always be present in a working tree, but tests
        # shouldn't crash a stripped checkout).
        if _current_bas_sha256() is None:
            pytest.skip("SolveHeadless.bas not in repo")
        path = _baseline_xlsm(tmp_path)
        result = run_preflight(path)
        d17 = [f for f in result.findings if f.code == "D17"]
        assert len(d17) == 1
        assert d17[0].severity == "warning"
        assert "no" in d17[0].message.lower() or "stamp" in d17[0].message.lower()

    def test_matching_stamp_no_d17_finding(self, tmp_path):
        current = _current_bas_sha256()
        if current is None:
            pytest.skip("SolveHeadless.bas not in repo")
        path = _baseline_xlsm(tmp_path)
        _inject_bas_hash_stamp(path, current)
        result = run_preflight(path)
        d17 = [f for f in result.findings if f.code == "D17"]
        assert d17 == [], (
            f"Expected no D17 finding when hash matches; got: {d17}"
        )

    def test_mismatched_stamp_produces_d17_error(self, tmp_path):
        if _current_bas_sha256() is None:
            pytest.skip("SolveHeadless.bas not in repo")
        path = _baseline_xlsm(tmp_path)
        # Plant a hash that definitely won't match the real .bas
        _inject_bas_hash_stamp(path, "0" * 64)
        result = run_preflight(path)
        d17 = [f for f in result.findings if f.code == "D17"]
        assert len(d17) == 1
        assert d17[0].severity == "error"
        assert "does not match" in d17[0].message
        # Tranche 7.13: hash-drift D17 is auto_fixable. Orchestrator
        # re-imports the macro into the _FIXED.xlsm sibling under
        # --auto-fix so the source workbook is never mutated.
        assert d17[0].auto_fixable is True, (
            "D17 (hash drift error) must be auto_fixable=True per "
            "Tranche 7.13 — operator must not need a separate flag for "
            "the common macro-drift case"
        )

    def test_no_stamp_d17_is_auto_fixable(self, tmp_path):
        """The 'no stamp' warning variant of D17 is also auto_fixable —
        a macro re-import plants the stamp as a side effect, so
        --auto-fix should clear this warning automatically.
        """
        if _current_bas_sha256() is None:
            pytest.skip("SolveHeadless.bas not in repo")
        path = _baseline_xlsm(tmp_path)
        result = run_preflight(path)
        d17 = [f for f in result.findings if f.code == "D17"]
        assert len(d17) == 1
        assert d17[0].severity == "warning"
        assert d17[0].auto_fixable is True, (
            "D17 (no-stamp warning) must be auto_fixable=True per "
            "Tranche 7.13"
        )

    def test_xlsx_skips_macro_hash_check(self, tmp_path):
        # .xlsx is macro-free by spec — D17 must not fire on it even
        # without a stamp, otherwise every non-macro test workbook would
        # surface a spurious warning.
        path = _baseline_workbook(tmp_path)  # .xlsx
        result = run_preflight(path)
        assert not any(f.code == "D17" for f in result.findings)


# --- D16 v2 (PROJECT-stream live-module list) + D18 (signature drift) -----

class TestVbaDecompress:
    """MS-OVBA 2.4.1 CompressedContainer decompression — hand-built
    vectors covering the literal, copy-token, and overlapping-copy paths."""

    def test_rejects_non_container(self):
        from dn38_solver.shadow.preflight import _decompress_vba
        with pytest.raises(ValueError):
            _decompress_vba(b"\x02whatever")
        with pytest.raises(ValueError):
            _decompress_vba(b"")

    def test_uncompressed_chunk(self):
        from dn38_solver.shadow.preflight import _decompress_vba
        # Raw chunk: header 0x3FFF (field=0xFFF, not compressed) + 4096 bytes
        payload = bytes(range(256)) * 16
        data = b"\x01" + (0x3FFF).to_bytes(2, "little") + payload
        assert _decompress_vba(data) == payload

    def test_compressed_chunk_with_copy_token(self):
        from dn38_solver.shadow.preflight import _decompress_vba
        # literals 'a','b','c' then copy token offset=3 len=6 -> "abcabcabc"
        # flag byte 0b00001000 (4th token is a copy); token = (offset-1)<<12 | (len-3)
        chunk = b"\x08abc" + (0x2003).to_bytes(2, "little")
        header = (len(chunk) + 2 - 3) | 0x3000 | 0x8000
        data = b"\x01" + header.to_bytes(2, "little") + chunk
        assert _decompress_vba(data) == b"abcabcabc"

    def test_overlapping_copy(self):
        from dn38_solver.shadow.preflight import _decompress_vba
        # literal 'a' then copy offset=1 len=7 -> "aaaaaaaa"
        chunk = b"\x02a" + (0x0004).to_bytes(2, "little")
        header = (len(chunk) + 2 - 3) | 0x3000 | 0x8000
        data = b"\x01" + header.to_bytes(2, "little") + chunk
        assert _decompress_vba(data) == b"a" * 8


class TestParamCounts:
    def test_basic_signatures(self):
        from dn38_solver.shadow.preflight import _parse_param_counts
        src = (
            "Public Sub NoArgs()\r\nEnd Sub\r\n"
            "Private Function TwoArgs(ByVal a As Integer, b As String) As Long\r\n"
            "End Function\r\n"
            "Sub OneArg(ByVal colIdx As Integer)\r\nEnd Sub\r\n"
        )
        counts = _parse_param_counts(src)
        assert counts == {"NoArgs": 0, "TwoArgs": 2, "OneArg": 1}

    def test_line_continuation_folded(self):
        from dn38_solver.shadow.preflight import _parse_param_counts
        src = (
            "Public Sub StampActiveProjectColumnHL(ByVal colIdx As Integer, _\r\n"
            "                                      ByVal dscrRestore As Double)\r\n"
            "End Sub\r\n"
        )
        assert _parse_param_counts(src) == {"StampActiveProjectColumnHL": 2}

    def test_array_param_does_not_truncate(self):
        from dn38_solver.shadow.preflight import _parse_param_counts
        src = "Sub TakesArray(ByRef arr() As Double, ByVal n As Integer)\r\nEnd Sub\r\n"
        assert _parse_param_counts(src) == {"TakesArray": 2}

    def test_call_sites_not_matched(self):
        from dn38_solver.shadow.preflight import _parse_param_counts
        # A call inside a body must not register as a definition.
        src = (
            "Sub Caller()\r\n"
            "    DoWork 1, 2\r\n"
            "    x = Helper(3)\r\n"
            "End Sub\r\n"
        )
        assert _parse_param_counts(src) == {"Caller": 0}


class TestD16ProjectStream:
    """D16 must trust the PROJECT stream's live-module list when present:
    removed modules leave residual name strings in the binary's identifier
    table, and the legacy substring scan false-positives on them
    (observed on SolarStone 2026-06-04 after VBProject cleanup)."""

    def _xlsm_with_vba(self, tmp_path, vba_bin: bytes):
        import zipfile as zf
        path = _baseline_workbook(tmp_path)
        xlsm = path.with_suffix(".xlsm")
        with zf.ZipFile(path, "r") as zin, zf.ZipFile(xlsm, "w") as zout:
            for item in zin.infolist():
                zout.writestr(item, zin.read(item.filename))
            zout.writestr("xl/vbaProject.bin", vba_bin)
        return xlsm

    _ALL_REQUIRED = (
        b"\x00".join(
            name.encode("ascii")
            for name in (
                "SolveHeadless", "InitSolveEnvHL", "SolveOneProjectByColHL",
                "FinalizeSolveEnvHL", "CalcSheetsForAppraisal",
                "CalcSheetsForNPP", "CalcSheetsForDSCR",
                "ClassifyConvergenceHL", "StampActiveProjectColumnHL",
                "ProjectElapsedHL", "HardStampNumericHL",
                "SetSkipOutputRecalcHL",
            )
        )
    )

    def test_residual_names_ignored_when_project_stream_clean(self, tmp_path):
        # PROJECT stream lists only modSolveHeadless; the stale names appear
        # elsewhere in the binary (identifier-table residue). No D16.
        vba = (
            self._ALL_REQUIRED
            + b"\r\nModule=modSolveHeadless\r\n"
            + b"\x00residue\x00Module2_Optimized\x00Module3\x00Module4\x00"
        )
        result = run_preflight(self._xlsm_with_vba(tmp_path, vba))
        assert not [f for f in result.findings if f.code == "D16"], (
            "D16 must not fire on identifier-table residue when the "
            "PROJECT stream shows the module was removed"
        )

    def test_live_stale_module_still_flagged(self, tmp_path):
        # PROJECT stream still lists the stale module -> D16 fires, and the
        # numeric-suffix family match catches Module2_Optimized3.
        vba = (
            self._ALL_REQUIRED
            + b"\r\nModule=modSolveHeadless\r\nModule=Module2_Optimized3\r\n"
        )
        result = run_preflight(self._xlsm_with_vba(tmp_path, vba))
        d16 = [f for f in result.findings if f.code == "D16"]
        assert len(d16) == 1
        assert "Module2_Optimized" in d16[0].message

    def test_legacy_fallback_without_project_stream(self, tmp_path):
        # No Module= lines anywhere -> fall back to the substring scan
        # (keeps synthetic-bin behavior; see test_d16_stale_modules_detected).
        vba = self._ALL_REQUIRED + b"\x00Module2_Optimized\x00"
        result = run_preflight(self._xlsm_with_vba(tmp_path, vba))
        assert [f for f in result.findings if f.code == "D16"]


class TestMacroSignatures:
    """D18 — parameter-count drift between embedded macro and repo .bas."""

    def test_skips_when_extraction_impossible(self, tmp_path):
        # Synthetic bins aren't valid OLE containers; extraction returns {}
        # and D18 stays silent rather than guessing.
        import zipfile as zf
        path = _baseline_workbook(tmp_path)
        xlsm = path.with_suffix(".xlsm")
        with zf.ZipFile(path, "r") as zin, zf.ZipFile(xlsm, "w") as zout:
            for item in zin.infolist():
                zout.writestr(item, zin.read(item.filename))
            zout.writestr("xl/vbaProject.bin", b"\x00SolveHeadless\x00garbage")
        from dn38_solver.shadow.preflight import check_macro_signatures
        assert check_macro_signatures(xlsm) == []

    def test_skips_xlsx(self, tmp_path):
        from dn38_solver.shadow.preflight import check_macro_signatures
        assert check_macro_signatures(_baseline_workbook(tmp_path)) == []

    def test_mismatch_detection_logic(self):
        # The comparison core: same name, different arity -> flagged.
        from dn38_solver.shadow.preflight import _parse_param_counts
        embedded = _parse_param_counts(
            "Public Sub StampActiveProjectColumnHL(ByVal colIdx As Integer)\r\nEnd Sub\r\n"
        )
        repo = _parse_param_counts(
            "Public Sub StampActiveProjectColumnHL(ByVal colIdx As Integer, _\r\n"
            "    ByVal dscrRestore As Double)\r\nEnd Sub\r\n"
        )
        name = "StampActiveProjectColumnHL"
        assert embedded[name] == 1
        assert repo[name] == 2
        assert embedded[name] != repo[name]
