"""Tests for dn38_solver.validation.post_merge.verify_merged_file.

Locks in the Tranche 7.4 fix: the verifier must NOT diff hard-stamped
cells for projects the Tranche 7.2 fast-skip path bypassed. Skipped
projects leave their post-solve cells untouched (in whatever state the
prior project's iterative calc left them — typically `#NUM!` on
placeholder columns), and comparing those against the worker-reported
sentinel-coded error value produced 10 false-positive mismatches on
SMP run id=18 before the fix.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from dn38_solver.types import CellAddress, SolveTask
from dn38_solver.validation.post_merge import verify_merged_file


def _make_merged_workbook(tmp_path: Path) -> Path:
    """Build a minimal Project Inputs sheet with one real + one
    placeholder column. Real (col H) has clean numeric hard-stamps;
    placeholder (col O) has an Excel error sentinel cached as a string
    — exactly the shape openpyxl returns when re-reading a #NUM! cell
    in data_only mode.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Inputs"
    # Real project at col H — clean numeric values
    ws.cell(row=31, column=8, value=0.0723)   # Appraisal IRR
    ws.cell(row=32, column=8, value=2.0901)   # Dev Fee $/W
    ws.cell(row=33, column=8, value=4.1102)   # FMV $/W
    ws.cell(row=37, column=8, value=0.1100)   # Live IRR
    ws.cell(row=38, column=8, value=0.8205)   # NPP $/W
    ws.cell(row=39, column=8, value=5_700_000) # NPP total $
    # Placeholder project at col O — left as #NUM! (openpyxl serializes
    # error tokens as strings on re-read).
    for row in (31, 32, 33, 37, 38, 39):
        ws.cell(row=row, column=15, value="#NUM!")
    out = tmp_path / "merged.xlsx"
    wb.save(out)
    return out


def _make_task(*, name: str, col_letter: str, col_idx: int) -> SolveTask:
    return SolveTask(
        workbook_path="/dev/null",
        project_offset=col_idx - 7,  # PI_BASE_COL = 7
        project_col_letter=col_letter,
        project_name=name,
        project_index_cell=CellAddress(sheet="Project Inputs", address="$F$2"),
        read_cells=(),
    )


def test_verifier_passes_skipped_project_without_diffing_stale_cells(
    tmp_path: Path,
) -> None:
    """The Tranche 7.4 contract: when meta['mode'] starts with 'skipped:'
    OR status == 'skipped', the verifier must not compare hard-stamped
    cells for that project. Otherwise the #NUM! left behind on a
    placeholder column produces a false-positive mismatch every run.
    """
    merged = _make_merged_workbook(tmp_path)
    real_task = _make_task(name="Greenlee I", col_letter="H", col_idx=8)
    placeholder_task = _make_task(name="Project 8", col_letter="O", col_idx=15)

    # Worker results mirror what direct_runner.run_direct produces:
    # the real project carries clean hard-stamp values that match the
    # merged file; the placeholder carries an error-sentinel float
    # (-2146826252 is the actual value pywin32 returns for a #NUM!
    # cell on Caroline's Office 365 build).
    worker_results = {
        0: {
            "project_results": [
                {
                    "project_name": "Greenlee I",
                    "project_offset": 1,
                    "status": "converged",
                    "solved_values": {
                        "Project Inputs!F31": 0.0723,
                        "Project Inputs!H32": 2.0901,
                        "Project Inputs!H33": 4.1102,
                        "Project Inputs!F37": 0.1100,
                        "Project Inputs!H38": 0.8205,
                        "Project Inputs!H39": 5_700_000,
                    },
                    "meta": {"project_offset": 1, "mode": "ok"},
                },
                {
                    "project_name": "Project 8",
                    "project_offset": 8,
                    "status": "skipped",  # Tranche 7.5 surface
                    "solved_values": {
                        "Project Inputs!F31": -2146826252.0,
                        "Project Inputs!O32": -2146826252.0,
                        "Project Inputs!O33": -2146826252.0,
                        "Project Inputs!F37": -2146826252.0,
                        "Project Inputs!O38": -2146826252.0,
                        "Project Inputs!O39": -2146826252.0,
                    },
                    "meta": {
                        "project_offset": 8,
                        "mode": "skipped:no_rc1_revenue",
                    },
                },
            ],
        },
    }

    mismatches = verify_merged_file(
        final_path=merged,
        worker_results=worker_results,
        partitions=[[real_task, placeholder_task]],
    )

    assert mismatches == [], (
        f"Expected zero mismatches (placeholder is skipped). Got:\n  "
        + "\n  ".join(mismatches)
    )


def test_verifier_honors_mode_sentinel_even_if_status_missed_translation(
    tmp_path: Path,
) -> None:
    """Defense-in-depth: a future worker version that forgets to set
    status='skipped' on Tranche 7.2 skips will still carry the
    meta['mode']='skipped:*' sentinel that the VBA macro writes
    directly. The verifier must honor either signal.
    """
    merged = _make_merged_workbook(tmp_path)
    placeholder_task = _make_task(name="Project 8", col_letter="O", col_idx=15)
    worker_results = {
        0: {
            "project_results": [
                {
                    "project_name": "Project 8",
                    "project_offset": 8,
                    # Stale status pre-Tranche 7.5 (worker bug regression case)
                    "status": "not_converged",
                    "solved_values": {
                        "Project Inputs!F31": -2146826252.0,
                        "Project Inputs!O32": -2146826252.0,
                        "Project Inputs!O33": -2146826252.0,
                        "Project Inputs!F37": -2146826252.0,
                        "Project Inputs!O38": -2146826252.0,
                        "Project Inputs!O39": -2146826252.0,
                    },
                    "meta": {
                        "project_offset": 8,
                        "mode": "skipped:no_mwdc",
                    },
                },
            ],
        },
    }

    mismatches = verify_merged_file(
        final_path=merged,
        worker_results=worker_results,
        partitions=[[placeholder_task]],
    )

    assert mismatches == []


def test_verifier_still_catches_real_corruption_on_converged_project(
    tmp_path: Path,
) -> None:
    """The skip bypass must NOT swallow genuine corruption on a real
    project. Worker reports NPP $0.821; merged file has $0.500
    (simulated partial-merge corruption). Verifier must flag.
    """
    merged = _make_merged_workbook(tmp_path)
    # Overwrite H38 to simulate a corrupted merge of NPP $/W
    wb = openpyxl.load_workbook(merged)
    wb["Project Inputs"].cell(row=38, column=8, value=0.500)
    wb.save(merged)

    real_task = _make_task(name="Greenlee I", col_letter="H", col_idx=8)
    worker_results = {
        0: {
            "project_results": [
                {
                    "project_name": "Greenlee I",
                    "project_offset": 1,
                    "status": "converged",
                    "solved_values": {
                        "Project Inputs!F31": 0.0723,
                        "Project Inputs!H32": 2.0901,
                        "Project Inputs!H33": 4.1102,
                        "Project Inputs!F37": 0.1100,
                        "Project Inputs!H38": 0.8205,
                        "Project Inputs!H39": 5_700_000,
                    },
                    "meta": {"project_offset": 1},
                },
            ],
        },
    }

    mismatches = verify_merged_file(
        final_path=merged,
        worker_results=worker_results,
        partitions=[[real_task]],
    )

    assert any("H38" in m for m in mismatches), (
        f"Expected H38 mismatch flagged (NPP corruption). Got: {mismatches}"
    )
