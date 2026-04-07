"""Toggle on 5 projects, run the solver, then restore the original toggles.

This tests the multi-project flow end-to-end without permanently
modifying the workbook (toggles are restored after the test).
"""
from __future__ import annotations

import shutil
import sys
import time
from pathlib import Path

from dn38_solver.config import DEFAULT_WORKBOOK, BASE_COL


def main() -> None:
    wb_path = DEFAULT_WORKBOOK
    if not wb_path.exists():
        print(f"ERROR: Workbook not found: {wb_path}")
        sys.exit(1)

    # Make a backup before modifying toggles
    backup_path = wb_path.parent / (wb_path.stem + "_BACKUP_TEST" + wb_path.suffix)
    print(f"Backing up to: {backup_path.name}")
    shutil.copy2(str(wb_path), str(backup_path))

    # Toggle on first 5 projects using openpyxl (no COM)
    import openpyxl
    print("Opening workbook to set toggles...")
    wbo = openpyxl.load_workbook(str(wb_path), keep_vba=True)
    ws = wbo["Project Inputs"]

    # Save original toggles and set first 5 ON
    originals: dict[int, int | None] = {}
    test_cols = [8, 9, 10, 11, 12]  # H through L
    for c in test_cols:
        name = ws.cell(row=4, column=c).value
        originals[c] = ws.cell(row=7, column=c).value
        ws.cell(row=7, column=c).value = 1
        print(f"  Toggled ON: col {chr(64+c)} — {name} (was {originals[c]})")

    wbo.save(str(wb_path))
    wbo.close()
    print(f"Saved with 5 projects toggled ON\n")

    # Run the solver
    from dn38_solver.solver.orchestrator import solve_all
    import logging
    logging.basicConfig(level=logging.INFO, format="%(message)s", stream=sys.stdout)

    print("=" * 60)
    start = time.time()
    record = solve_all(wb_path, batch_id="multitest")
    elapsed = time.time() - start
    print("=" * 60)
    print(f"\nMulti-project test: {len(record.projects)} projects in {elapsed:.1f}s")
    print(f"Status: {record.status}")

    for p in record.projects:
        npp = f"${p.npp_per_w:.4f}" if p.npp_per_w else "—"
        dev = f"${p.dev_fee_per_w:.4f}" if p.dev_fee_per_w else "—"
        fmv = f"${p.fmv_per_w:.4f}" if p.fmv_per_w else "—"
        dscr = f"{p.dscr_multiple:.4f}x" if p.dscr_multiple else "—"
        eq = f"{p.equity_pct:.2%}" if p.equity_pct else "—"
        ok = "OK" if p.converged else "CHECK"
        print(f"  {p.name:<25} NPP={npp:>10} Dev={dev:>10} FMV={fmv:>10} DSCR={dscr:>8} Eq={eq:>8} [{ok}]")

    # Restore original toggles
    print(f"\nRestoring original toggles...")
    wbo = openpyxl.load_workbook(str(wb_path), keep_vba=True)
    ws = wbo["Project Inputs"]
    for c, val in originals.items():
        ws.cell(row=7, column=c).value = val
    wbo.save(str(wb_path))
    wbo.close()
    print("Toggles restored.")

    # Clean up backup
    backup_path.unlink(missing_ok=True)
    print("Backup removed. Test complete.")


if __name__ == "__main__":
    main()
