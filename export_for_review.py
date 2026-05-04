"""
38DN Excel Macro Runner — Export for VP Review App
Exports macro run results from SQLite as structured .xlsx files that
the VP Review App can load directly via file upload.

The exported file contains a "Project Inputs" sheet formatted to match
the pricing model layout (same row numbers, same column positions) so
the VP app's load_pricing_model() can parse it natively.

Usage:
    python export_for_review.py                           # Export latest run
    python export_for_review.py --db results.db           # Custom DB path
    python export_for_review.py --batch <batch_id>        # Export a batch
    python export_for_review.py --workbook "38DN*.xlsm"   # Filter by name
    python export_for_review.py --output solved.xlsx      # Custom output path
"""
import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment

from dn38_solver.config import DB_PATH, OUTPUT_ROWS, PROJECT_NAME_ROW, PROJECT_TOGGLE_ROW
from dn38_solver.storage.database import get_connection, get_batch_runs, get_latest_run, get_runs


# Row-label mapping (macro runner output labels -> Project Inputs row numbers)
_LABEL_TO_ROW = {
    "Dev Fee ($/W)": 32,
    "FMV Calculated ($/W)": 33,
    "NPP ($/W)": 38,
    "NPP ($)": 39,
    "FMV WACC (Target)": 30,
    "Live Appraisal IRR": 31,
    "Target IRR": 36,
    "Live Levered Pre-Tax IRR": 37,
}

# Row labels for the exported sheet (matches VP app's INPUT_ROW_LABELS + OUTPUT_ROWS)
_ROW_LABELS = {
    4: "Project Name",
    7: "Toggle (On/Off)",
    11: "Size MWDC",
    12: "Size MWAC",
    30: "FMV WACC (Target)",
    31: "Live Appraisal IRR",
    32: "Dev Fee ($/W)",
    33: "FMV Calculated ($/W)",
    36: "Target IRR",
    37: "Live Levered Pre-Tax IRR",
    38: "NPP ($/W)",
    39: "NPP ($)",
}


def _runs_to_project_list(runs):
    """Convert a list of DB rows to project dicts with row-keyed data."""
    projects = []
    for r in runs:
        raw = r["raw_outputs"]
        raw_dict = json.loads(raw) if raw else {}
        project_outputs = raw_dict.get("project_outputs", {})

        data = {}
        for label, value in project_outputs.items():
            row_num = _LABEL_TO_ROW.get(label)
            if row_num is not None:
                data[row_num] = value

        # Fill from top-level DB columns as fallback
        if r["npp_per_w"] is not None:
            data.setdefault(38, r["npp_per_w"])
        if r["npp_total"] is not None:
            data.setdefault(39, r["npp_total"])
        if r["fmv_per_w"] is not None:
            data.setdefault(33, r["fmv_per_w"])
        if r["dev_fee_per_w"] is not None:
            data.setdefault(32, r["dev_fee_per_w"])
        if r["target_irr"] is not None:
            data.setdefault(36, r["target_irr"])
        if r["live_irr"] is not None:
            data.setdefault(37, r["live_irr"])

        projects.append({
            "name": r["project_name"] or "Unknown",
            "col": r["project_col"] or 6,
            "data": data,
        })

    return projects


def _write_project_inputs_sheet(ws, projects):
    """Write project data into the sheet matching the pricing model layout.

    Row numbers and column positions match the real pricing model so that
    load_pricing_model() in the VP app can parse this file.
    """
    header_font = Font(bold=True, size=10)
    label_font = Font(bold=True, size=9, color="444444")

    # Column A header
    ws.cell(row=1, column=1, value="Project Inputs (Macro Runner Export)")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)

    # Write row labels in columns A-E (column 1-5)
    for row_num, label in sorted(_ROW_LABELS.items()):
        cell = ws.cell(row=row_num, column=1, value=label)
        cell.font = label_font
        cell.alignment = Alignment(horizontal="right")

    # Write each project in its original column position
    for proj in projects:
        col = proj["col"]

        # Row 4: Project name
        ws.cell(row=PROJECT_NAME_ROW, column=col, value=proj["name"])
        ws.cell(row=PROJECT_NAME_ROW, column=col).font = header_font

        # Row 7: Toggle (always "1" since macro runner only saves active projects)
        ws.cell(row=PROJECT_TOGGLE_ROW, column=col, value=1)

        # Data rows
        for row_num, value in proj["data"].items():
            if value is not None:
                ws.cell(row=row_num, column=col, value=value)

    # Auto-width column A
    ws.column_dimensions["A"].width = 28


def export_solved_results(db_path, output_path, workbook_filter=None):
    """Export the latest macro run results as a structured .xlsx file.

    Args:
        db_path: Path to the SQLite database.
        output_path: Where to write the .xlsx file.
        workbook_filter: If provided, only export runs from this workbook name.

    Returns:
        Path to the written file, or None if no data found.
    """
    conn = get_connection(Path(db_path))

    if workbook_filter:
        # Get the latest batch for this workbook
        batch_row = conn.execute(
            "SELECT batch_id FROM macro_runs "
            "WHERE workbook_name = ? AND status = 'success' AND batch_id IS NOT NULL "
            "ORDER BY run_timestamp DESC LIMIT 1",
            (workbook_filter,)
        ).fetchone()

        if batch_row and batch_row["batch_id"]:
            runs = conn.execute(
                "SELECT * FROM macro_runs WHERE batch_id = ? AND status = 'success' "
                "ORDER BY project_col ASC",
                (batch_row["batch_id"],)
            ).fetchall()
        else:
            # Fall back: latest single run for this workbook
            runs = conn.execute(
                "SELECT * FROM macro_runs "
                "WHERE workbook_name = ? AND status = 'success' "
                "ORDER BY run_timestamp DESC LIMIT 1",
                (workbook_filter,)
            ).fetchall()
    else:
        # Get the most recent batch or single run
        latest = conn.execute(
            "SELECT batch_id FROM macro_runs "
            "WHERE status = 'success' "
            "ORDER BY run_timestamp DESC LIMIT 1"
        ).fetchone()

        if latest and latest["batch_id"]:
            runs = conn.execute(
                "SELECT * FROM macro_runs WHERE batch_id = ? AND status = 'success' "
                "ORDER BY project_col ASC",
                (latest["batch_id"],)
            ).fetchall()
        else:
            runs = conn.execute(
                "SELECT * FROM macro_runs "
                "WHERE status = 'success' "
                "ORDER BY run_timestamp DESC LIMIT 1"
            ).fetchall()

    conn.close()

    if not runs:
        print("No successful runs found.")
        return None

    projects = _runs_to_project_list(runs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Inputs"
    _write_project_inputs_sheet(ws, projects)

    output_path = Path(output_path)
    wb.save(str(output_path))
    wb.close()
    print(f"Exported {len(projects)} project(s) to {output_path}")
    return output_path


def export_batch_results(db_path, batch_id, output_path):
    """Export all results from a specific batch as a structured .xlsx file.

    Args:
        db_path: Path to the SQLite database.
        batch_id: The batch identifier.
        output_path: Where to write the .xlsx file.

    Returns:
        Path to the written file, or None if no data found.
    """
    conn = get_connection(Path(db_path))
    runs = get_batch_runs(conn, batch_id)
    conn.close()

    # Filter to successful runs only
    runs = [r for r in runs if r["status"] == "success"]

    if not runs:
        print(f"No successful runs found for batch '{batch_id}'.")
        return None

    projects = _runs_to_project_list(runs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Inputs"
    _write_project_inputs_sheet(ws, projects)

    output_path = Path(output_path)
    wb.save(str(output_path))
    wb.close()
    print(f"Exported {len(projects)} project(s) from batch '{batch_id}' to {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Export macro run results for VP Review App")
    parser.add_argument("--db", default=str(DB_PATH), help="Path to SQLite database")
    parser.add_argument("--output", "-o", default=None, help="Output .xlsx path")
    parser.add_argument("--batch", default=None, help="Export a specific batch by ID")
    parser.add_argument("--workbook", default=None, help="Filter by workbook name")
    args = parser.parse_args()

    if args.batch:
        out = args.output or f"macro_results_batch_{args.batch[:8]}.xlsx"
        export_batch_results(args.db, args.batch, out)
    else:
        out = args.output or "macro_results_latest.xlsx"
        export_solved_results(args.db, out, workbook_filter=args.workbook)


if __name__ == "__main__":
    main()
