"""Export per-project checkpoint rows for a given batch_id to .xlsx.

Use after a chunked solve crashes mid-portfolio (e.g. Excel dies at
SaveAs after the macro converges) so the in-memory converged values are
lost but the per-project checkpoints landed in SQLite. Pull them out
into a clean workbook for paste-back into the source pricing model.

Usage:
    python export_checkpoints.py <batch_id> [output.xlsx]

The output sheet has one row per checkpointed project with columns
matching what the macro writes to Project Inputs:
    NPP $/W (row 38), Dev Fee $/W (row 32), FMV $/W (row 33),
    Live Levered IRR (row 37), Appraisal IRR (row 31),
    DSCR (PT Returns!F129), Equity %, convergence tier.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dn38_solver.storage.database import get_checkpointed_projects, get_connection
from dn38_solver.types import convergence_label


def export(batch_id: str, output_path: Path) -> None:
    conn = get_connection()
    try:
        projects = get_checkpointed_projects(conn, batch_id)
    finally:
        conn.close()

    if not projects:
        print(f"No checkpoints found for batch_id={batch_id}.")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checkpoints"

    headers = [
        "Project",
        "Col (PI)",
        "PI col letter",
        "NPP $/W (row 38)",
        "Dev Fee $/W (row 32)",
        "FMV $/W (row 33)",
        "Live Levered IRR (row 37)",
        "Appraisal IRR (row 31)",
        "DSCR",
        "Equity %",
        "Convergence",
        "Iterations",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", name="Aptos Narrow", size=10)
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    body_font = Font(name="Aptos Narrow", size=10, color="050D25")
    body_align = Alignment(horizontal="center", vertical="center")

    for p in projects:
        ws.append([
            p.name,
            p.col,
            get_column_letter(p.col),
            p.npp_per_w,
            p.dev_fee_per_w,
            p.fmv_per_w,
            p.live_irr,
            p.appraisal_live,
            p.dscr_multiple,
            p.equity_pct,
            convergence_label(p),
            p.iterations,
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = body_align

    # Number formats — em-dash zero per 38DN convention
    fmt_dollar_3 = '_([$$]#,##0.000_);[Red]_([$$](#,##0.000);_("–")_;_(@_)'
    fmt_pct_2 = '_(#,##0.00%_);(#,##0.00%);_("–")_;_(@_)'
    fmt_dscr = '0.000\\x'

    for col_letter, fmt in [
        ("D", fmt_dollar_3),  # NPP
        ("E", fmt_dollar_3),  # Dev Fee
        ("F", fmt_dollar_3),  # FMV
        ("G", fmt_pct_2),     # Live IRR
        ("H", fmt_pct_2),     # Appraisal IRR
        ("I", fmt_dscr),      # DSCR
        ("J", fmt_pct_2),     # Equity %
    ]:
        for cell in ws[col_letter][1:]:
            cell.number_format = fmt

    widths = [28, 8, 12, 14, 14, 14, 18, 16, 10, 10, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 38

    ws.freeze_panes = "A2"

    wb.save(str(output_path))
    print(f"Wrote {len(projects)} project(s) -> {output_path}")


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python export_checkpoints.py <batch_id> [output.xlsx]")
        sys.exit(2)
    batch_id = sys.argv[1]
    if len(sys.argv) >= 3:
        out = Path(sys.argv[2])
    else:
        out = Path.cwd() / f"checkpoints_{batch_id}.xlsx"
    export(batch_id, out)


if __name__ == "__main__":
    main()
