"""dn38_solver.com.parallel_runner — Spawn N worker processes for parallel solves.

Each worker runs in its own subprocess with its own Excel COM session, its
own temp workbook copy, and a round-robin slice of the project list.
Parent collects per-worker results and merges them into a single
`<workbook>_SOLVED.xlsm` next to the source.

Why a separate parallel runner instead of an N=N>1 path inside run_direct:
- run_direct holds an Excel COM apartment; spawning subprocesses from
  inside a COM-using process is messy and risks hidden EXCEL.EXE leaks
  when the parent crashes.
- Parent stays COM-free — only worker processes touch Excel — so PID
  tracking, timeout enforcement, and cleanup all work via stdlib subprocess.
- Each worker reuses run_direct unchanged. No code duplication.

Caroline's correctness contract (Issue #8 + her clarification):
- Per-project NPP / Dev Fee / DSCR / FMV / Live IRR / Appraisal IRR must
  match single-worker output to ~1e-4.
- Portfolio-level aggregates (Portfolio sheet, cross-column sums) may be
  stale in the merged _SOLVED.xlsm; Excel will recalc them on next
  interactive open since the underlying project columns are correct.
- Critical sheets per spec (Dashboard, Table, PT Returns, NPP Calc,
  Appraisal, Perm Debt, Tax Equity, CL, Project Inputs) recalc inside
  each worker's own Excel session for its assigned projects.
"""
from __future__ import annotations

import contextlib
import json
import logging
import os
import shutil
import subprocess
import sys
import tempfile
import threading
import time
from pathlib import Path
from typing import Sequence

import msgspec

from dn38_solver.types import SolveTask

log = logging.getLogger(__name__)

# Hard cap so a user typo (`--workers 64`) doesn't open 64 Excel processes
# and consume all RAM. The Plan agent's analysis suggested cpu_count // 2
# as a sane upper bound; 8 is a portable conservative cap.
MAX_WORKERS = 8


class _WorkerLogTailer(threading.Thread):
    """Background thread that tails each worker's stderr.log and forwards
    new lines to the parent log so the terminal isn't silent during long
    cold solves.

    Tracks per-file byte offsets so each line is forwarded exactly once.
    Lines already carry the worker id via the `[wN]` prefix the worker
    sets in `logging.basicConfig(format=...)`, so we don't re-decorate.
    """

    def __init__(
        self,
        *,
        sources: list[tuple[int, Path]],
        poll_interval: float = 2.0,
    ) -> None:
        super().__init__(daemon=True, name="WorkerLogTailer")
        self._sources = list(sources)
        self._interval = poll_interval
        self._offsets: dict[Path, int] = {p: 0 for _, p in sources}
        self._stop_evt = threading.Event()

    def run(self) -> None:
        while not self._stop_evt.is_set():
            self._poll_once()
            if self._stop_evt.wait(timeout=self._interval):
                break
        # Final flush so we don't miss the last few lines after stop
        self._poll_once()

    def _poll_once(self) -> None:
        for _wid, path in self._sources:
            try:
                size = path.stat().st_size
            except FileNotFoundError:
                continue
            offset = self._offsets[path]
            if size <= offset:
                continue
            try:
                with open(path, "rb") as fh:
                    fh.seek(offset)
                    chunk = fh.read(size - offset)
            except OSError:
                continue
            self._offsets[path] = size
            text = chunk.decode("utf-8", errors="replace")
            for line in text.splitlines():
                if line.strip():
                    log.info("  %s", line)

    def stop(self, *, join_timeout: float = 5.0) -> None:
        self._stop_evt.set()
        self.join(timeout=join_timeout)


def _partition_round_robin(
    tasks: Sequence[SolveTask], n_workers: int
) -> list[list[SolveTask]]:
    """Distribute tasks round-robin so warm/cold projects spread evenly.

    Block partitioning (chunks of consecutive indices) tends to put slow
    cold-mode projects in the same chunk, leaving one worker as the
    long-tail bottleneck. Round-robin avoids that.
    """
    slices: list[list[SolveTask]] = [[] for _ in range(n_workers)]
    for i, t in enumerate(tasks):
        slices[i % n_workers].append(t)
    return slices


def _kill_worker_excel_children(worker_pid: int) -> None:
    """Terminate any EXCEL.EXE under a worker PID. No-op if psutil missing."""
    from dn38_solver.com.cleanup import kill_excel_children
    with contextlib.suppress(Exception):
        kill_excel_children(worker_pid)


def run_parallel(
    workbook_path: str,
    tasks: list[SolveTask],
    *,
    workers: int,
    original_f2: int = 1,
    timeout_sec: int = 3600,
    use_chunked: bool = True,
    skip_output_recalc: bool = False,
    strip_sheets: tuple[str, ...] = (),
    excel_threads_per_worker: int | None = None,
) -> dict:
    """Spawn `workers` subprocess workers and merge their results.

    Returns the same shape as `run_direct` so the orchestrator can branch
    on workers > 1 without other changes downstream:
        {status, project_results, duration_sec, saved_to, error,
         macro_used, open_time_sec, warmup_time_sec, solve_time_sec,
         read_time_sec, solver_heartbeat, validation,
         merge_path, estimated_sequential_sec}

    `merge_path` is one of "openpyxl" | "vba_fallback" | "copy_master" |
    None and tells the orchestrator how authoritative the merged file is.
    `estimated_sequential_sec` is the sum of per-project solve seconds
    (from each worker's __SolverResults timings), so the caller can
    print real speedup instead of guessing.
    """
    n = max(1, min(workers, MAX_WORKERS))
    if n != workers:
        log.warning("Clamping workers from %d to %d (MAX_WORKERS)", workers, n)

    start = time.time()
    parent_tmp = Path(tempfile.mkdtemp(prefix="38dn_parallel_"))

    # Tracked across the try/finally so cleanup can always tear them down,
    # even if Ctrl+C lands mid-spawn or mid-wait. Without this, an interrupt
    # during a 20-minute solve leaves N hidden Excel processes running until
    # the user reboots.
    procs: list[tuple[int, subprocess.Popen, Path, Path, Path, Path]] = []
    aggregator = None
    progress = None
    saved_to: str | None = None
    merge_path: str | None = None
    worker_errors: list[str] = []
    worker_results: dict[int, dict] = {}
    project_results: list[dict] = []

    # Cap Excel threads per worker so N × Excel doesn't oversubscribe CPU.
    # Validated empirically on SMP WalkTEST: cpu_count // n_workers (24/2=12)
    # produced ZERO speedup vs single-worker because the 24 calc threads
    # competed with the OS, pywin32, the aggregator, and the live-progress
    # tailer for the same 24 cores. Excel's multi-threaded recalc has
    # sharply diminishing returns past ~4 threads anyway, so the default
    # is a conservative cap that leaves headroom for coordination overhead.
    # User can override via the CLI flag if their workload differs.
    #
    # Semantics of `excel_threads_per_worker`:
    #   None    -> use the default 1/3-of-cores cap (per-worker)
    #   0       -> no cap; let Excel use its built-in default (cpu_count)
    #   1..32   -> set Excel's ThreadCount to exactly that value
    #   >32     -> clamped to 32 (sanity cap; Excel's own implementation
    #              caps at 1024 but anything past 32 is wasted on the kind
    #              of dependency-chain heavy recalc this model does)
    cpu = os.cpu_count() or 4
    SANITY_MAX_THREADS = 32
    if excel_threads_per_worker is None:
        # Default: 1/3 of available cores per worker, min 2, max 4.
        # On a 24-core box with 2 workers: 4 + 4 = 8 calc threads, leaving
        # 16 cores for OS / coordination / file I/O.
        threads_per_worker: int | None = max(2, min(4, cpu // (n * 3)))
    elif excel_threads_per_worker == 0:
        threads_per_worker = None  # signals "no cap" to direct_runner
    else:
        if excel_threads_per_worker > SANITY_MAX_THREADS:
            log.warning(
                "  Clamping excel_threads_per_worker from %d to %d (sanity cap)",
                excel_threads_per_worker, SANITY_MAX_THREADS,
            )
            threads_per_worker = SANITY_MAX_THREADS
        else:
            threads_per_worker = int(excel_threads_per_worker)
    threads_label = (
        "Excel default (no cap)" if threads_per_worker is None
        else f"{threads_per_worker} Excel threads each"
    )
    log.info(
        "  Parallel mode: %d workers × %s (cpu_count=%d)",
        n, threads_label, cpu,
    )

    partitions = _partition_round_robin(tasks, n)
    log.info(
        "  Task split (round-robin): %s",
        ", ".join(f"w{i}={len(p)}" for i, p in enumerate(partitions)),
    )

    try:
        # Spawn workers. Each worker's stdout/stderr is redirected to a file
        # in its temp dir, NOT to a subprocess.PIPE — Windows anonymous pipes
        # have a ~65KB buffer, and a long cold solve can easily emit more than
        # that to stderr, deadlocking the worker (writes to a full pipe block
        # until the parent reads, but proc.communicate() is what reads). Log
        # files round-trip cleanly and let the parent tail them for live
        # progress without any blocking risk.
        worker_status_paths: list[Path] = []
        for worker_id, slice_tasks in enumerate(partitions):
            if not slice_tasks:
                continue
            wdir = parent_tmp / f"w{worker_id}"
            wdir.mkdir(parents=True, exist_ok=True)
            config_path = wdir / "config.json"
            result_path = wdir / "result.json"
            output_path = wdir / f"{Path(workbook_path).stem}_SOLVED_w{worker_id}.xlsm"
            status_path = wdir / f"solver_status_w{worker_id}.json"
            worker_status_paths.append(status_path)

            tasks_json = msgspec.json.encode(slice_tasks).decode("utf-8")
            config = {
                "workbook_path": workbook_path,
                "tasks_json": tasks_json,
                "output_path": str(output_path),
                "status_path": str(status_path),
                "worker_id": worker_id,
                "original_f2": int(original_f2),
                "timeout_sec": int(timeout_sec),
                "use_chunked": bool(use_chunked),
                "skip_output_recalc": bool(skip_output_recalc),
                "strip_sheets": list(strip_sheets),
                "excel_threads": threads_per_worker,
            }
            config_path.write_text(json.dumps(config), encoding="utf-8")

            cmd = [
                sys.executable, "-u", "-m", "dn38_solver.com.worker",
                str(config_path), str(result_path),
            ]
            # CREATE_NO_WINDOW = 0x08000000 — suppress console flash for each
            # worker subprocess. Windows-only flag; ignored on other OS.
            creationflags = 0
            if sys.platform == "win32":
                creationflags = 0x08000000
            log.info("  Spawning worker %d (%d projects)...", worker_id, len(slice_tasks))
            # Open log files unbuffered (binary 'wb' with no flush deferral)
            # so the parent's tailing thread sees worker writes immediately.
            stdout_path = wdir / "stdout.log"
            stderr_path = wdir / "stderr.log"
            stdout_fh = open(stdout_path, "wb", buffering=0)
            stderr_fh = open(stderr_path, "wb", buffering=0)
            proc = subprocess.Popen(
                cmd,
                stdout=stdout_fh,
                stderr=stderr_fh,
                creationflags=creationflags,
            )
            # We keep handles open in the child; the parent doesn't write to
            # them. Close our parent-side handles so we don't pin the inode.
            stdout_fh.close()
            stderr_fh.close()
            procs.append((worker_id, proc, config_path, result_path, stdout_path, stderr_path))

        # Start the status aggregator. Each worker writes its own status JSON;
        # the aggregator merges them into the canonical `solver_status.json`
        # the Streamlit tracker reads, so dashboards work without changes.
        from dn38_solver.com.direct_runner import STATUS_FILE
        from dn38_solver.com.status_aggregator import StatusAggregator
        aggregator = StatusAggregator(
            worker_status_paths=worker_status_paths,
            output_path=STATUS_FILE,
            workbook_path=workbook_path,
            poll_interval=1.0,
        )
        aggregator.start()

        # Live-progress thread: tails each worker's stderr.log and forwards
        # new lines to the parent log so the terminal isn't silent during
        # a 10-20 minute cold solve. Without this the user sees nothing
        # between "Spawning worker N" and "Worker N completed".
        progress = _WorkerLogTailer(
            sources=[(wid, err_path) for (wid, _, _, _, _, err_path) in procs],
            poll_interval=2.0,
        )
        progress.start()

        # Wait for all workers. Per-worker timeout is the parent-level timeout
        # since each worker handles only its slice; the cap should be generous
        # enough that no worker hits it before its slice's COM session does.
        parent_timeout = max(timeout_sec, 60)
        deadline = time.time() + parent_timeout + 60  # +60s for cleanup grace

        for worker_id, proc, config_path, result_path, stdout_path, stderr_path in procs:
            remaining = max(1, int(deadline - time.time()))
            try:
                # stdout/stderr are redirected to files, so communicate()
                # blocks only on the process exit — no PIPE buffer to drain.
                proc.communicate(timeout=remaining)
            except subprocess.TimeoutExpired:
                log.warning(
                    "  Worker %d timeout after %ds — terminating and reaping "
                    "child Excel processes by PID", worker_id, remaining,
                )
                _kill_worker_excel_children(proc.pid)
                proc.terminate()
                with contextlib.suppress(subprocess.TimeoutExpired):
                    proc.communicate(timeout=10)
                if proc.poll() is None:
                    proc.kill()
                worker_results[worker_id] = {
                    "status": "error",
                    "error": f"Worker {worker_id} timed out",
                    "project_results": [],
                }
                continue

            # Worker stderr is now in stderr_path and the live tailer has
            # already forwarded it during the run. Drain any final lines
            # for completeness.
            with contextlib.suppress(OSError):
                tail = stderr_path.read_text(encoding="utf-8", errors="replace")
                # Only forward lines the tailer hasn't already printed;
                # tailer tracks per-file byte offsets so we just print the
                # recent tail as INFO if the worker errored, otherwise the
                # lines were already streamed live.
                if proc.returncode != 0 and tail.strip():
                    log.info("  --- worker %d stderr tail ---", worker_id)
                    for line in tail.splitlines()[-30:]:
                        log.info("  %s", line)

            if proc.returncode != 0:
                log.error("  Worker %d exited with code %d", worker_id, proc.returncode)
                _kill_worker_excel_children(proc.pid)
                try:
                    worker_results[worker_id] = json.loads(
                        result_path.read_text(encoding="utf-8")
                    )
                except Exception:
                    worker_results[worker_id] = {
                        "status": "error",
                        "error": f"Worker {worker_id} exited non-zero, no result file",
                        "project_results": [],
                    }
                continue

            try:
                worker_results[worker_id] = json.loads(
                    result_path.read_text(encoding="utf-8")
                )
            except Exception as exc:
                log.error("  Could not read worker %d result: %s", worker_id, exc)
                worker_results[worker_id] = {
                    "status": "error",
                    "error": f"Could not read worker {worker_id} result: {exc}",
                    "project_results": [],
                }

        # Aggregate per-project results in original task order.
        by_offset: dict[int, dict] = {}
        for wid, wresult in worker_results.items():
            if wresult.get("error"):
                worker_errors.append(f"w{wid}: {wresult['error']}")
            for pr in wresult.get("project_results", []):
                # Use the top-level project_offset stamped by direct_runner.
                # Falling back to name-based lookup is unsafe — portfolios
                # can have duplicate project names across LLCs.
                offset = pr.get("project_offset")
                if offset is None:
                    meta = pr.get("meta") or {}
                    offset = meta.get("project_offset") or meta.get("offset")
                if offset is None:
                    log.warning(
                        "Worker %d returned a project_result with no offset "
                        "key; skipping rather than name-matching (unsafe with "
                        "duplicate names). project_name=%r",
                        wid, pr.get("project_name"),
                    )
                    continue
                by_offset[int(offset)] = pr

        for t in tasks:
            pr = by_offset.get(t.project_offset)
            if pr is None:
                project_results.append({
                    "project_name": t.project_name,
                    "status": "not_attempted",
                    "solved_values": {},
                    "iterations_used": 0,
                    "duration_sec": 0,
                    "meta": {},
                })
            else:
                project_results.append(pr)

        # Merge the worker _SOLVED.xlsm files into one canonical output.
        # Strategy: pick a successfully-converged worker as master, copy
        # converged cells (the per-project columns) from each other worker's
        # file via openpyxl. Cross-project portfolio aggregates may be stale
        # per Caroline's spec — Excel recalcs them on next interactive open.
        # Prefer the lowest-id worker whose status is "converged" AND who
        # actually saved a file. Falling back to procs[0] regardless of
        # outcome risks merging on top of an errored worker's empty/corrupt
        # output.
        candidate_masters = sorted(
            (wid for wid, r in worker_results.items()
             if r.get("status") == "converged" and r.get("saved_to")
             and Path(r["saved_to"]).exists()),
        )
        if candidate_masters:
            master_worker_id = candidate_masters[0]
            master_result = worker_results.get(master_worker_id, {})
            master_src = master_result.get("saved_to")
            if master_src and Path(master_src).exists():
                wb_path = Path(workbook_path)
                final_path = wb_path.parent / f"{wb_path.stem}_SOLVED{wb_path.suffix}"
                try:
                    _merge_solved_workbooks(
                        master_src=Path(master_src),
                        others=[
                            Path(worker_results[wid].get("saved_to", ""))
                            for wid in worker_results
                            if wid != master_worker_id
                            and worker_results[wid].get("saved_to")
                        ],
                        final_path=final_path,
                        partitions=partitions,
                        master_worker_id=master_worker_id,
                    )
                    saved_to = str(final_path)
                    merge_path = "openpyxl"
                    log.info("  Merged %d worker output(s) into %s",
                             len(worker_results), final_path)
                except Exception as merge_exc:
                    # openpyxl merge failed — most likely keep_vba=True did
                    # not survive the round-trip on this workbook's macro
                    # project. Fall back to a VBA-side merge that uses Excel
                    # COM (which handles .xlsm natively) to stamp converged
                    # column values from peer workers into the master.
                    log.warning(
                        "  openpyxl merge failed (%s) — trying VBA-helper "
                        "fallback via Excel COM", merge_exc,
                    )
                    others_paths = [
                        Path(worker_results[wid].get("saved_to", ""))
                        for wid in worker_results
                        if wid != master_worker_id
                        and worker_results[wid].get("saved_to")
                    ]
                    try:
                        _merge_via_vba_fallback(
                            master_src=Path(master_src),
                            others=others_paths,
                            final_path=final_path,
                            partitions=partitions,
                            worker_results=worker_results,
                            master_worker_id=master_worker_id,
                        )
                        saved_to = str(final_path)
                        merge_path = "vba_fallback"
                        log.info(
                            "  VBA-helper fallback succeeded: merged %d output(s) into %s",
                            len(worker_results), final_path,
                        )
                    except Exception as vba_exc:
                        log.warning(
                            "  VBA-helper fallback also failed (%s) — copying "
                            "master worker's output as-is. Per-project columns "
                            "owned by other workers will reflect their PRE-solve "
                            "values; consult the per-worker _SOLVED.xlsm files "
                            "in %s for forensic recovery.",
                            vba_exc, parent_tmp,
                        )
                        with contextlib.suppress(Exception):
                            shutil.copy2(master_src, final_path)
                            saved_to = str(final_path)
                            merge_path = "copy_master"
                        # Force run-level error so the user knows the merged
                        # file is not authoritative and so parent_tmp is
                        # preserved for forensics (see cleanup gate below).
                        worker_errors.append(
                            "merge fell back to master-only — per-project "
                            "columns owned by non-master workers reflect "
                            "PRE-solve values, not converged values"
                        )

                # Post-merge sanity gate. Re-open the merged file and
                # confirm the hard-stamped convergence cells match what
                # each worker reported. If they don't, the merge silently
                # corrupted data and we must surface that as an error so
                # the user doesn't ship a file with wrong NPP / FMV /
                # Dev Fee in some columns. Skip the gate when we already
                # know the merge fell back to master-only — its mismatches
                # are expected and would just spam the log.
                if saved_to and merge_path in ("openpyxl", "vba_fallback"):
                    mismatches = _verify_merged_file(
                        final_path=Path(saved_to),
                        worker_results=worker_results,
                        partitions=partitions,
                    )
                    if mismatches:
                        log.error(
                            "  Post-merge verification: %d cell mismatch(es) "
                            "between merged file and worker reports",
                            len(mismatches),
                        )
                        for m in mismatches[:10]:
                            log.error("    %s", m)
                        if len(mismatches) > 10:
                            log.error("    ... (%d more)", len(mismatches) - 10)
                        worker_errors.append(
                            f"merged file failed sanity gate "
                            f"({len(mismatches)} cell mismatch(es)) — "
                            f"per-worker outputs in {parent_tmp} are authoritative"
                        )
                    else:
                        log.info("  Post-merge verification: OK")
        else:
            # No worker produced a saved, converged output — nothing to merge.
            # Surface this clearly so the orchestrator marks the run as error
            # and keeps parent_tmp for forensics.
            log.error(
                "  No worker produced a converged _SOLVED.xlsm — skipping merge. "
                "Worker outcomes: %s",
                {wid: r.get("status") for wid, r in worker_results.items()},
            )
            worker_errors.append(
                "no converged worker output to merge — all workers errored or "
                "produced no saved file"
            )

    finally:
        # Always tear down workers and the helper threads, including on
        # KeyboardInterrupt or any unhandled exception during merge. Order
        # matters: kill workers first so their final stderr writes have
        # a chance to land before the tailer stops.
        for wid, p, _cfg, _res, _out, _err in procs:
            if p.poll() is None:
                log.warning(
                    "  Cleanup: terminating worker %d (pid=%d) and its "
                    "Excel children", wid, p.pid,
                )
                _kill_worker_excel_children(p.pid)
                with contextlib.suppress(Exception):
                    p.terminate()
                with contextlib.suppress(Exception):
                    p.wait(timeout=5)
                if p.poll() is None:
                    with contextlib.suppress(Exception):
                        p.kill()
        if progress is not None:
            with contextlib.suppress(Exception):
                progress.stop(join_timeout=5.0)
        if aggregator is not None:
            with contextlib.suppress(Exception):
                aggregator.stop(join_timeout=5.0)

    # Cleanup worker temp dirs (keep parent tmp_dir only if a worker errored
    # — useful for forensics; otherwise drop everything).
    # DN38_KEEP_WORKER_TMP=1 forces retention regardless (debugging hatch).
    keep_tmp = bool(worker_errors) or os.environ.get("DN38_KEEP_WORKER_TMP") == "1"
    if not keep_tmp:
        with contextlib.suppress(OSError):
            shutil.rmtree(parent_tmp)
    else:
        log.info(
            "  Worker errors detected — preserving %s for forensics",
            parent_tmp,
        )

    total = time.time() - start
    batch_status = "error" if worker_errors else "converged"
    error_msg = "; ".join(worker_errors) if worker_errors else None

    # Compose the run_direct-compatible result. Per-stage timings are
    # summed across workers (open_time, solve_time, read_time); the parent
    # didn't directly measure them so we accumulate worker reports.
    def _sum_stage(key: str) -> float:
        return sum(
            float(w.get(key, 0) or 0) for w in worker_results.values()
        )

    # Estimated sequential time = sum of every project's solve_seconds
    # (from VBA's __SolverResults!M, surfaced as meta["solve_seconds"]).
    # Comparing the parallel wall time to this gives the actual speedup,
    # which is more useful than "elapsed vs. previous run" since project
    # mix differs across runs.
    estimated_sequential = 0.0
    for wresult in worker_results.values():
        for pr in wresult.get("project_results", []):
            meta = pr.get("meta") or {}
            try:
                estimated_sequential += float(meta.get("solve_seconds") or 0.0)
            except (TypeError, ValueError):
                pass

    return {
        "status": batch_status,
        "project_results": project_results,
        "duration_sec": round(total, 2),
        "saved_to": saved_to,
        "error": error_msg,
        "macro_used": "SolveHeadless",
        "open_time_sec": round(_sum_stage("open_time_sec"), 2),
        "warmup_time_sec": round(_sum_stage("warmup_time_sec"), 2),
        "solve_time_sec": round(_sum_stage("solve_time_sec"), 2),
        "read_time_sec": round(_sum_stage("read_time_sec"), 2),
        "solver_heartbeat": None,
        "validation": None,
        "workers_used": n,
        "merge_path": merge_path,
        "estimated_sequential_sec": round(estimated_sequential, 2),
    }


def _verify_merged_file(
    *,
    final_path: Path,
    worker_results: dict[int, dict],
    partitions: list[list[SolveTask]],
    abs_tol: float = 0.01,
) -> list[str]:
    """Re-open the merged xlsm and assert the hard-stamped convergence
    cells match what each worker reported.

    Returns a list of mismatch strings (empty list = clean merge).

    Why this exists: every merge path (openpyxl AND VBA-helper) iterates
    per-project, copying convergence values from peer worker files into
    the master. A silent failure inside that loop — wrong column letter,
    swallowed COM exception, peer file corrupted — would leave the merged
    file with PRE-solve values in some columns and the user would only
    notice when an IC memo derived from the file produced wrong numbers.
    Better to surface the inconsistency at run-end and force the user to
    consult the per-worker outputs.

    Tolerance defaults to $0.01/W since the cells are hard-stamped exact
    floats (cell-self-assign in VBA), but we leave a small slack for any
    openpyxl serialization rounding.
    """
    import openpyxl
    from openpyxl.utils import column_index_from_string

    HARD_STAMPED_ROWS = (31, 32, 33, 37, 38, 39)
    mismatches: list[str] = []

    try:
        wb = openpyxl.load_workbook(
            str(final_path), data_only=True, read_only=True, keep_vba=True,
        )
    except Exception as exc:
        return [f"merged file unreadable: {exc}"]

    try:
        if "Project Inputs" not in wb.sheetnames:
            return ["merged file has no 'Project Inputs' sheet"]
        ws = wb["Project Inputs"]

        # Build name -> result lookup. Worker results carry project_name;
        # we use that to find the worker's solved_values dict for each
        # task. Duplicate names across LLCs are unlikely at the per-worker
        # slice level (round-robin partition spreads them), but we log a
        # warning if we hit one rather than picking arbitrarily.
        name_to_result: dict[str, dict] = {}
        for wresult in worker_results.values():
            for pr in wresult.get("project_results", []):
                pname = pr.get("project_name")
                if not pname:
                    continue
                if pname in name_to_result:
                    log.warning(
                        "  Verify: duplicate project_name %r across workers; "
                        "verification may be ambiguous", pname,
                    )
                name_to_result[pname] = pr

        for tasks_slice in partitions:
            for task in tasks_slice:
                pr = name_to_result.get(task.project_name)
                if pr is None:
                    continue
                # Only verify converged projects. Non-converged projects
                # legitimately have stale or missing values in the worker
                # report, so a mismatch there is meaningless noise.
                if pr.get("status") != "converged":
                    continue
                sv = pr.get("solved_values", {})
                col_idx = column_index_from_string(task.project_col_letter)
                for row in HARD_STAMPED_ROWS:
                    expected = sv.get(
                        f"Project Inputs!{task.project_col_letter}{row}"
                    )
                    if expected is None:
                        continue
                    try:
                        expected_f = float(expected)
                    except (TypeError, ValueError):
                        continue
                    actual = ws.cell(row=row, column=col_idx).value
                    if actual is None:
                        mismatches.append(
                            f"{task.project_name} "
                            f"{task.project_col_letter}{row}: "
                            f"merged value missing (None); "
                            f"expected {expected_f:.4f}"
                        )
                        continue
                    try:
                        actual_f = float(actual)
                    except (TypeError, ValueError):
                        mismatches.append(
                            f"{task.project_name} "
                            f"{task.project_col_letter}{row}: "
                            f"merged value not numeric ({actual!r}); "
                            f"expected {expected_f:.4f}"
                        )
                        continue
                    if abs(actual_f - expected_f) > abs_tol:
                        mismatches.append(
                            f"{task.project_name} "
                            f"{task.project_col_letter}{row}: "
                            f"merged={actual_f:.4f} vs "
                            f"worker-reported={expected_f:.4f} "
                            f"(diff={actual_f - expected_f:+.4f})"
                        )
    finally:
        with contextlib.suppress(Exception):
            wb.close()

    return mismatches


def _merge_solved_workbooks(
    *,
    master_src: Path,
    others: list[Path],
    final_path: Path,
    partitions: list[list[SolveTask]],
    master_worker_id: int,
) -> None:
    """Copy per-project converged column values from other workers' SOLVED
    workbooks into the master, then save to final_path.

    Only the project-column cells are copied (rows 31, 32, 33, 37, 38, 39
    on Project Inputs — the cached convergence outputs). The rest of the
    workbook in the master is left as-is; per Caroline's spec, portfolio
    aggregates may be stale and will refresh on next interactive open.

    keep_vba=True is critical — without it, openpyxl strips the macro
    project on save. Verified working on real xlsm round-trips during
    development of the speed-win strip-sheets feature.
    """
    import openpyxl

    # Convergence-output rows that get hard-stamped by VBA into the per-
    # project column cells (see SolveHeadless.bas lines around 678-679,
    # 968-969). Reading from VBA-written values means we don't need to
    # recalc anything in openpyxl.
    OUTPUT_ROWS = (31, 32, 33, 37, 38, 39)

    # Load master with keep_vba so the .xlsm round-trips with macros intact.
    # If openpyxl couldn't grab the macro project blob, raise so the caller
    # falls back to the VBA-helper merge path (Excel COM) — saving without
    # macros would produce a silently broken .xlsm.
    wb_master = openpyxl.load_workbook(str(master_src), keep_vba=True)
    if getattr(wb_master, "vba_archive", None) is None:
        wb_master.close()
        raise RuntimeError(
            "openpyxl could not load the workbook's VBA project; "
            "saving via openpyxl would strip macros. Falling back."
        )
    ws_pi_master = wb_master["Project Inputs"] if "Project Inputs" in wb_master.sheetnames else None

    if ws_pi_master is None:
        # Nothing to merge into — just save master to final_path
        wb_master.save(str(final_path))
        return

    for other_path in others:
        if not other_path.exists():
            continue
        try:
            wb_other = openpyxl.load_workbook(str(other_path), keep_vba=True, data_only=True)
        except Exception:
            continue
        ws_pi_other = wb_other["Project Inputs"] if "Project Inputs" in wb_other.sheetnames else None
        if ws_pi_other is None:
            wb_other.close()
            continue

        # Figure out which worker this is and which task columns it owned
        # by matching against the file name's worker id suffix.
        wid = None
        stem = other_path.stem  # e.g., "<name>_SOLVED_w2"
        if "_w" in stem:
            try:
                wid = int(stem.rsplit("_w", 1)[1])
            except ValueError:
                wid = None
        if wid is None or wid >= len(partitions):
            wb_other.close()
            continue

        # Copy the convergence-output cells for each project this worker owned
        for task in partitions[wid]:
            # project_col_letter is the Excel column letter (e.g., "L"); convert
            # to numeric index for openpyxl.
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(task.project_col_letter)
            for row in OUTPUT_ROWS:
                src_val = ws_pi_other.cell(row=row, column=col_idx).value
                ws_pi_master.cell(row=row, column=col_idx).value = src_val

        wb_other.close()

    wb_master.save(str(final_path))
    wb_master.close()


def _merge_via_vba_fallback(
    *,
    master_src: Path,
    others: list[Path],
    final_path: Path,
    partitions: list[list[SolveTask]],
    worker_results: dict[int, dict],
    master_worker_id: int,
) -> None:
    """VBA-helper merge path. Used when openpyxl can't round-trip the .xlsm.

    Opens the master in a fresh Excel COM session, reads converged column
    values from each peer worker's SOLVED file via openpyxl (read-only, no
    save), and calls SolveHeadless's StampConvergedValuesHL via
    Application.Run to write them into the master. Excel handles the .xlsm
    save natively so the macro project stays intact.

    Requires SolveHeadless.bas's StampConvergedValuesHL to be present in
    the master workbook (it is, since master is a worker's own _SOLVED
    file and workers all imported the module before solving).
    """
    import openpyxl
    import pythoncom
    import win32com.client
    from openpyxl.utils import column_index_from_string

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        with contextlib.suppress(Exception):
            excel.AutomationSecurity = 1
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False

        wb = excel.Workbooks.Open(str(master_src), ReadOnly=False, UpdateLinks=0)

        # Verify StampConvergedValuesHL is actually callable before the
        # per-project loop. Without this check, every per-project Application.Run
        # would raise, the except-and-continue at the bottom of the loop would
        # swallow each error silently, and we'd SaveAs a workbook that looks
        # right but has only worker-0's converged values.
        try:
            vbp = wb.VBProject
            found = False
            for i in range(1, vbp.VBComponents.Count + 1):
                comp = vbp.VBComponents.Item(i)
                try:
                    cm = comp.CodeModule
                    if cm.Find("StampConvergedValuesHL", 1, 1, cm.CountOfLines, 999, True, False, False):
                        found = True
                        break
                except Exception:
                    continue
            if not found:
                raise RuntimeError(
                    "StampConvergedValuesHL not found in workbook VBA. "
                    "The module is required for the merge fallback path; "
                    "re-import SolveHeadless.bas via import_vba_module.py."
                )
        except Exception as verify_exc:
            log.error("  VBA merge precondition failed: %s", verify_exc)
            raise

        for other_path in others:
            if not other_path.exists():
                continue
            # Extract worker id from filename: "<stem>_SOLVED_w{id}.xlsm"
            stem = other_path.stem
            if "_w" not in stem:
                continue
            try:
                wid = int(stem.rsplit("_w", 1)[1])
            except ValueError:
                continue
            if wid >= len(partitions):
                continue

            # Read peer worker's converged values via openpyxl (read-only).
            try:
                wb_other = openpyxl.load_workbook(
                    str(other_path), data_only=True, read_only=True,
                )
            except Exception:
                continue
            ws_pi_other = wb_other["Project Inputs"] if "Project Inputs" in wb_other.sheetnames else None
            if ws_pi_other is None:
                wb_other.close()
                continue

            for task in partitions[wid]:
                col_idx = column_index_from_string(task.project_col_letter)
                npp = ws_pi_other.cell(row=38, column=col_idx).value
                dev_fee = ws_pi_other.cell(row=32, column=col_idx).value
                fmv = ws_pi_other.cell(row=33, column=col_idx).value
                live_irr = ws_pi_other.cell(row=37, column=col_idx).value
                appr_live = ws_pi_other.cell(row=31, column=col_idx).value
                npp_total = ws_pi_other.cell(row=39, column=col_idx).value
                # Pass zeros for cells the peer didn't populate so the VBA
                # Sub doesn't trip on Variant/Empty across the COM boundary
                try:
                    excel.Application.Run(
                        f"'{wb.Name}'!StampConvergedValuesHL",
                        int(col_idx),
                        float(npp or 0),
                        float(dev_fee or 0),
                        float(fmv or 0),
                        float(live_irr or 0),
                        float(appr_live or 0),
                        float(npp_total or 0),
                    )
                except Exception as stamp_exc:
                    log.warning(
                        "  StampConvergedValuesHL failed for col %d (%s): %s",
                        col_idx, task.project_name, stamp_exc,
                    )
            wb_other.close()

        # SaveAs to final canonical path with explicit FileFormat=52
        # (xlOpenXMLWorkbookMacroEnabled) so the VBA project survives.
        wb.SaveAs(str(final_path), FileFormat=52)

    finally:
        if wb is not None:
            with contextlib.suppress(Exception):
                wb.Close(SaveChanges=False)
        if excel is not None:
            with contextlib.suppress(Exception):
                excel.Quit()
        with contextlib.suppress(Exception):
            pythoncom.CoUninitialize()
