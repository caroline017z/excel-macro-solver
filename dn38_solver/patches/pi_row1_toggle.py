"""dn38_solver.patches.pi_row1_toggle — Restore Project Inputs row 1 toggle multipliers.

Background
----------
The 38DN pricing-model template uses a two-row toggle pattern in Project
Inputs:

  Row 1 (per-column multiplier) = =IF(Project_Toggle={col}2,1,0)
  Row 2 (per-column index)      = {col-1}2 + 1     (project index chain)

Downstream summary cells on the Table tab consume row 1 as a live/cached
selector:

  =IF('Project Inputs'!{col}1, Dashboard!$P$8, Table!{col}{row})

The template ships with row 1 populated only through column Q (10
projects). When a portfolio extends past that — SolarStone (28), Azimuth
(20+) — the row 2 chain is extended correctly, but row 1 is not. The
unset cells fall into a circular self-reference branch on Table, and two
accidentally-populated cells (V1=`=V19`, W1=`=W19`) point at project-name
strings, producing #VALUE! errors that propagate across hundreds of
Table/summary cells.

The errors are purely cosmetic — NPP, Dev Fee, FMV, DSCR are sourced via
column F live-pull (and row 371 hardcoded per project as of Tranche
7.12+), not via Table row 51/52. But they pollute the pre-flight scan
and any Table-derived deliverable.

Fix
---
For every PI column where row 2 holds a project index, ensure row 1
holds the canonical `=IF(Project_Toggle={col_letter}2,1,0)` formula.
Overwrite any stray non-matching formula. Skip cells already correct.
Idempotent; re-running on a clean workbook is a no-op.

Usage
-----
Currently exposed via the standalone `scripts.patch_pi_row1_toggle` CLI
for one-off file cleanup. Not wired into the orchestrator's preflight
framework — the issue is cosmetic-only, so wiring is optional. Future
work can promote this to a preflight finding (auto-fixable, similar to
A1 iterateDelta) inside `dn38_solver/shadow/preflight.py`.
"""
from __future__ import annotations

import logging
from pathlib import Path

import msgspec
import openpyxl

from dn38_solver.config import PROJECT_COL_RANGE
from dn38_solver.convert import col_letter

log = logging.getLogger(__name__)

# Row indices on the Project Inputs sheet.
_TOGGLE_MULTIPLIER_ROW: int = 1
_PROJECT_INDEX_ROW: int = 2

# Cells in the index row count as "populated" — meaning this column
# belongs to the project range and row 1 must carry a multiplier — when
# either an integer literal or a formula is present. Empty cells past
# the end of the chain are skipped.
_SHEET = "Project Inputs"


class PatchResult(msgspec.Struct, frozen=True, kw_only=True):
    """Outcome of a single pi_row1_toggle patch pass."""
    status: str  # "patched" | "clean" | "patch_failed"
    cells_filled: int          # missing → filled with canonical formula
    cells_overwritten: int     # stray formula replaced
    cells_already_correct: int
    columns_inspected: int
    error: str | None = None

    @property
    def total_changes(self) -> int:
        return self.cells_filled + self.cells_overwritten


def _expected_formula(col_letter_str: str) -> str:
    """Canonical row 1 formula for a given column."""
    return f"=IF(Project_Toggle={col_letter_str}2,1,0)"


def patch_pi_row1_toggle(workbook_path: Path | str) -> PatchResult:
    """Restore the Project Inputs row 1 toggle-multiplier chain in place.

    Loads with keep_vba=True so .xlsm macro storage survives the save.
    Saves only if at least one cell was modified.
    """
    p = Path(workbook_path)
    if not p.exists():
        return PatchResult(
            status="patch_failed",
            cells_filled=0,
            cells_overwritten=0,
            cells_already_correct=0,
            columns_inspected=0,
            error=f"File not found: {p}",
        )

    try:
        wb = openpyxl.load_workbook(str(p), keep_vba=True)
    except Exception as exc:
        return PatchResult(
            status="patch_failed",
            cells_filled=0,
            cells_overwritten=0,
            cells_already_correct=0,
            columns_inspected=0,
            error=f"Failed to load workbook: {exc}",
        )

    try:
        if _SHEET not in wb.sheetnames:
            return PatchResult(
                status="patch_failed",
                cells_filled=0,
                cells_overwritten=0,
                cells_already_correct=0,
                columns_inspected=0,
                error=f"Sheet '{_SHEET}' not found",
            )

        ws = wb[_SHEET]
        filled = 0
        overwritten = 0
        already_correct = 0
        inspected = 0

        for col in PROJECT_COL_RANGE:
            # Skip columns past the populated index chain — row 2 empty
            # means this column isn't part of the project range.
            index_cell_value = ws.cell(row=_PROJECT_INDEX_ROW, column=col).value
            if index_cell_value is None:
                continue

            inspected += 1
            col_str = col_letter(col)
            expected = _expected_formula(col_str)

            target_cell = ws.cell(row=_TOGGLE_MULTIPLIER_ROW, column=col)
            current = target_cell.value

            if current == expected:
                already_correct += 1
                continue

            if current is None:
                target_cell.value = expected
                filled += 1
                log.debug("Filled %s!%s1 (was empty)", _SHEET, col_str)
            else:
                target_cell.value = expected
                overwritten += 1
                log.warning(
                    "Overwrote %s!%s1: was %r, now %r",
                    _SHEET,
                    col_str,
                    current,
                    expected,
                )

        total_changes = filled + overwritten
        if total_changes > 0:
            wb.save(str(p))
            status = "patched"
        else:
            status = "clean"

        return PatchResult(
            status=status,
            cells_filled=filled,
            cells_overwritten=overwritten,
            cells_already_correct=already_correct,
            columns_inspected=inspected,
        )
    except Exception as exc:
        return PatchResult(
            status="patch_failed",
            cells_filled=0,
            cells_overwritten=0,
            cells_already_correct=0,
            columns_inspected=0,
            error=f"Patch failed: {exc}",
        )
    finally:
        wb.close()


def format_patch_report(r: PatchResult) -> str:
    """Render a one-line text report for log output."""
    if r.status == "patch_failed":
        return f"PI row 1 toggle patch: FAILED — {r.error}"
    if r.status == "clean":
        return (
            f"PI row 1 toggle: clean "
            f"({r.cells_already_correct}/{r.columns_inspected} columns correct)"
        )
    parts = [
        f"PI row 1 toggle: patched "
        f"({r.total_changes} change(s) across {r.columns_inspected} columns)"
    ]
    if r.cells_filled:
        parts.append(f"    filled {r.cells_filled} empty cell(s)")
    if r.cells_overwritten:
        parts.append(f"    overwrote {r.cells_overwritten} stray formula(s)")
    if r.cells_already_correct:
        parts.append(f"    {r.cells_already_correct} already correct")
    return "\n".join(parts)
