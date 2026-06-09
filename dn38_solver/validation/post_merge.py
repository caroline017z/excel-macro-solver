"""dn38_solver.validation.post_merge — Verify the merged _SOLVED.xlsm.

After parallel mode merges per-worker `_SOLVED.xlsm` files into one
canonical output, this module re-opens the merged file via openpyxl and
asserts that the hard-stamped convergence cells (Project Inputs rows
32/33/38/39/371 per project) match what each worker reported. Any
mismatch is a silent-corruption finding that forces the run to ERROR
and preserves the parent_tmp directory for forensics.

Rows 31 (Live Appraisal IRR) and 37 (Live Levered Pre-Tax IRR) are
deliberately NOT verified as of Tranche 7.12. The macro no longer
hardcodes those cells — they hold sticky-IF circular formulas
(IF(<col>2=$F$2, $F$<row>, <col><row>)). openpyxl reading with
data_only=True sees only cached values, and the cache is not
populated until the workbook is recalc-and-saved through Excel
interactively (which is the first thing Caroline does after the
merge anyway). Including 31/37 in the verifier would produce a
false "cached value missing" mismatch on every run.

Row 371 (Min Equity DSCR Multiple) IS hardcoded per project by the
post-solve stamp pass — that's the input lock Tranche 7.12 added to
preserve the audit chain. Verifying it catches silent-corruption in
the merge of the DSCR multiple just as the original rows 32/33/38/39
verification catches it for NPP/Dev Fee/FMV/NPP-total.

Why this lives in `validation/` and not `com/`:
- It does not touch COM. It's a pure openpyxl read against the merged
  artifact and could be called against ANY post-solve workbook (single-
  worker `_SOLVED.xlsm` included, where there is currently no equivalent
  gate — wired up in a follow-up).
- Post-merge is the structural sibling of `validation.parallel_correctness`
  (which compares parallel vs sequential outputs). Both validate
  end-state artifacts; neither runs the solver.
- Keeping `com/parallel_runner.py` focused on spawn + lifecycle + return
  shape makes the parallel runner legible at a glance.
"""
from __future__ import annotations

import contextlib
import logging
from pathlib import Path

from openpyxl.utils import column_index_from_string

from dn38_solver.types import SolveTask

log = logging.getLogger(__name__)

# The five rows VBA hard-stamps via cell-self-assign at end of each
# project's solve, post-Tranche-7.12. These are the cells the verifier
# checks against worker-reported expected values; their values are
# exactly the ones that ship to IC (or, for row 371, the upstream input
# lock that drives the IRR cascade by formula), so silent corruption
# here is the highest-cost bug class the merge step can produce.
#
# Rows 31 and 37 are deliberately excluded — Tranche 7.12 left them as
# sticky-IF formulas; openpyxl read with data_only=True can't see their
# values until the workbook is recalc-and-saved through Excel.
HARD_STAMPED_ROWS: tuple[int, ...] = (32, 33, 38, 39, 371)

# Per-row tolerance for the post-merge gate. Rows 32/33/38/371 are
# dollar-per-watt or unitless multiplier values (typical magnitude < $5
# or < 2x); 1¢ / 0.01x slack absorbs openpyxl serialization rounding
# without missing real bugs. Row 39 is NPP $ total — typical $5M-$50M
# for the portfolios this tool sees. A flat $1 tolerance would silently
# let a six-figure corruption slip past on a $30M project ($0.5M =
# 1.7e-5 relative). Row 39 uses a scaled tolerance (see
# `tolerance_for_row`).
_VERIFY_TOL_BY_ROW: dict[int, float] = {
    32: 0.01,
    33: 0.01,
    38: 0.01,
    # Row 39 entry is unused — `tolerance_for_row` short-circuits with
    # the scaled formula. Kept for symmetry / discoverability.
    39: 1.0,
    # Row 371: Min Equity DSCR Multiple (Tranche 7.12). Typical solved
    # values are 1.0x-2.0x; 0.01x slack absorbs roundoff without
    # missing a corrupt copy that lands the wrong project's DSCR.
    371: 0.01,
}


def tolerance_for_row(row: int, expected: float) -> float:
    """Return per-row absolute tolerance, scaled for row 39's magnitude.

    Row 39 (NPP $ total) is in the millions; a flat $1 tolerance would
    miss six-figure corruption on a $30M project. max($1, 1e-6 ×
    |expected|) keeps the absolute floor for noise on small projects
    while the relative ceiling catches partial-cell-corruption on large
    ones. Other rows are flat $0.01/W, which absorbs openpyxl roundtrip
    noise without missing real bugs (the values are typically <$5/W).
    """
    if row == 39:
        return max(1.0, 1e-6 * abs(expected))
    return _VERIFY_TOL_BY_ROW.get(row, 0.01)


def expected_address_for_row(row: int, col_letter: str) -> str:
    """Return the `solved_values` key the verifier should look up for a
    given hard-stamped row in a project's column.

    Row 371 (Min Equity DSCR Multiple) is hard-stamped by the macro to
    the per-column cell, but the worker only captures the LIVE single
    cell `PT Returns!F129` (that's what READ_CELLS_TEMPLATES holds).
    The macro reads PT!F129 mid-solve and stamps it onto PI!<col>!371
    at end-of-solve, so the worker-reported PT!F129 IS the expected
    value for the merged PI!<col>!371. Map row 371 to that key.

    Rows 31 (Live Appraisal IRR) and 37 (Live IRR) were similarly
    mapped to F31/F37 before Tranche 7.12, but they're no longer in
    HARD_STAMPED_ROWS — the macro leaves them as sticky-IF formulas
    so the audit chain stays intact. They'll never reach this function
    via the verifier loop today, but defensive callers can still pass
    them and get the F-column key back.
    """
    if row == 371:
        return "PT Returns!F129"
    if row in (31, 37):
        return f"Project Inputs!F{row}"
    return f"Project Inputs!{col_letter}{row}"


def verify_merged_file(
    *,
    final_path: Path,
    worker_results: dict[int, dict],
    partitions: list[list[SolveTask]],
) -> list[str]:
    """Re-open the merged xlsm and assert hard-stamped cells match.

    Returns a list of mismatch strings (empty list = clean merge).

    Why this exists: every merge path (openpyxl AND VBA-helper) iterates
    per-project, copying convergence values from peer worker files into
    the master. A silent failure inside that loop — wrong column letter,
    swallowed COM exception, peer file corrupted — would leave the merged
    file with PRE-solve values in some columns and the user would only
    notice when an IC memo derived from the file produced wrong numbers.
    Better to surface the inconsistency at run-end and force the user to
    consult the per-worker outputs.

    The check covers BOTH converged and non-converged projects: a project
    that didn't converge still got its row 32/33/38/39/371 cells
    hard-stamped by VBA (cell-self-assign in SolveOneProjectByColHL), so
    a corrupt merge of a non-converged project's column is just as bad
    as a converged one's. Projects that were `not_attempted` (worker
    crashed before reaching them) are skipped — they have no expected
    value to check against.
    """
    import openpyxl

    mismatches: list[str] = []

    try:
        # C26: read-only value verify, never saved — no keep_vba (avoids the
        # leaked in-memory vba_archive ZipFile handle).
        wb = openpyxl.load_workbook(
            str(final_path), data_only=True, read_only=True,
        )
    except Exception as exc:
        return [f"merged file unreadable: {exc}"]

    try:
        if "Project Inputs" not in wb.sheetnames:
            return ["merged file has no 'Project Inputs' sheet"]
        ws = wb["Project Inputs"]
        # Capture sheet height once so the per-cell error message can
        # distinguish "row exists but value cached as None" from "row
        # past sheet end" — the two have different remediation. In
        # read_only mode max_row may be None for sheets without a known
        # dimensions tag; treat that as "unknown, fall back to None
        # message" rather than crashing.
        ws_max_row = ws.max_row or 0

        # Build offset -> result lookup. project_offset is the column-
        # specific identity (set in direct_runner from
        # SolveTask.project_offset) and is GUARANTEED unique per project
        # — project_name is not, and duplicate names across LLCs land on
        # different workers under round-robin partitioning. Keying on
        # name silently let a corrupt merge of the FIRST duplicate's
        # column slip past unchecked because the LAST encountered worker
        # overwrote the name lookup. (Bug found in round-3 review.)
        offset_to_result: dict[int, dict] = {}
        for wresult in worker_results.values():
            for pr in wresult.get("project_results", []):
                offset = pr.get("project_offset")
                if offset is None:
                    meta = pr.get("meta") or {}
                    offset = meta.get("project_offset") or meta.get("offset")
                if offset is None:
                    log.warning(
                        "  Verify: project_result has no project_offset "
                        "(name=%r); cannot include in verification.",
                        pr.get("project_name"),
                    )
                    continue
                offset_int = int(offset)
                if offset_int in offset_to_result:
                    log.error(
                        "  Verify: duplicate project_offset %d across "
                        "workers — this is a worker logic bug; using "
                        "first occurrence.", offset_int,
                    )
                    continue
                offset_to_result[offset_int] = pr

        # Read all hard-stamped rows in a single iter_rows pass.
        # values_only=True with manual offset tracking — values_only=False
        # returns EmptyCell stubs for blank cells in read_only mode and
        # EmptyCell has no .row / .column attributes.
        cell_values: dict[tuple[int, int], object] = {}
        first_row = min(HARD_STAMPED_ROWS)
        for row_offset, row_values in enumerate(ws.iter_rows(
            min_row=first_row,
            max_row=max(HARD_STAMPED_ROWS),
            values_only=True,
        )):
            actual_row = first_row + row_offset
            if actual_row not in HARD_STAMPED_ROWS:
                continue
            for col_offset, value in enumerate(row_values):
                cell_values[(actual_row, col_offset + 1)] = value

        cells_checked = 0
        for tasks_slice in partitions:
            for task in tasks_slice:
                pr = offset_to_result.get(task.project_offset)
                if pr is None:
                    continue
                if pr.get("status") in ("not_attempted", "skipped"):
                    continue
                # Defense-in-depth: bypass on meta["mode"] sentinel too,
                # in case a future worker version forgets to translate
                # skip-mode into status="skipped". The fast-skip leaves
                # the project's hard-stamped cells untouched; the
                # worker's post-solve F-column read returns an Excel
                # error that COM marshals as a numeric sentinel while
                # openpyxl reads it as '#NUM!' — comparing the two
                # produces false-positive mismatches.
                #
                # stamp_skipped:<context> is written by HardStampNumericHL
                # when a per-cell stamp fails mid-project, leaving the
                # hard-stamped row in a partial state. Same bypass logic
                # applies.
                meta = pr.get("meta") or {}
                mode = meta.get("mode")
                if isinstance(mode, str) and (
                    mode.startswith("skipped:") or mode.startswith("stamp_skipped:")
                ):
                    continue
                sv = pr.get("solved_values", {})
                col_idx = column_index_from_string(task.project_col_letter)
                for row in HARD_STAMPED_ROWS:
                    expected_key = expected_address_for_row(
                        row, task.project_col_letter,
                    )
                    expected = sv.get(expected_key)
                    if expected is None:
                        # Worker didn't report this cell at all — happens
                        # for non-converged projects whose late-stage reads
                        # legitimately produced None. Skip silently.
                        continue
                    try:
                        expected_f = float(expected)
                    except (TypeError, ValueError):
                        # Non-numeric expected value (e.g., worker
                        # captured an Excel error string like "#DIV/0!").
                        # Flag so the operator knows a pre-existing
                        # error is being merged into the canonical file.
                        mismatches.append(
                            f"{task.project_name} "
                            f"{task.project_col_letter}{row}: "
                            f"worker-reported value is non-numeric "
                            f"({expected!r}); merged file likely contains "
                            f"the same error — investigate at source"
                        )
                        continue
                    cells_checked += 1
                    tol = tolerance_for_row(row, expected_f)
                    actual = cell_values.get((row, col_idx))
                    if actual is None:
                        if ws_max_row and row > ws_max_row:
                            mismatches.append(
                                f"{task.project_name} "
                                f"{task.project_col_letter}{row}: "
                                f"row {row} is past sheet end "
                                f"(max_row={ws_max_row}) — workbook variant "
                                f"or stripped sheet; expected {expected_f:.4f}"
                            )
                        else:
                            mismatches.append(
                                f"{task.project_name} "
                                f"{task.project_col_letter}{row}: "
                                f"merged value is None (cached value missing "
                                f"— the file may need a one-time interactive "
                                f"open + save in Excel); expected {expected_f:.4f}"
                            )
                        continue
                    try:
                        actual_f = float(actual)
                    except (TypeError, ValueError):
                        mismatches.append(
                            f"{task.project_name} "
                            f"{task.project_col_letter}{row}: "
                            f"merged value not numeric ({actual!r}); "
                            f"expected {expected_f:.4f}"
                        )
                        continue
                    if abs(actual_f - expected_f) > tol:
                        mismatches.append(
                            f"{task.project_name} "
                            f"{task.project_col_letter}{row}: "
                            f"merged={actual_f:.4f} vs "
                            f"worker-reported={expected_f:.4f} "
                            f"(diff={actual_f - expected_f:+.4f}, tol={tol})"
                        )
        if not mismatches:
            log.info("    Verification covered %d cells", cells_checked)
    finally:
        with contextlib.suppress(Exception):
            wb.close()

    return mismatches
