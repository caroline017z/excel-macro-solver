"""dn38_solver.reporting.export_xlsx — Clean summary .xlsx for sharing.

Generates a branded one-pager with solve results, suitable for email.
Includes NPP, Dev Fee, DSCR Multiple, and all convergence details
so the run is fully auditable and replicable.
"""
from __future__ import annotations

import io
import logging

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import absolute_coordinate, get_column_letter

from dn38_solver.config import CELL_DSCR_MULTIPLE, LABEL_TO_ROW
from dn38_solver.types import RunRecord

log = logging.getLogger(__name__)

# Sheet names — single source so cross-sheet check formulas can't drift.
SUMMARY_SHEET_NAME = "Solve Summary"
CHECKS_SHEET_NAME = "Checks"

# First per-project data row on the Solve Summary tab. Shared by the summary
# writer and the Checks tab's cross-sheet links so the two layouts can't drift.
SUMMARY_FIRST_DATA_ROW = 5

# Project Inputs row that holds the per-project Min Equity DSCR Multiple.
# Pulled from config (LABEL_TO_ROW) so a model re-map updates this in one place.
DSCR_MULTIPLE_ROW = LABEL_TO_ROW["Min Equity DSCR Multiple"]

# Convergence tolerances — DISPLAY MIRROR of the Private Const values in
# SolveHeadless.bas (the canonical source the macro actually solves against).
# Surfaced as input cells on the Checks tab so the in-sheet PASS/FAIL ties are
# self-documenting. Keep in sync with the .bas if the macro's bands change;
# they are display/audit values here, not the values the solver reads.
TOL_IRR = 0.0003              # 3 bps — IRR and appraisal gap acceptance tol
EQUITY_TARGET = 0.10          # Min Equity target (10%)
EQUITY_BAND_STRICT = 0.0025   # +/-0.25pp — strict convergence band (reference)
EQUITY_BAND_RELAXED = 0.005   # +/-0.50pp — relaxed-tier / acceptance band

# ---------------------------------------------------------------------------
# 38DN brand styles (Aptos Narrow / navy / teal)
# ---------------------------------------------------------------------------

_FONT_TITLE = Font(name="Aptos Display", size=14, bold=True, color="FFFFFF")
_FONT_SUBTITLE = Font(name="Aptos Narrow", size=10, color="A0A8C0")
_FONT_HEADER = Font(name="Aptos Display", size=10, bold=True, color="FFFFFF")
_FONT_LABEL = Font(name="Aptos Narrow", size=10, color="050D25")
_FONT_DATA = Font(name="Aptos Narrow", size=10, color="050D25")
_FONT_DATA_BOLD = Font(name="Aptos Narrow", size=10, bold=True, color="050D25")
_FONT_PASS = Font(name="Aptos Narrow", size=10, bold=True, color="3A7D44")
_FONT_FAIL = Font(name="Aptos Narrow", size=10, bold=True, color="B83230")
# Cell-semantic colors (38DN + xlsx-author convention):
#   green = cross-sheet link, blue italic = hardcoded input.
_FONT_LINK = Font(name="Aptos Narrow", size=10, color="008000")
_FONT_INPUT = Font(name="Aptos Narrow", size=10, italic=True, color="0000FF")

_FILL_NAVY = PatternFill(start_color="050D25", end_color="050D25", fill_type="solid")
_FILL_NAVY2 = PatternFill(start_color="212B48", end_color="212B48", fill_type="solid")
_FILL_LIGHT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_FILL_WHITE = PatternFill(fill_type=None)

_THIN = Side(style="thin", color="D0D4DC")
_BORDER = Border(bottom=_THIN)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Number formats
_NF_DPW = '"$"0.0000'
_NF_PCT = "0.00%"
_NF_DSCR = "0.0000\"x\""
_NF_DOLLAR = '"$"#,##0'
_NF_SEC = '0.0"s"'
_NF_PCT_FINE = "0.0000%"   # bps-resolution gaps


def generate_summary_xlsx(record: RunRecord) -> io.BytesIO:
    """Generate a branded summary .xlsx from a RunRecord.

    Layout:
      Row 1-2: Navy banner with title + timestamp
      Row 3: Blank spacer
      Row 4: Column headers (navy)
      Row 5+: One row per project with key metrics
      Bottom: Run metadata (batch ID, solver mode, total duration)

    Columns:
      Project | State | MW(DC) | NPP ($/W) | Dev Fee ($/W) | FMV ($/W) |
      DSCR Multiple | Target IRR | Live IRR | Appraisal IRR | WACC Target |
      Equity % | Iterations | Status
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SUMMARY_SHEET_NAME

    # --- Column definitions ---
    columns = [
        ("Project", 30, _ALIGN_LEFT, None),
        ("NPP ($/W)", 14, _ALIGN_RIGHT, _NF_DPW),
        ("Step-Up Dev Fee ($/W)", 20, _ALIGN_RIGHT, _NF_DPW),
        ("FMV ($/W)", 14, _ALIGN_RIGHT, _NF_DPW),
        ("NPP ($)", 16, _ALIGN_RIGHT, _NF_DOLLAR),
        ("DSCR Multiple", 16, _ALIGN_RIGHT, _NF_DSCR),
        ("Target IRR", 13, _ALIGN_RIGHT, _NF_PCT),
        ("Live IRR", 13, _ALIGN_RIGHT, _NF_PCT),
        ("IRR Gap", 11, _ALIGN_RIGHT, _NF_PCT),
        ("Appraisal IRR", 15, _ALIGN_RIGHT, _NF_PCT),
        ("WACC Target", 13, _ALIGN_RIGHT, _NF_PCT),
        ("Appraisal Gap", 14, _ALIGN_RIGHT, _NF_PCT),
        ("Equity %", 11, _ALIGN_RIGHT, _NF_PCT),
        ("Iterations", 11, _ALIGN_CENTER, None),
        ("Status", 14, _ALIGN_CENTER, None),
    ]

    n_cols = len(columns)

    # Set column widths
    for i, (_, width, _, _) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- Row 1: Navy banner ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "38DN Pricing Model — Solve Summary"
    title_cell.font = _FONT_TITLE
    title_cell.fill = _FILL_NAVY
    title_cell.alignment = _ALIGN_LEFT
    ws.row_dimensions[1].height = 32

    for c in range(2, n_cols + 1):
        ws.cell(row=1, column=c).fill = _FILL_NAVY

    # --- Row 2: Subtitle with metadata ---
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub_cell = ws.cell(row=2, column=1)
    ts_display = record.run_timestamp[:19].replace("T", " ")
    sub_cell.value = (
        f"Workbook: {record.workbook_name}  |  "
        f"Run: {ts_display} UTC  |  "
        f"Batch: {record.batch_id}  |  "
        f"Mode: {record.solver_mode}  |  "
        f"Duration: {record.total_duration_sec:.1f}s"
    )
    sub_cell.font = _FONT_SUBTITLE
    sub_cell.fill = _FILL_NAVY2
    sub_cell.alignment = _ALIGN_LEFT
    ws.row_dimensions[2].height = 22

    for c in range(2, n_cols + 1):
        ws.cell(row=2, column=c).fill = _FILL_NAVY2

    # --- Row 3: Spacer ---
    ws.row_dimensions[3].height = 6

    # --- Row 4: Column headers ---
    for i, (label, _, align, _) in enumerate(columns, 1):
        cell = ws.cell(row=4, column=i, value=label)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_NAVY
        cell.alignment = align
        cell.border = _BORDER
    ws.row_dimensions[4].height = 24

    # --- Data rows (one per project) ---
    # Gap columns are written as live formulas (=ABS(...)) referencing the
    # adjacent input cells, per 38DN convention "every output formula-linked":
    #   IRR Gap (col I) = ABS(Live IRR H - Target IRR G)
    #   Appraisal Gap (col L) = ABS(Appraisal IRR J - WACC Target K)
    # When either input is missing, the formula is omitted so the cell stays
    # visually blank instead of computing 0 from blanks.
    for row_idx, proj in enumerate(record.projects, SUMMARY_FIRST_DATA_ROW):
        irr_gap = (
            f"=ABS(H{row_idx}-G{row_idx})"
            if proj.live_irr is not None and proj.target_irr is not None
            else None
        )
        appr_gap = (
            f"=ABS(J{row_idx}-K{row_idx})"
            if proj.appraisal_live is not None and proj.wacc_target is not None
            else None
        )

        values = [
            proj.name,
            proj.npp_per_w,
            proj.dev_fee_per_w,
            proj.fmv_per_w,
            proj.npp_total,
            proj.dscr_multiple,
            proj.target_irr,
            proj.live_irr,
            irr_gap,
            proj.appraisal_live,
            proj.wacc_target,
            appr_gap,
            proj.equity_pct,
            proj.iterations,
            "CONVERGED" if proj.converged else "CHECK",
        ]

        fill = _FILL_LIGHT if row_idx % 2 == 1 else _FILL_WHITE
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _, _, align, nf = columns[col_idx - 1]
            cell.font = _FONT_DATA
            cell.alignment = align
            cell.fill = fill
            cell.border = _BORDER
            if nf and val is not None:
                cell.number_format = nf

        # Style the status column
        status_cell = ws.cell(row=row_idx, column=n_cols)
        status_cell.font = _FONT_PASS if proj.converged else _FONT_FAIL

        # Bold the key solve outputs
        for bold_col in (2, 3, 6):  # NPP, Dev Fee, DSCR
            ws.cell(row=row_idx, column=bold_col).font = _FONT_DATA_BOLD

    # --- Bottom: Convergence reference ---
    bottom_row = SUMMARY_FIRST_DATA_ROW + len(record.projects) + 1
    ws.merge_cells(start_row=bottom_row, start_column=1, end_row=bottom_row, end_column=n_cols)
    ref_cell = ws.cell(row=bottom_row, column=1)
    ref_cell.value = (
        "Tolerances — IRR: 0.03% (3 bps)  |  "
        "Equity: +/-0.50pp of 10%  |  "
        "DSCR bounds: 0.5x-5.0x  |  "
        "Max iterations: 8 outer x 6 inner"
    )
    ref_cell.font = Font(name="Aptos Narrow", size=9, italic=True, color="7A8291")
    ref_cell.alignment = _ALIGN_LEFT

    # Freeze panes (header row)
    ws.freeze_panes = "A5"

    # Print settings
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # --- Audit Checks tab (convergence ties + DSCR reconciliation) ---
    _build_checks_sheet(wb, record)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log.info("Generated summary XLSX (%d projects)", len(record.projects))
    return buf


def _build_checks_sheet(wb: openpyxl.Workbook, record: RunRecord) -> None:
    """Append an audit "Checks" tab to ``wb``.

    Two sections:

    1. Per-project convergence ties — IRR, appraisal, and Min Equity gaps
       recomputed in-sheet as PASS/FAIL against tolerance input cells. Live
       values are GREEN cross-sheet links to the Solve Summary tab (so the
       checks recompute if the summary is edited); tolerances are BLUE input
       cells mirroring SolveHeadless.bas. These ties are a documentation /
       audit cross-check, not an independent re-solve — they restate the
       macro's verdict (the authoritative "Solver Tier" column) in auditable
       form, against the same acceptance tolerances shown on the summary.

    2. DSCR Multiple reconciliation — each project's converged DSCR multiple as
       a STATIC, copy-pasteable value alongside its paste target
       (``Project Inputs!<col>371``) and the live dynamic source
       (``='PT Returns'!$F$129``). Lets an auditor reproduce a project's solved
       Min Equity state by pasting one value — without reading and resetting
       the live PT Returns DSCR — and shows the formula to restore for a fully
       dynamic Min Equity solve.
    """
    ws = wb.create_sheet(CHECKS_SHEET_NAME)

    # Per-project checks table: (header, width, alignment). Column letters used
    # by the formulas below: B Live IRR, C Target, D IRR Gap, E IRR OK,
    # F Appraisal, G WACC, H Appr Gap, I Appr OK, J Equity, K Eq Gap,
    # L Eq OK, M Solver Tier, N All Pass.
    check_cols = [
        ("Project", 30, _ALIGN_LEFT),
        ("Live IRR", 12, _ALIGN_RIGHT),
        ("Target IRR", 12, _ALIGN_RIGHT),
        ("IRR Gap", 11, _ALIGN_RIGHT),
        ("IRR OK?", 9, _ALIGN_CENTER),
        ("Appraisal IRR", 13, _ALIGN_RIGHT),
        ("WACC Target", 12, _ALIGN_RIGHT),
        ("Appr Gap", 11, _ALIGN_RIGHT),
        ("Appr OK?", 9, _ALIGN_CENTER),
        ("Equity %", 11, _ALIGN_RIGHT),
        ("Equity Gap", 11, _ALIGN_RIGHT),
        ("Equity OK?", 10, _ALIGN_CENTER),
        ("Solver Tier", 13, _ALIGN_CENTER),
        ("All Checks Pass?", 16, _ALIGN_CENTER),
    ]
    n_cols = len(check_cols)
    for i, (_, width, _) in enumerate(check_cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- Row 1: banner ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value="38DN Solve — Audit Checks & DSCR Reconciliation")
    c.font = _FONT_TITLE
    c.fill = _FILL_NAVY
    c.alignment = _ALIGN_LEFT
    ws.row_dimensions[1].height = 32
    for col in range(2, n_cols + 1):
        ws.cell(row=1, column=col).fill = _FILL_NAVY

    # --- Row 2: subtitle ---
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c = ws.cell(row=2, column=1, value=(
        "Convergence ties recomputed in-sheet from green Solve Summary links; "
        "tolerances (blue) mirror SolveHeadless.bas. 'Solver Tier' is the "
        "macro's authoritative verdict. DSCR multiples below are static "
        "paste-values for live-model audit."
    ))
    c.font = _FONT_SUBTITLE
    c.fill = _FILL_NAVY2
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30
    for col in range(2, n_cols + 1):
        ws.cell(row=2, column=col).fill = _FILL_NAVY2

    # --- Tolerance input block (blue inputs; mirror of SolveHeadless.bas) ---
    tc = ws.cell(row=4, column=1, value="Convergence Tolerances (mirror of SolveHeadless.bas)")
    tc.font = _FONT_DATA_BOLD
    tc.alignment = _ALIGN_LEFT
    tol_rows = [
        ("IRR / Appraisal gap tolerance", TOL_IRR, _NF_PCT_FINE),
        ("Min Equity target", EQUITY_TARGET, _NF_PCT),
        ("Equity strict band (+/-, reference)", EQUITY_BAND_STRICT, _NF_PCT),
        ("Equity acceptance band (+/-)", EQUITY_BAND_RELAXED, _NF_PCT),
    ]
    tol_start = 5
    for i, (label, val, nf) in enumerate(tol_rows):
        r = tol_start + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _FONT_LABEL
        lc.alignment = _ALIGN_LEFT
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = _FONT_INPUT
        vc.alignment = _ALIGN_RIGHT
        vc.number_format = nf
    # Absolute refs for the per-project tie formulas.
    irr_tol_ref = f"$B${tol_start}"
    eq_target_ref = f"$B${tol_start + 1}"
    eq_band_ref = f"$B${tol_start + 3}"   # acceptance band drives Equity OK?

    # --- Per-project convergence checks ---
    hdr_row = tol_start + len(tol_rows) + 1
    for i, (label, _, align) in enumerate(check_cols, 1):
        cell = ws.cell(row=hdr_row, column=i, value=label)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_NAVY
        cell.alignment = align
        cell.border = _BORDER
    ws.row_dimensions[hdr_row].height = 24

    first_data = hdr_row + 1
    sq = f"'{SUMMARY_SHEET_NAME}'"
    for p, proj in enumerate(record.projects):
        r = first_data + p
        sr = SUMMARY_FIRST_DATA_ROW + p   # matching Solve Summary data row
        fill = _FILL_LIGHT if r % 2 == 1 else _FILL_WHITE

        name_c = ws.cell(row=r, column=1, value=proj.name)
        name_c.font = _FONT_DATA
        name_c.alignment = _ALIGN_LEFT
        name_c.fill = fill
        name_c.border = _BORDER

        tier_c = ws.cell(row=r, column=13, value=proj.convergence_tier)
        tier_c.font = _FONT_DATA_BOLD
        tier_c.alignment = _ALIGN_CENTER
        tier_c.fill = fill
        tier_c.border = _BORDER

        if proj.convergence_tier in ("skipped", "not_attempted"):
            # No solved metrics to tie — flag and move on.
            for col in range(2, n_cols + 1):
                if col == 13:
                    continue
                blank = ws.cell(row=r, column=col)
                blank.fill = fill
                blank.border = _BORDER
            pass_c = ws.cell(row=r, column=n_cols, value=proj.convergence_tier.upper())
            pass_c.font = _FONT_FAIL
            pass_c.alignment = _ALIGN_CENTER
            continue

        # Linked live values (green = cross-sheet link).
        links = {
            2: f"={sq}!H{sr}",    # Live IRR
            3: f"={sq}!G{sr}",    # Target IRR
            6: f"={sq}!J{sr}",    # Appraisal IRR
            7: f"={sq}!K{sr}",    # WACC Target
            10: f"={sq}!M{sr}",   # Equity %
        }
        for col, formula in links.items():
            cc = ws.cell(row=r, column=col, value=formula)
            cc.font = _FONT_LINK
            cc.alignment = _ALIGN_RIGHT
            cc.number_format = _NF_PCT
            cc.fill = fill
            cc.border = _BORDER

        # Computed gaps (black formulas).
        gaps = {
            4: f"=ABS(B{r}-C{r})",                  # IRR gap
            8: f"=ABS(F{r}-G{r})",                  # Appraisal gap
            11: f"=ABS(J{r}-{eq_target_ref})",      # Equity gap vs target
        }
        for col, formula in gaps.items():
            gc = ws.cell(row=r, column=col, value=formula)
            gc.font = _FONT_DATA
            gc.alignment = _ALIGN_RIGHT
            gc.number_format = _NF_PCT_FINE
            gc.fill = fill
            gc.border = _BORDER

        # PASS/FAIL ties (black formulas → TRUE/FALSE).
        oks = {
            5: f"=D{r}<={irr_tol_ref}",
            9: f"=H{r}<={irr_tol_ref}",
            12: f"=K{r}<={eq_band_ref}",
        }
        for col, formula in oks.items():
            kc = ws.cell(row=r, column=col, value=formula)
            kc.font = _FONT_DATA
            kc.alignment = _ALIGN_CENTER
            kc.fill = fill
            kc.border = _BORDER

        all_c = ws.cell(row=r, column=14, value=f"=AND(E{r},I{r},L{r})")
        all_c.font = _FONT_DATA_BOLD
        all_c.alignment = _ALIGN_CENTER
        all_c.fill = fill
        all_c.border = _BORDER

    # --- DSCR Multiple reconciliation ---
    dscr_title_row = first_data + len(record.projects) + 1
    ws.merge_cells(start_row=dscr_title_row, start_column=1, end_row=dscr_title_row, end_column=n_cols)
    dt = ws.cell(row=dscr_title_row, column=1, value="DSCR Multiple — Audit Reconciliation")
    dt.font = _FONT_TITLE
    dt.fill = _FILL_NAVY
    dt.alignment = _ALIGN_LEFT
    ws.row_dimensions[dscr_title_row].height = 26
    for col in range(2, n_cols + 1):
        ws.cell(row=dscr_title_row, column=col).fill = _FILL_NAVY

    note_row = dscr_title_row + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=n_cols)
    note = ws.cell(row=note_row, column=1, value=(
        "To reproduce a project's solved Min Equity state: copy its Converged "
        "DSCR Multiple (static value) and paste-special as a VALUE into the "
        "Paste Target cell in the live model, then re-solve — this avoids "
        "reading and resetting the live 'PT Returns' DSCR. To return a project "
        "to a fully dynamic Min Equity solve, enter the Live Dynamic Source "
        "formula into the Paste Target cell instead."
    ))
    note.font = Font(name="Aptos Narrow", size=9, italic=True, color="7A8291")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[note_row].height = 42

    dscr_hdr_row = note_row + 1
    dscr_cols = [
        ("Project", _ALIGN_LEFT),
        ("Model Col", _ALIGN_CENTER),
        ("Converged DSCR Multiple", _ALIGN_RIGHT),
        ("Paste Target (live model)", _ALIGN_LEFT),
        ("Live Dynamic Source (restore as =)", _ALIGN_LEFT),
    ]
    for i, (label, align) in enumerate(dscr_cols, 1):
        cell = ws.cell(row=dscr_hdr_row, column=i, value=label)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_NAVY
        cell.alignment = align
        cell.border = _BORDER
    ws.row_dimensions[dscr_hdr_row].height = 24

    # Live dynamic source shown as literal text (leading "'" of the quoted
    # sheet name keeps openpyxl from parsing it as a formula). Auditors prepend
    # "=" to restore. Sheet + cell pulled from config so a re-map propagates.
    dynamic_src = (
        f"'{CELL_DSCR_MULTIPLE.sheet}'!"
        f"{absolute_coordinate(CELL_DSCR_MULTIPLE.address)}"
    )
    for p, proj in enumerate(record.projects):
        r = dscr_hdr_row + 1 + p
        fill = _FILL_LIGHT if r % 2 == 1 else _FILL_WHITE

        name_c = ws.cell(row=r, column=1, value=proj.name)
        name_c.font = _FONT_DATA
        name_c.alignment = _ALIGN_LEFT
        name_c.fill = fill
        name_c.border = _BORDER

        col_c = ws.cell(row=r, column=2, value=proj.col_letter)
        col_c.font = _FONT_DATA
        col_c.alignment = _ALIGN_CENTER
        col_c.fill = fill
        col_c.border = _BORDER

        # Static converged DSCR — copy/paste-ready (not a formula, not a link).
        dscr_c = ws.cell(row=r, column=3, value=proj.dscr_multiple)
        dscr_c.font = _FONT_DATA_BOLD
        dscr_c.alignment = _ALIGN_RIGHT
        dscr_c.fill = fill
        dscr_c.border = _BORDER
        if proj.dscr_multiple is not None:
            dscr_c.number_format = _NF_DSCR

        target_c = ws.cell(
            row=r, column=4,
            value=f"'Project Inputs'!{proj.col_letter}{DSCR_MULTIPLE_ROW}",
        )
        target_c.font = _FONT_DATA
        target_c.alignment = _ALIGN_LEFT
        target_c.fill = fill
        target_c.border = _BORDER

        src_c = ws.cell(row=r, column=5, value=dynamic_src)
        src_c.font = _FONT_DATA
        src_c.alignment = _ALIGN_LEFT
        src_c.fill = fill
        src_c.border = _BORDER

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1).coordinate
