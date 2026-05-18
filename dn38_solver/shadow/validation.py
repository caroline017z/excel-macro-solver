"""dn38_solver.shadow.validation — Pure-Python formula-error scan.

Scans a workbook's cached values for Excel errors (#REF!, #DIV/0!, etc.)
without requiring Excel COM or a LibreOffice install. Output shape mirrors
the xlsx skill's `recalc.py` so the two can be swapped if a host adds
LibreOffice (recalc + scan) later.

Two roles in the solver pipeline:
- **Pre-flight** (orchestrator Phase 1): catch a broken input workbook
  before paying the COM startup + multi-minute solve cost.
- **Post-export gate** (direct_runner after SaveAs): enforce the
  conventions.md "ZERO formula errors" rule on every saved _SOLVED.xlsx.
"""
from __future__ import annotations

import logging
from pathlib import Path

import msgspec
import openpyxl

log = logging.getLogger(__name__)

# Error tokens recognized by Excel. Matches xlsx skill's recalc.py set.
EXCEL_ERROR_TOKENS: tuple[str, ...] = (
    "#VALUE!",
    "#DIV/0!",
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
)

# Cap per-error-type location list so reports stay bounded on heavily
# broken workbooks. 20 is enough to triage; full counts are still tracked.
_MAX_LOCATIONS_PER_TYPE = 20


class ErrorTypeDetail(msgspec.Struct, frozen=True, kw_only=True):
    """Per-error-token rollup."""
    count: int
    locations: tuple[str, ...]  # "{sheet}!{coord}" form, capped


class WorkbookValidation(msgspec.Struct, frozen=True, kw_only=True):
    """Result of a workbook formula-error scan."""
    status: str  # "success" | "errors_found" | "scan_failed"
    total_errors: int
    total_formulas: int
    error_summary: dict[str, ErrorTypeDetail]
    error: str | None = None  # populated when status == "scan_failed"

    @property
    def ok(self) -> bool:
        return self.status == "success"


def scan_workbook_errors(
    path: Path | str,
    *,
    wb_vals: openpyxl.Workbook | None = None,
) -> WorkbookValidation:
    """Scan a workbook for cached Excel error tokens.

    Two passes over the file:
    1. data_only=True — read cached values, look for error tokens.
    2. data_only=False — count formula cells (denominator for triage).

    Returns a WorkbookValidation. Never raises on per-cell decode issues;
    a top-level load failure produces status="scan_failed" with error msg.

    `wb_vals` is an optional pre-loaded value-pass workbook handle. When
    provided, this function skips its own openpyxl.load_workbook() call and
    iterates the caller's handle instead — used by run_preflight to share
    one workbook load across scan_workbook_errors + check_critical_path_-
    errors (~3-5 min saved on a 13MB pricing model). Caller owns the
    handle's lifecycle; this function never closes a handle it didn't open.
    """
    p = Path(path)
    if not p.exists():
        return WorkbookValidation(
            status="scan_failed",
            total_errors=0,
            total_formulas=0,
            error_summary={},
            error=f"File not found: {p}",
        )

    error_locations: dict[str, list[str]] = {tok: [] for tok in EXCEL_ERROR_TOKENS}
    error_counts: dict[str, int] = {tok: 0 for tok in EXCEL_ERROR_TOKENS}
    total_errors = 0

    close_vals = False
    if wb_vals is None:
        try:
            wb_vals = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
            close_vals = True
        except Exception as exc:
            return WorkbookValidation(
                status="scan_failed",
                total_errors=0,
                total_formulas=0,
                error_summary={},
                error=f"Failed to load (values pass): {exc}",
            )

    try:
        for sheet_name in wb_vals.sheetnames:
            ws = wb_vals[sheet_name]
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    val = cell.value
                    if val is None or not isinstance(val, str):
                        continue
                    for tok in EXCEL_ERROR_TOKENS:
                        if tok in val:
                            error_counts[tok] += 1
                            if len(error_locations[tok]) < _MAX_LOCATIONS_PER_TYPE:
                                error_locations[tok].append(
                                    f"{sheet_name}!{cell.coordinate}"
                                )
                            total_errors += 1
                            break
    finally:
        if close_vals:
            wb_vals.close()

    # Formula count — only run when errors were found, since it's an
    # expensive second openpyxl pass (3-5 min on a 13MB xlsm) used only
    # as the triage denominator in the report. On clean workbooks (the
    # common case) skipping this saves ~half the scan time.
    formula_count = 0
    if total_errors > 0:
        try:
            wb_formulas = openpyxl.load_workbook(str(p), data_only=False, read_only=True)
            try:
                for sheet_name in wb_formulas.sheetnames:
                    ws = wb_formulas[sheet_name]
                    for row in ws.iter_rows(values_only=True):
                        for val in row:
                            if isinstance(val, str) and val.startswith("="):
                                formula_count += 1
            finally:
                wb_formulas.close()
        except Exception as exc:
            log.debug("Formula count pass failed (non-fatal): %s", exc)

    summary = {
        tok: ErrorTypeDetail(
            count=error_counts[tok],
            locations=tuple(error_locations[tok]),
        )
        for tok in EXCEL_ERROR_TOKENS
        if error_counts[tok] > 0
    }

    status = "success" if total_errors == 0 else "errors_found"
    return WorkbookValidation(
        status=status,
        total_errors=total_errors,
        total_formulas=formula_count,
        error_summary=summary,
    )


def format_validation_report(v: WorkbookValidation, label: str = "Workbook") -> str:
    """Render a one-screen text report for log output."""
    if v.status == "scan_failed":
        return f"{label}: scan failed — {v.error}"
    if v.ok:
        return (
            f"{label}: clean ({v.total_formulas} formula(s), 0 errors)"
        )
    parts = [
        f"{label}: {v.total_errors} formula error(s) "
        f"across {len(v.error_summary)} type(s) "
        f"(of {v.total_formulas} formulas total)"
    ]
    for tok, detail in v.error_summary.items():
        head = ", ".join(detail.locations[:5])
        more = (
            f" (+{detail.count - 5} more)"
            if detail.count > 5
            else ""
        )
        parts.append(f"  {tok} x{detail.count}: {head}{more}")
    return "\n".join(parts)
