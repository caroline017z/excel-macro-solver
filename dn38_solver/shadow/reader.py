"""dn38_solver.shadow.reader — Workbook reader via openpyxl (no COM).

data_only=True reads cached formula values. read_only=False enables
fast random cell access (~3s vs ~456s with read_only=True on 8.5MB xlsm).
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Self

import openpyxl

from dn38_solver.types import ProjectInfo
from dn38_solver.config import (
    BASE_COL,
    OUTPUT_ROWS,
    PROJECT_COL_RANGE,
    PROJECT_NAME_ROW,
    PROJECT_TOGGLE_ROW,
    SNAPSHOT_RANGES,
)
from dn38_solver.convert import col_letter, safe_float, safe_value

log = logging.getLogger(__name__)

_TOGGLE_ON_VALUES = frozenset({"1", "on", "true"})


class WorkbookReader:
    """Openpyxl-based workbook reader. Pure Python, zero COM dependency."""

    __slots__ = ("path", "_wb")

    def __init__(self, path: Path) -> None:
        self.path = path
        self._wb = openpyxl.load_workbook(
            str(path), data_only=True, read_only=False,
        )
        log.info(
            "Opened workbook: %s (%d sheets)",
            path.name,
            len(self._wb.sheetnames),
        )

    def close(self) -> None:
        self._wb.close()

    def __enter__(self) -> Self:
        return self

    def __exit__(self, *_: object) -> None:
        self.close()

    # --- Single cell reads ---

    def cell_value(self, sheet: str, row: int, col: int) -> float | str | None:
        """Read a single cell as float (preferred) or str."""
        return safe_value(self._wb[sheet].cell(row=row, column=col).value)

    def cell_float(self, sheet: str, row: int, col: int) -> float | None:
        """Read a single cell as float only. None if not numeric."""
        return safe_float(self._wb[sheet].cell(row=row, column=col).value)

    # --- Project discovery ---

    def extract_active_projects(self) -> tuple[ProjectInfo, ...]:
        """THE one canonical project extraction function.

        Scans Project Inputs row 7 for toggled-on projects.
        Returns frozen tuple of ProjectInfo structs.
        """
        ws = self._wb["Project Inputs"]
        projects: list[ProjectInfo] = []

        for c in PROJECT_COL_RANGE:
            raw_name = ws.cell(row=PROJECT_NAME_ROW, column=c).value
            if not raw_name or not str(raw_name).strip():
                continue

            toggle_val = ws.cell(row=PROJECT_TOGGLE_ROW, column=c).value
            is_on = (
                str(toggle_val).strip().lower() in _TOGGLE_ON_VALUES
                if toggle_val is not None
                else False
            )
            if not is_on:
                continue

            clean_name = " | ".join(
                line.strip()
                for line in str(raw_name).strip().splitlines()
                if line.strip()
            )
            projects.append(
                ProjectInfo(
                    name=clean_name,
                    col=c,
                    col_letter=col_letter(c),
                    offset=c - BASE_COL,
                    toggle=True,
                )
            )

        result = tuple(projects)
        log.info("Found %d active projects", len(result))
        return result

    # --- Output extraction ---

    def read_output_rows(self, col: int) -> dict[int, float | None]:
        """Read all OUTPUT_ROWS for a given project column."""
        ws = self._wb["Project Inputs"]
        return {
            row: safe_float(ws.cell(row=row, column=col).value)
            for row in OUTPUT_ROWS
        }

    # --- Snapshot for diff ---

    def snapshot(self) -> dict[tuple[str, int, int], float | str | None]:
        """Capture a snapshot of all configured snapshot ranges.

        Returns {(sheet, row, col): value} for non-None cells.
        Used for pre/post diff comparison.
        """
        result: dict[tuple[str, int, int], float | str | None] = {}

        for sheet_name, (min_r, max_r, min_c, max_c) in SNAPSHOT_RANGES.items():
            if sheet_name not in self._wb.sheetnames:
                log.warning("Snapshot sheet %s not found, skipping", sheet_name)
                continue
            ws = self._wb[sheet_name]
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    val = safe_value(ws.cell(row=r, column=c).value)
                    if val is not None:
                        result[(sheet_name, r, c)] = val

        # Also snapshot output rows for all active projects
        ws_pi = self._wb["Project Inputs"]
        for col in PROJECT_COL_RANGE:
            raw_name = ws_pi.cell(row=PROJECT_NAME_ROW, column=col).value
            if not raw_name or not str(raw_name).strip():
                continue
            for row in OUTPUT_ROWS:
                val = safe_value(ws_pi.cell(row=row, column=col).value)
                if val is not None:
                    result[("Project Inputs", row, col)] = val

        log.info("Snapshot captured: %d cells", len(result))
        return result
