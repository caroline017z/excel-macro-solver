"""dn38_solver.shadow.preflight — Bank-grade pre-flight validation.

A single workbook check that fails LOUD before paying COM startup +
multi-minute solve cost on a workbook that won't converge.

Three categories of checks, each with a stable error code:

  A. Workbook calc-property failures. These weaken or disable the
     iterative-calc engine that the macro's row 31 self-circular hard-
     stamp pattern relies on. Failures are subtle — convergence may
     still appear to work but with non-deterministic last-wins values.
       A1  iterateDelta missing or > 0.0001
       A2  iterate=False
       A3  calcMode=auto
       A4  fullCalcOnLoad=False
  B. Workbook structure failures (macro crashes mid-run with cryptic
     VBA errors when sheets/cells/protection don't match expectations).
       B5  Required sheets missing
       B7  Required cells missing or wrong type
       B8  Workbook or sheet structurally protected
       B9  No active project flagged in row 7
  C. Critical-path cached-error checks (errors on cells the macro reads
     during convergence — XIRR returns garbage when feeder rows have
     #VALUE!/#DIV/0!/#REF! tokens).
       C10 Errors on Project Inputs F30:F39 master column
       C11 Errors on Appraisal cash flow rows 155-159
       C12 Errors on Appraisal!H161 (the XIRR readout)
  D. Embedded VBA macro version. The repo's SolveHeadless.bas evolves
     (yesterday alone added phase-scoped recalc, hard-stamp post-read pass,
     parallel-runner trust gates). Workbooks where the macro hasn't
     been re-imported run an OUTDATED macro version missing critical
     functions; the orchestrator calls them via Application.Run and
     they silently fail (or fall through On Error Resume Next),
     leaving the solve in a half-broken state.
       D15 Embedded macro is missing required functions
       D16 Embedded macro has leftover stale modules
  E. Input-range checks against the macro's hardcoded bounds. The chunked
     entry point (SolveOneProjectByColHL) resets pre-solve Dev Fee / NPP
     to seed values when they fall outside [DEV_FEE_MIN..MAX]/[NPP_MIN..
     MAX]. For models where the natural converged Dev Fee is well above
     DEV_FEE_MAX (e.g., utility solar where Dev Fee can be $1.50-$2.50/W),
     this reset destroys the starting state and GoalSeek may not recover.
       E13 Pre-solve Dev Fee outside macro bounds
       E14 Pre-solve NPP outside macro bounds

Findings carry a remediation string and an auto_fixable flag. Today
only A1 (iterateDelta) is auto-fixable; A2/E13/E14 could be added but
mutating model inputs / iterate flag without operator audit is risky.

Bank-grade defaults: errors always abort the run. Warnings are visible
in the log but proceed unless --strict-preflight is set.
"""
from __future__ import annotations

import hashlib
import logging
import re
import shutil
import zipfile
from collections.abc import Iterator
from pathlib import Path

import msgspec
import openpyxl
from openpyxl.utils import column_index_from_string

from dn38_solver.shadow.validation import (
    EXCEL_ERROR_TOKENS,
    WorkbookValidation,
    scan_workbook_errors,
)

log = logging.getLogger(__name__)

# Tolerance ceiling for the iterative-calc engine. Below this value
# GoalSeek's Appraisal inner loop converges reliably. Above it (Excel's
# default 0.001) the loop bisects on a half-converged H161 and slides
# Dev Fee to the floor.
ITERATE_DELTA_CEILING = 0.0001

# Sheets the macro requires. Missing any of these crashes SolveHeadless or
# the chunked entry points with a "Subscript out of range" — surface that
# upfront with a named sheet rather than letting the macro die mid-run.
REQUIRED_SHEETS = (
    "Project Inputs",
    "Appraisal",
    "NPP Calc",
    "Operations",
    "PT Returns",
    "Tax Equity",
    "Perm Debt",
    "CL",
    "Capex",
    "Rate Curves",
    "Global",
)

# Master-column cells the macro reads/writes during the solve. F2 is the
# active-project index, F30/F36 are targets (WACC / Equity), F31/F37 are
# live readouts the GoalSeek operates on, F32 is the changing Dev Fee.
REQUIRED_PI_CELLS = {
    "F2": "active project index",
    "F30": "FMV WACC target",
    "F31": "Live Appraisal IRR (GoalSeek target var)",
    "F32": "Dev Fee (GoalSeek changing cell)",
    "F36": "Equity IRR target",
    "F37": "Live Levered Pre-Tax IRR",
}

# Ranges the macro reads or that feed the GoalSeek convergence variable.
# Critical-path errors here corrupt the inner loop irrespective of model
# inputs — distinct from the broad workbook-wide scan in validation.py.
PI_CRITICAL_PATH_ROWS = (30, 31, 32, 33, 36, 37, 38, 39)
APPRAISAL_CASHFLOW_ROWS = (155, 156, 157, 158, 159)

# Hardcoded bounds in SolveHeadless.bas (lines 46-51). Mirror them here so
# pre-flight can warn before the macro's chunked path resets out-of-range
# values to seed values. If the .bas constants change, update these too —
# there's no clean import path because the macro is not Python.
NPP_BOUNDS = (-0.2, 2.0)         # NPP_MIN, NPP_MAX — bounds widened for utility-scale ranges
DEV_FEE_BOUNDS = (0.05, 5.5)     # DEV_FEE_MIN, DEV_FEE_MAX — bounds widened to cover concentrated-portfolio Dev Fees
PI_ROW_NPP = 38
PI_ROW_DEV_FEE = 32
PI_ROW_TOGGLE = 7
# H..BG = 8..59, mirrors SolveHeadless COL_SCAN_LIMIT=60. The range
# must reach the macro's scan limit; a shorter range would silently
# miss projects in late columns. Inactive columns past the last
# toggled-on project are no-ops (row 7 != 1).
PI_PROJECT_COL_RANGE = ("H", "BG")
APPRAISAL_CASHFLOW_MAX_COL = "UC"  # right-edge bound for the cashflow error scan

# Rate Component sub-block layout in Project Inputs. Six RCs, one block
# each, at rows 157/167/177/187/197/207. Within each block:
# +1 = Rate Name, +2 = Custom/Generic toggle, +5 = Term. The E15
# check uses Rate Name and Toggle to detect Toggle="Custom" with empty
# Rate Name — leaves the revenue stream unfunded, which the model
# compensates for with an inflated Dev Fee at solve time.
RC_BLOCK_BASES = (157, 167, 177, 187, 197, 207)  # RC1..RC6
RC_OFFSET_NAME = 1
RC_OFFSET_TOGGLE = 2
RC_OFFSET_GEN_RATE = 3

# Equity master on/off toggles for each RC. These sit ABOVE the per-RC
# sub-blocks under the "Revenue Rate Components for Equity Model" header
# (row 149). When the master toggle is 0, the RC contributes zero
# revenue regardless of the per-block Custom/Generic config below — so
# E15a/b/c must skip those (col, rc) pairs to avoid false positives on
# the common pattern of a workbook with RC5/RC6 master-off but cosmetic
# values still in the sub-blocks (typical IL CS deal config — RC5/RC6
# master-off for all projects but rows 197-215 carry leftover template
# data). Caught on the 2026-05-18 SolarStone next-wave run after the
# E15c block produced an actionable false positive.
RC_MASTER_TOGGLE_ROWS = (150, 151, 152, 153, 154, 155)  # RC1..RC6, Equity

# Marker function names that must exist in the embedded macro for the
# current orchestrator to drive the workbook correctly. Absence implies
# the workbook hasn't had the latest macro re-imported. Scanned as ASCII
# tokens in vbaProject.bin — VBA's compressed-storage format leaves
# identifier strings readable. ~10ms on the 215KB binary, no Excel COM
# needed at preflight time.
REQUIRED_MACRO_FUNCTIONS = (
    "SolveHeadless",                # entry point (single-shot)
    "InitSolveEnvHL",               # entry point (chunked init)
    "SolveOneProjectByColHL",       # entry point (chunked per-project)
    "FinalizeSolveEnvHL",           # entry point (chunked finalize)
    "CalcSheetsForAppraisal",       # phase-scoped recalc
    "CalcSheetsForNPP",             # phase-scoped recalc
    "CalcSheetsForDSCR",            # phase-scoped recalc
    "ClassifyConvergenceHL",        # strict/relaxed/none tier classifier
    "StampActiveProjectColumnHL",   # post-read hard-stamps
    "ProjectElapsedHL",             # timer wraparound-safe elapsed
    "HardStampNumericHL",           # IsError-guarded numeric stamping
    "SetSkipOutputRecalcHL",        # output-recalc skip flag
)

# Names of stale module artifacts that linger in workbooks where macros
# were imported then partially replaced. Each is a hint that the workbook
# has been through several import cycles without the cleanup pass that
# removes the prior versions. Not a hard blocker but worth surfacing.
STALE_MACRO_MODULES = (
    "Module2_Optimized",
    "Module3",
    "Module4",
)

# Custom-doc-property name and source .bas path for the D17 hash drift
# check. import_vba_module.py stamps the SHA256 of SolveHeadless.bas into
# this property on every successful import; check_macro_hash re-reads it
# and compares against the current repo .bas. Mismatch = re-import is
# required. Stronger guard than the function-presence scan in D15: a .bas
# can carry all required function NAMES but have updated bodies that the
# orchestrator depends on (e.g., changed loop bounds, new heartbeat
# emits). Hash equality is the only reliable drift signal.
BAS_HASH_PROP = "DN38_BAS_SHA256"
_REPO_BAS_PATH = Path(__file__).parent.parent.parent / "SolveHeadless.bas"
_BAS_HASH_RE = re.compile(
    r'<property[^>]+name="' + BAS_HASH_PROP + r'"[^>]*>\s*<vt:[^>]+>([0-9a-fA-F]+)</vt:',
    re.IGNORECASE,
)


class PreflightFinding(msgspec.Struct, frozen=True, kw_only=True):
    """One actionable finding from the pre-flight pass."""
    code: str
    severity: str  # "error" | "warning" | "info"
    location: str
    message: str
    impact: str
    remediation: str
    auto_fixable: bool = False


class PreflightResult(msgspec.Struct, frozen=True, kw_only=True):
    """Aggregate result of a workbook pre-flight pass."""
    workbook_path: str
    findings: tuple[PreflightFinding, ...]
    error_scan: WorkbookValidation

    @property
    def errors(self) -> tuple[PreflightFinding, ...]:
        return tuple(f for f in self.findings if f.severity == "error")

    @property
    def warnings(self) -> tuple[PreflightFinding, ...]:
        return tuple(f for f in self.findings if f.severity == "warning")

    @property
    def auto_fixable(self) -> tuple[PreflightFinding, ...]:
        return tuple(f for f in self.findings if f.auto_fixable)

    @property
    def ok(self) -> bool:
        """No error-severity findings."""
        return len(self.errors) == 0


# ---------------------------------------------------------------------------
# Individual check functions. Each returns a list (possibly empty) of
# PreflightFindings. Kept as pure functions over an open openpyxl workbook
# so they're trivially testable with synthetic in-memory workbooks.
# ---------------------------------------------------------------------------

def iter_active_project_cols(ws) -> Iterator[int]:
    """Yield 1-based column indices for active projects on `ws`
    (Project Inputs). Active = row 7 (PI_ROW_TOGGLE) == 1.
    Shared by every check that iterates project columns.
    """
    start_col = column_index_from_string(PI_PROJECT_COL_RANGE[0])
    end_col = column_index_from_string(PI_PROJECT_COL_RANGE[1])
    for col in range(start_col, end_col + 1):
        if ws.cell(row=PI_ROW_TOGGLE, column=col).value == 1:
            yield col


def check_calc_properties(wb: openpyxl.Workbook) -> list[PreflightFinding]:
    """A1-A4: workbook calculation properties."""
    findings: list[PreflightFinding] = []
    cp = wb.calculation

    # A1: iterateDelta. Absence (or a value above the ceiling) weakens
    # iterative-calc convergence on the row 31 self-circular cells the
    # macro depends on. Healthy workbooks set this explicitly to 1E-4;
    # absence is anomalous and worth flagging loud.
    if cp.iterateDelta is None or cp.iterateDelta > ITERATE_DELTA_CEILING:
        findings.append(PreflightFinding(
            code="A1",
            severity="error",
            location="workbook calcPr",
            message=(
                f"iterateDelta is {cp.iterateDelta!r} (must be <= {ITERATE_DELTA_CEILING})"
            ),
            impact=(
                "Iterative-calc engine may exit before circular references "
                "settle, causing the macro's row 31 self-circular per-project "
                "hard-stamps to capture stale values. All baseline pricing "
                "models on disk have this set explicitly to 1E-4; missing "
                "value indicates non-standard save state."
            ),
            remediation=(
                "In Excel: File > Options > Formulas > Maximum Change = 0.0001. "
                "Save. Or rerun with --auto-fix to patch a copy."
            ),
            auto_fixable=True,
        ))

    # A2: iterative calc must be enabled. Row 31 self-circular cells
    # (=IF(H2=$F$2,$F$31,H31)) become #REF! when iterate=False.
    if not cp.iterate:
        findings.append(PreflightFinding(
            code="A2",
            severity="error",
            location="workbook calcPr",
            message="iterative calculation is disabled (iterate=False)",
            impact=(
                "Project Inputs row 31 self-circular formulas error out, "
                "preventing per-column appraisal IRR capture. The post-solve "
                "hard-stamps will read #REF!."
            ),
            remediation=(
                "In Excel: File > Options > Formulas > "
                "Enable iterative calculation (checked). Save."
            ),
            auto_fixable=False,  # Doable but risky to flip without model audit
        ))

    # A3: calcMode auto causes uncontrolled recalc during macro runs.
    if cp.calcMode and cp.calcMode != "manual":
        findings.append(PreflightFinding(
            code="A3",
            severity="warning",
            location="workbook calcPr",
            message=f"calcMode is {cp.calcMode!r} (expected 'manual')",
            impact=(
                "Excel may trigger background recalcs during the macro, racing "
                "with the macro's own targeted .Calculate calls. Convergence "
                "behavior becomes nondeterministic."
            ),
            remediation=(
                "In Excel: Formulas tab > Calculation Options > Manual. Save."
            ),
            auto_fixable=False,
        ))

    # A4: fullCalcOnLoad ensures the cached values seen by openpyxl scans
    # actually reflect the formulas' current state.
    if cp.fullCalcOnLoad is False:  # explicit False, not None
        findings.append(PreflightFinding(
            code="A4",
            severity="warning",
            location="workbook calcPr",
            message="fullCalcOnLoad is False",
            impact=(
                "Excel won't recalculate on open, so cached values may not "
                "reflect current formulas. Pre-flight error scans (#VALUE!, "
                "#DIV/0!, etc.) become unreliable."
            ),
            remediation=(
                "Open in Excel, press F9 to force full recalc, save. "
                "Alternative: have the model owner enable Workbook > "
                "Calculation > 'Recalculate workbook before saving'."
            ),
            auto_fixable=False,
        ))

    return findings


def check_required_sheets(wb: openpyxl.Workbook) -> list[PreflightFinding]:
    """B5: required sheets must be present."""
    present = set(wb.sheetnames)
    missing = [s for s in REQUIRED_SHEETS if s not in present]
    if not missing:
        return []
    return [PreflightFinding(
        code="B5",
        severity="error",
        location="workbook",
        message=f"missing required sheet(s): {', '.join(missing)}",
        impact=(
            "Macro will fail with VBA 'Subscript out of range' when it "
            "calls Sheets(\"<missing>\")."
        ),
        remediation=(
            f"Restore the missing sheet(s) from a baseline pricing model. "
            f"Required: {', '.join(REQUIRED_SHEETS)}."
        ),
        auto_fixable=False,
    )]


def check_required_cells(wb: openpyxl.Workbook) -> list[PreflightFinding]:
    """B7: master-column cells the macro depends on must exist."""
    if "Project Inputs" not in wb.sheetnames:
        return []  # B5 already flagged this
    ws = wb["Project Inputs"]
    findings: list[PreflightFinding] = []
    for cell, role in REQUIRED_PI_CELLS.items():
        v = ws[cell].value
        if v is None:
            findings.append(PreflightFinding(
                code="B7",
                severity="error",
                location=f"Project Inputs!{cell}",
                message=f"required cell is empty (role: {role})",
                impact=(
                    "Macro reads or writes this cell every iteration. "
                    "Empty value will produce wrong results or VBA error."
                ),
                remediation=(
                    f"Restore Project Inputs!{cell} from a baseline model. "
                    f"This cell holds the {role}."
                ),
                auto_fixable=False,
            ))
    return findings


def check_workbook_protection(wb: openpyxl.Workbook) -> list[PreflightFinding]:
    """B8: workbook and critical-sheet protection."""
    findings: list[PreflightFinding] = []
    # Workbook structure protection (prevents adding/renaming sheets — the
    # macro adds __SolverResults).
    try:
        wb_protected = wb.security and (
            getattr(wb.security, "lockStructure", False)
            or getattr(wb.security, "lockWindows", False)
        )
    except Exception:
        wb_protected = False
    if wb_protected:
        findings.append(PreflightFinding(
            code="B8",
            severity="error",
            location="workbook",
            message="workbook structure is protected",
            impact=(
                "Macro will fail when adding the __SolverResults telemetry "
                "sheet at startup. Protection blocks all sheet add/rename."
            ),
            remediation=(
                "In Excel: Review tab > Protect Workbook (toggle off). "
                "Provide password if one was set."
            ),
            auto_fixable=False,
        ))

    # Per-sheet protection on critical sheets the macro writes to.
    for sname in ("Project Inputs", "Appraisal", "NPP Calc"):
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        if ws.protection.sheet:
            findings.append(PreflightFinding(
                code="B8",
                severity="error",
                location=sname,
                message=f"sheet '{sname}' is protected",
                impact=(
                    f"Macro writes to '{sname}' during solve (project index, "
                    f"hard-stamps, calc triggers). Protection blocks writes."
                ),
                remediation=(
                    f"In Excel: select sheet '{sname}' > Review tab > "
                    f"Unprotect Sheet."
                ),
                auto_fixable=False,
            ))
    return findings


def check_active_projects(wb: openpyxl.Workbook) -> list[PreflightFinding]:
    """B9: at least one project must be flagged active in row 7."""
    if "Project Inputs" not in wb.sheetnames:
        return []  # B5 covered this
    ws = wb["Project Inputs"]
    # Project columns are H..S (8..19). Row 7 is the active flag (1=on).
    active_cols = []
    for col in range(column_index_from_string("H"), column_index_from_string("S") + 1):
        v = ws.cell(row=7, column=col).value
        if v == 1:
            active_cols.append(col)
    if not active_cols:
        return [PreflightFinding(
            code="B9",
            severity="error",
            location="Project Inputs!H7:S7",
            message="no active projects flagged (row 7 = 1 in any project column)",
            impact="Solver will exit immediately with 'No active projects found'.",
            remediation=(
                "Set row 7 to 1 for the project columns you want to solve. "
                "Inactive projects must be 0 or blank."
            ),
            auto_fixable=False,
        )]
    return []


def check_critical_path_errors(
    wb_values: openpyxl.Workbook,
    error_scan: WorkbookValidation,
) -> list[PreflightFinding]:
    """C10-C12: cached error tokens on cells the convergence depends on.

    Uses the data_only=True workbook to read cached error tokens. Falls
    back to the full error scan's reported locations to detect criticality
    in O(broken-cell-count) rather than re-scanning the whole workbook.
    """
    findings: list[PreflightFinding] = []

    def is_error(val: object) -> bool:
        return isinstance(val, str) and any(tok in val for tok in EXCEL_ERROR_TOKENS)

    # C10: PI master column (F30:F39) and active project columns (H..S)
    if "Project Inputs" in wb_values.sheetnames:
        ws = wb_values["Project Inputs"]
        bad_cells: list[str] = []
        for row in PI_CRITICAL_PATH_ROWS:
            for col in range(6, column_index_from_string("S") + 1):  # F..S
                v = ws.cell(row=row, column=col).value
                if is_error(v):
                    bad_cells.append(f"Project Inputs!{ws.cell(row=row, column=col).coordinate}={v}")
                    if len(bad_cells) >= 10:
                        break
            if len(bad_cells) >= 10:
                break
        if bad_cells:
            findings.append(PreflightFinding(
                code="C10",
                severity="error",
                location="Project Inputs F30:S39",
                message=(
                    f"{len(bad_cells)} cached error(s) on convergence-critical cells "
                    f"(showing first {min(10, len(bad_cells))}: {'; '.join(bad_cells[:5])}"
                    f"{'; ...' if len(bad_cells) > 5 else ''})"
                ),
                impact=(
                    "Macro reads these rows for the GoalSeek target/changing "
                    "cells and per-project hard-stamps. Errors here corrupt "
                    "the entire convergence loop."
                ),
                remediation=(
                    "Open the workbook in Excel, navigate to the listed cells, "
                    "and trace the precedent chain to find the source of the "
                    "error. Common causes: missing rate curves, broken named "
                    "ranges, deleted source rows."
                ),
                auto_fixable=False,
            ))

    # C11/C12: Appraisal cash flow rows + H161 readout
    if "Appraisal" in wb_values.sheetnames:
        ws = wb_values["Appraisal"]
        bad_cf: list[str] = []
        # Active cash-flow region only — the OFFSET picks columns to the
        # right of where the CT comes online.
        max_col = min(ws.max_column, column_index_from_string(APPRAISAL_CASHFLOW_MAX_COL))
        # Use iter_rows for bulk scanning, NOT ws.cell() in a loop. The
        # caller's wb_values handle was opened with read_only=True for
        # memory efficiency; in that mode each random ws.cell(row, col)
        # call re-streams the sheet part from the start, making the
        # nested loop O(rows × cols × cols) instead of O(rows × cols).
        # On a workbook with the Appraisal cashflow spanning ~550 cols
        # (post-many-projects-solved state), this turned a sub-second
        # check into a multi-minute hang that timed out pre-flight
        # entirely. (Caught 2026-05-18 on the 25-project SolarStone
        # re-run — the same workbook ran cleanly through the same check
        # earlier in the day when fewer columns were populated.)
        min_col = column_index_from_string("J")
        appr_row_set = set(APPRAISAL_CASHFLOW_ROWS)
        # iter_rows is bulk-streamed and respects min_row/max_row, so
        # we get one efficient pass over just the cashflow rows.
        first_row = min(APPRAISAL_CASHFLOW_ROWS)
        last_row = max(APPRAISAL_CASHFLOW_ROWS)
        for row_idx, row_vals in enumerate(
            ws.iter_rows(min_row=first_row, max_row=last_row,
                         min_col=1, max_col=max_col, values_only=True),
            start=first_row,
        ):
            if row_idx not in appr_row_set:
                continue
            for col_offset, v in enumerate(row_vals, start=1):
                if col_offset < min_col:
                    continue
                if is_error(v):
                    coord = ws.cell(row=row_idx, column=col_offset).coordinate
                    bad_cf.append(f"Appraisal!{coord}={v}")
                    if len(bad_cf) >= 10:
                        break
            if len(bad_cf) >= 10:
                break
        if bad_cf:
            findings.append(PreflightFinding(
                code="C11",
                severity="error",
                location="Appraisal rows 155-159",
                message=(
                    f"{len(bad_cf)} cached error(s) on Appraisal cash flow rows "
                    f"({'; '.join(bad_cf[:5])}{'; ...' if len(bad_cf) > 5 else ''})"
                ),
                impact=(
                    "Cash flow row 159 = SUM(155:158) feeds the XIRR readout "
                    "at H161. Errors here propagate to H161 and corrupt the "
                    "GoalSeek target value."
                ),
                remediation=(
                    "Trace precedents on the listed cells. Often caused by "
                    "missing inputs in Project Inputs row 32 (Dev Fee), row 11 "
                    "(System Size), or rate curve gaps in Rate Curves."
                ),
                auto_fixable=False,
            ))

        h161 = ws["H161"].value
        if is_error(h161):
            findings.append(PreflightFinding(
                code="C12",
                severity="error",
                location="Appraisal!H161",
                message=f"H161 (Live Appraisal IRR XIRR readout) is {h161}",
                impact=(
                    "F31 = Appraisal!H161 is the GoalSeek target variable. "
                    "An error here means GoalSeek has no valid value to drive "
                    "and will fail immediately."
                ),
                remediation=(
                    "Fix the upstream errors first (likely C11 findings) and "
                    "force a full recalc (F9) before re-running."
                ),
                auto_fixable=False,
            ))

    return findings


def check_macro_version(workbook_path: Path) -> list[PreflightFinding]:
    """D15/D16: embedded macro must have all required functions and no
    stale leftover modules.

    Scans xl/vbaProject.bin for ASCII function-name tokens. VBA's
    compressed-storage format leaves identifier strings readable as
    plain ASCII; a substring match for `\\x00<name>` (the typical
    surrounding bytes) is more precise than a bare substring search,
    but a bare search is good enough for our marker functions which
    don't appear elsewhere in the binary.
    """
    findings: list[PreflightFinding] = []
    # .xlsx is macro-free by spec; skip D-tier entirely. Tests use .xlsx
    # synthetic workbooks, and the solver targets .xlsm only — checking
    # for a missing macro on .xlsx would be noise.
    if workbook_path.suffix.lower() != ".xlsm":
        return findings

    try:
        with zipfile.ZipFile(workbook_path, "r") as z:
            try:
                vba_bin = z.read("xl/vbaProject.bin")
            except KeyError:
                # No VBA project at all in an .xlsm. Solver can't drive a
                # macro-less workbook; flag as a structural error.
                return [PreflightFinding(
                    code="D15",
                    severity="error",
                    location="xl/vbaProject.bin",
                    message="workbook contains no VBA project",
                    impact=(
                        "Solver requires the SolveHeadless macro to be "
                        "embedded. Without it, no entry point exists for "
                        "Application.Run."
                    ),
                    remediation=(
                        "Run: python import_vba_module.py "
                        f'"{workbook_path}"'
                    ),
                    auto_fixable=False,
                )]
    except (zipfile.BadZipFile, OSError) as exc:
        # Defer to scan_workbook_errors's X0 handling; don't double-report.
        log.debug("D-tier scan: cannot open zip (%s): %s", workbook_path, exc)
        return findings

    # Scan as latin-1 so every byte maps to a 1-char string (identifier
    # tokens are ASCII either way; latin-1 is just the safe decode).
    text = vba_bin.decode("latin-1", errors="replace")

    missing = [name for name in REQUIRED_MACRO_FUNCTIONS if name not in text]
    if missing:
        findings.append(PreflightFinding(
            code="D15",
            severity="error",
            location="xl/vbaProject.bin",
            message=(
                f"embedded macro is missing {len(missing)} of "
                f"{len(REQUIRED_MACRO_FUNCTIONS)} required function(s): "
                f"{', '.join(missing)}"
            ),
            impact=(
                "The orchestrator calls these functions via Application.Run "
                "during the chunked solve path. Missing functions either "
                "fail silently (On Error Resume Next) or raise a VBA error, "
                "leaving the solve in a half-broken state."
            ),
            remediation=(
                f'Re-import the latest macro: python import_vba_module.py '
                f'"{workbook_path}"'
            ),
            # Per Tranche 7.13: --auto-fix can recover from this by
            # re-importing the macro into the _FIXED.xlsm sibling
            # (orchestrator handles the COM call; preflight stays pure).
            # The original workbook is never mutated.
            auto_fixable=True,
        ))

    stale_present = [m for m in STALE_MACRO_MODULES if m in text]
    if stale_present:
        findings.append(PreflightFinding(
            code="D16",
            severity="warning",
            location="xl/vbaProject.bin",
            message=(
                f"workbook contains {len(stale_present)} stale macro "
                f"module(s): {', '.join(stale_present)}"
            ),
            impact=(
                "Indicates the workbook has been through several macro-"
                "import cycles without cleanup. Stale modules don't break "
                "the solve directly but may shadow current function names "
                "or contain references to obsolete cells. Worth removing."
            ),
            remediation=(
                "In Excel: Alt+F11 to open VBA editor; right-click each "
                "stale module under VBAProject and choose 'Remove'. "
                "Decline the export prompt unless you want a backup."
            ),
            auto_fixable=False,
        ))

    return findings


def _current_bas_sha256() -> str | None:
    """Return SHA256 of the repo's current SolveHeadless.bas, or None if
    the file is missing (developer working tree only — preflight should
    skip the drift check rather than fail loudly)."""
    if not _REPO_BAS_PATH.exists():
        return None
    return hashlib.sha256(_REPO_BAS_PATH.read_bytes()).hexdigest()


def _read_stamped_bas_hash(workbook_path: Path) -> str | None:
    """Pull DN38_BAS_SHA256 out of docProps/custom.xml via zip-level read.

    Avoids the openpyxl save round-trip (which would violate the openpyxl-
    xlsm save rule even on a read-then-close pattern, since openpyxl's
    `read_only=True` skips custom properties entirely and a normal load
    can mutate workbook state on close). Returns None when the property
    is absent or the file has no custom.xml at all.
    """
    try:
        with zipfile.ZipFile(workbook_path, "r") as z:
            try:
                xml = z.read("docProps/custom.xml").decode("utf-8")
            except KeyError:
                return None
    except (zipfile.BadZipFile, OSError):
        return None
    m = _BAS_HASH_RE.search(xml)
    return m.group(1) if m else None


def check_macro_hash(workbook_path: Path) -> list[PreflightFinding]:
    """D17: embedded macro hash must match the current SolveHeadless.bas.

    Compares the SHA256 stamped by import_vba_module.py against a fresh
    hash of the repo's .bas file. Cheap (~5ms) and catches drift that the
    function-name scan in D15 misses — most .bas updates preserve marker
    function names while changing implementation, so D15 would pass while
    a critical body change goes undetected.

    Semantics:
      - .bas missing in repo:  skip silently (orchestrator can still run)
      - .xlsx workbook:        skip (.xlsx has no macros by spec)
      - no stamp on workbook:  warning (workbook predates the stamp
                               convention; can't verify either way)
      - stamp ≠ current:       error (definite drift; re-import required)
    """
    findings: list[PreflightFinding] = []
    if workbook_path.suffix.lower() != ".xlsm":
        return findings
    current = _current_bas_sha256()
    if current is None:
        # Repo working tree without the source .bas. Don't blame the
        # workbook for a developer-environment issue.
        return findings
    stamped = _read_stamped_bas_hash(workbook_path)
    if stamped is None:
        findings.append(PreflightFinding(
            code="D17",
            severity="warning",
            location="docProps/custom.xml",
            message=(
                f"workbook has no {BAS_HASH_PROP} stamp — cannot verify "
                f"the embedded macro matches the current SolveHeadless.bas"
            ),
            impact=(
                "Without a stamp, drift between the repo .bas and the "
                "embedded macro is invisible. Macro drift is the #1 cause "
                "of cryptic solve failures."
            ),
            remediation=(
                f"Re-import the macro once to plant the stamp: "
                f'python import_vba_module.py "{workbook_path}"'
            ),
            # Per Tranche 7.13: --auto-fix re-imports the macro into the
            # _FIXED.xlsm sibling, which plants the stamp as a side effect.
            auto_fixable=True,
        ))
        return findings
    if stamped.lower() != current.lower():
        findings.append(PreflightFinding(
            code="D17",
            severity="error",
            location="docProps/custom.xml",
            message=(
                f"embedded macro hash {stamped[:12]}... does not match "
                f"current SolveHeadless.bas {current[:12]}..."
            ),
            impact=(
                "The .bas in the repo has changed since this workbook was "
                "last re-imported. Body changes that don't add or remove "
                "function names are invisible to D15 but still break the "
                "orchestrator's assumptions (loop bounds, heartbeat keys, "
                "error handling). Re-import before solving."
            ),
            remediation=(
                f'Re-import the macro: python import_vba_module.py '
                f'"{workbook_path}"'
            ),
            # Per Tranche 7.13: --auto-fix re-imports the macro into the
            # _FIXED.xlsm sibling and re-checks D15/D17 post-import.
            # Original workbook is never mutated.
            auto_fixable=True,
        ))
    return findings


def check_input_bounds(wb: openpyxl.Workbook) -> list[PreflightFinding]:
    """E13/E14: pre-solve Dev Fee / NPP per project flagged against the
    macro's hardcoded sanity bounds. The non-chunked SolveHeadless entry
    point resets out-of-range values to seed at iter 0 AND inside the
    inner loop; the chunked SolveOneProjectByColHL only seeds blanks.
    Either path benefits from a heads-up when inputs are unusual.
    """
    findings: list[PreflightFinding] = []
    if "Project Inputs" not in wb.sheetnames:
        return findings
    ws = wb["Project Inputs"]

    npp_lo, npp_hi = NPP_BOUNDS
    df_lo, df_hi = DEV_FEE_BOUNDS

    out_of_band_dev_fee: list[tuple[str, float]] = []
    out_of_band_npp: list[tuple[str, float]] = []

    for col in iter_active_project_cols(ws):
        coord_letter = ws.cell(row=PI_ROW_DEV_FEE, column=col).coordinate

        df = ws.cell(row=PI_ROW_DEV_FEE, column=col).value
        if isinstance(df, (int, float)) and not (df_lo <= df <= df_hi):
            out_of_band_dev_fee.append((coord_letter, float(df)))

        npp_coord = ws.cell(row=PI_ROW_NPP, column=col).coordinate
        npp = ws.cell(row=PI_ROW_NPP, column=col).value
        if isinstance(npp, (int, float)) and not (npp_lo <= npp <= npp_hi):
            out_of_band_npp.append((npp_coord, float(npp)))

    if out_of_band_dev_fee:
        cells = ", ".join(f"{c}={v:.2f}" for c, v in out_of_band_dev_fee)
        findings.append(PreflightFinding(
            code="E13",
            severity="warning",  # SMP empirically converges with E13 firing.
            location=f"Project Inputs row {PI_ROW_DEV_FEE}",
            message=(
                f"pre-solve Dev Fee outside macro bounds [{df_lo}..{df_hi}] "
                f"on {len(out_of_band_dev_fee)} active project(s): {cells}"
            ),
            impact=(
                "Behavior depends on entry point. Chunked path "
                "(SolveOneProjectByColHL, default for --chunked runs) only "
                "seeds blanks — pre-existing out-of-range values pass "
                "through to GoalSeek untouched, so convergence is unaffected. "
                "Non-chunked path (SolveHeadless) resets to DEV_FEE_SEED "
                "($0.20) at iter 0 and on every inner GoalSeek if the value "
                "drifts back out of range — which can trap legitimate "
                "answers ('0 iter / NOT CONVERGED'). The bound is a sanity "
                "check, not a model constraint."
            ),
            remediation=(
                "If the run fails to converge, two options: (1) Manually "
                "set the Project Inputs Dev Fee cells to a value within "
                f"[${df_lo}..${df_hi}/W] and re-run (risky if Dev Fee is a "
                "deal-side input). (2) Ask the model owner to raise "
                "DEV_FEE_MAX in SolveHeadless.bas to encompass the natural "
                "pricing range, re-import the macro, and re-run."
            ),
            auto_fixable=False,
        ))

    if out_of_band_npp:
        cells = ", ".join(f"{c}={v:.3f}" for c, v in out_of_band_npp)
        findings.append(PreflightFinding(
            code="E14",
            severity="warning",
            location=f"Project Inputs row {PI_ROW_NPP}",
            message=(
                f"pre-solve NPP outside macro bounds [{npp_lo}..{npp_hi}] "
                f"on {len(out_of_band_npp)} active project(s): {cells}"
            ),
            impact=(
                "Macro resets these to NPP_SEED ($0.20) at iteration 0. NPP "
                "is solved within the inner loop (driven by Equity IRR "
                "GoalSeek), so a bad seed is usually recoverable — but "
                "convergence may take more iterations than usual."
            ),
            remediation=(
                "If recurring, raise NPP_MIN/NPP_MAX in SolveHeadless.bas "
                "to match the deal pipeline's natural NPP range."
            ),
            auto_fixable=False,
        ))

    return findings


def check_rate_component_config(wb: openpyxl.Workbook) -> list[PreflightFinding]:
    """E15: Rate Component source-mode consistency check.

    Caroline's RC audit memory (feedback_revenue_component_audit) documents
    the failure mode: a project column with an RC sub-block Toggle set to
    "Custom" but an empty Rate Name effectively zeros out a revenue stream
    that the corresponding "Generic" row would have produced. The Appraisal
    GoalSeek then converges to an inflated Dev Fee to make IRR=WACC hold,
    producing a mathematically valid but economically nonsensical result.

    Concrete: a portfolio where half the projects had RC5 Toggle=
    "Generic" with a real Rate Name + rate produced Dev Fees of
    $1.49-$1.96/W; the other half had RC5 Toggle="Custom" with empty
    Rate Name and converged to $4.30-$5.34/W. Same deal, same EPC,
    same IX — the only input difference was the RC5 sub-block config.

    Two flag categories:
      E15a — Custom-with-empty-name: Toggle="Custom" but Rate Name is None
             or whitespace. Almost always a misconfiguration (intentional
             custom rates would have a name). Severity: warning.
      E15b — Cross-project mismatch: same RC slot has different Toggle
             values across toggled-on projects in the same workbook. Can
             be intentional (e.g., one project sourced from a custom
             tariff schedule) but warrants explicit operator confirmation
             before solving. Severity: warning.

    Both are warnings, not errors, because the underlying configurations
    CAN be intentional. Caller decides via --strict-preflight whether to
    halt on warnings.
    """
    findings: list[PreflightFinding] = []
    if "Project Inputs" not in wb.sheetnames:
        return findings
    ws = wb["Project Inputs"]

    # Collect (rc_idx, col_letter, toggle, rate_name, generic_rate) for
    # every active project × every RC. Skip (col, rc) pairs where the
    # Equity master toggle (rows 150-155) is 0 — those RCs don't
    # contribute to revenue regardless of what the per-block sub-block
    # at rows 157-215 contains, so flagging them as "asymmetric" or
    # "Custom-empty" produces false positives. The model's per-block
    # cells often retain leftover template values when the master is
    # off, especially in IL CS deals where only RC1/RC2 are typically
    # active. Used for E15a, E15b, and E15c.
    per_project_rc: list[tuple[int, str, object, object, object]] = []
    active_cols: list[int] = []
    skipped_master_off = 0
    for col in iter_active_project_cols(ws):
        active_cols.append(col)
        col_letter = ws.cell(row=PI_ROW_DEV_FEE, column=col).coordinate.rstrip("0123456789")
        for rc_idx, base in enumerate(RC_BLOCK_BASES, start=1):
            master = ws.cell(row=RC_MASTER_TOGGLE_ROWS[rc_idx - 1], column=col).value
            # Master toggle is 0/1 per project. Treat anything that
            # isn't an explicit truthy numeric as off — None / blank /
            # FALSE all mean "RC inactive for Equity," so the sub-block
            # below is irrelevant.
            try:
                master_on = float(master) == 1.0
            except (TypeError, ValueError):
                master_on = False
            if not master_on:
                skipped_master_off += 1
                continue
            rate_name = ws.cell(row=base + RC_OFFSET_NAME, column=col).value
            rc_toggle = ws.cell(row=base + RC_OFFSET_TOGGLE, column=col).value
            generic_rate = ws.cell(row=base + RC_OFFSET_GEN_RATE, column=col).value
            per_project_rc.append((rc_idx, col_letter, rc_toggle, rate_name, generic_rate))

    if not active_cols:
        return findings  # no projects to check; covered by B9

    # E15a — Custom with empty Rate Name
    misconfigured: list[tuple[int, str]] = []
    for rc_idx, col_letter, rc_toggle, rate_name, _gr in per_project_rc:
        if rc_toggle != "Custom":
            continue
        if rate_name is None or not str(rate_name).strip():
            misconfigured.append((rc_idx, col_letter))

    if misconfigured:
        # Group by RC for a tighter message
        by_rc: dict[int, list[str]] = {}
        for rc_idx, col_letter in misconfigured:
            by_rc.setdefault(rc_idx, []).append(col_letter)
        details = "; ".join(
            f"RC{rc_idx} on cols [{', '.join(cols)}]"
            for rc_idx, cols in sorted(by_rc.items())
        )
        findings.append(PreflightFinding(
            code="E15a",
            severity="warning",
            location="Project Inputs Rate Component sub-blocks",
            message=(
                f"{len(misconfigured)} RC sub-block(s) have Toggle='Custom' "
                f"but empty Rate Name: {details}"
            ),
            impact=(
                "A Custom toggle without a Rate Name effectively zeros out "
                "that revenue stream — the corresponding Custom rate rows in "
                "the Rate Curves tab are likely empty too. The Appraisal "
                "GoalSeek will compensate by inflating Dev Fee until IRR=WACC "
                "still holds, producing an economically nonsensical solution "
                "(typical: $4-5/W Dev Fees on projects with RC5 Custom + empty "
                "Name). The macro will converge — but to the wrong equilibrium."
            ),
            remediation=(
                "For each flagged RC: either (1) flip Toggle to 'Generic' "
                "and populate the Generic rate row in Project Inputs, or "
                "(2) keep 'Custom' but populate Rate Name AND the per-project "
                "rate vector in the Rate Curves tab for that project column. "
                "Re-run the RC audit (RC1-RC6 active state + term length) "
                "across equity/debt/appraisal before solving."
            ),
            auto_fixable=False,
        ))

    # E15b — Cross-project toggle mismatch (same RC, different mode across cols)
    mismatched_rcs: list[tuple[int, dict[str, list[str]]]] = []
    for rc_idx in range(1, 7):
        by_toggle: dict[str, list[str]] = {}
        for r, col_letter, rc_toggle, _, _gr in per_project_rc:
            if r != rc_idx:
                continue
            key = str(rc_toggle) if rc_toggle is not None else "(none)"
            by_toggle.setdefault(key, []).append(col_letter)
        if len(by_toggle) > 1:
            mismatched_rcs.append((rc_idx, by_toggle))

    if mismatched_rcs:
        lines = []
        for rc_idx, by_toggle in mismatched_rcs:
            parts = "; ".join(
                f"{mode}=[{', '.join(cols)}]"
                for mode, cols in sorted(by_toggle.items())
            )
            lines.append(f"RC{rc_idx}: {parts}")
        findings.append(PreflightFinding(
            code="E15b",
            severity="warning",
            location="Project Inputs Rate Component sub-blocks",
            message=(
                f"{len(mismatched_rcs)} Rate Component(s) have mixed source "
                f"modes across active projects: " + " | ".join(lines)
            ),
            impact=(
                "Mixed Generic/Custom modes across projects in the same "
                "workbook CAN be intentional (one project has a bespoke "
                "tariff, others use the default curve), but the pattern "
                "frequently signals a half-applied edit. Combined with "
                "E15a, the same pattern produced the inflated Dev Fees "
                "in the documented portfolio failure."
            ),
            remediation=(
                "Confirm whether the mixed configuration is intentional. "
                "If not, flip the outlier projects to match the majority "
                "mode and verify the Rate Curves tab populates correctly."
            ),
            auto_fixable=False,
        ))

    # E15c — Asymmetric revenue. Fires as ERROR when the same RC has BOTH:
    #   (a) at least one project with Toggle=Generic AND non-zero Generic
    #       Rate (a real revenue stream included in cap-stack sizing), AND
    #   (b) at least one project with Toggle=Custom AND empty Rate Name
    #       (effectively zero revenue for that project on the same RC).
    # The Appraisal GoalSeek then compensates by inflating Dev Fee on
    # the (b) projects until IRR=WACC still holds, producing the $4-5/W
    # Dev Fees this check is designed to prevent shipping.
    #
    # Crucially: uniform Custom-empty across ALL active projects is NOT
    # this case — it just means the RC is intentionally disabled for the
    # whole portfolio. That's the SMP pattern, and the macro produces
    # correct output for it. E15c distinguishes the two by requiring (a)
    # to be non-empty before firing.
    # Revenue signal: Generic toggle + non-empty Rate Name. The Rate Name
    # being non-empty is the operator-authored signal that this RC is
    # contributing a real revenue stream on this project. We deliberately
    # do NOT inspect the Generic Rate (Project Inputs!base+3) value: under
    # openpyxl data_only=False that cell may be a formula string rather
    # than a literal, and the value-pass load happens later in run_preflight.
    # Rate Name presence is a robust proxy that works regardless of load mode.
    #
    # Zero signal: Custom toggle + empty Rate Name (and by implication an
    # empty per-project Custom rate vector in the Rate Curves tab).
    asymmetric: list[tuple[int, list[str], list[str]]] = []
    for rc_idx in range(1, 7):
        revenue_cols: list[str] = []
        zero_cols: list[str] = []
        for r, col_letter, rc_toggle, rate_name, _gr in per_project_rc:
            if r != rc_idx:
                continue
            name_is_set = rate_name is not None and bool(str(rate_name).strip())
            if rc_toggle == "Generic" and name_is_set:
                revenue_cols.append(col_letter)
            elif rc_toggle == "Custom" and not name_is_set:
                zero_cols.append(col_letter)
        if revenue_cols and zero_cols:
            asymmetric.append((rc_idx, revenue_cols, zero_cols))

    if asymmetric:
        lines = []
        for rc_idx, rev, zero in asymmetric:
            lines.append(
                f"RC{rc_idx}: revenue=[{', '.join(rev)}] vs zeroed=[{', '.join(zero)}]"
            )
        findings.append(PreflightFinding(
            code="E15c",
            severity="error",
            location="Project Inputs Rate Component sub-blocks",
            message=(
                f"{len(asymmetric)} Rate Component(s) have ASYMMETRIC "
                "revenue across active projects: " + " | ".join(lines)
            ),
            impact=(
                "Some projects have a real revenue stream (Generic mode with "
                "non-zero rate); others on the same RC are effectively zeroed "
                "(Custom mode with empty Rate Name -> empty Custom rate rows). "
                "Appraisal GoalSeek will converge for both sets, but the "
                "zeroed projects' Dev Fees inflate to compensate for the "
                "missing revenue (observed: $4-5/W Dev Fees on the zeroed "
                "subset vs $1.50-$2.00/W on the revenue subset, same EPC + IX). "
                "Mathematically valid, economically nonsensical."
            ),
            remediation=(
                "Decide whether the zeroed projects SHOULD have the same "
                "revenue stream. If yes, flip them to Generic with the same "
                "rate as the revenue projects. If no (genuinely no merchant "
                "revenue for those projects), populate the Custom rate rows "
                "in the Rate Curves tab with the actual project-specific "
                "rates. Re-run the RC audit before solving."
            ),
            auto_fixable=False,
        ))

    return findings


def check_placeholder_projects(wb: openpyxl.Workbook) -> list[PreflightFinding]:
    """E16: Detect active project columns that look like un-populated
    template placeholders.

    Failure mode this catches: a workbook with row 7 = 1 (active) on
    template-name columns ("Project 7" .. "Project 16") that share an
    identical placeholder shape — RC1 toggle = "Generic" with Generic
    Energy Rate at COD = 0. Zero revenue propagates #DIV/0! through the
    cap stack; pre-skip the macro's chunked entry point raises HRESULT
    0x800a9c68 (Type Mismatch from Range.GoalSeek on an error cell).
    Even with the macro's defensive guards a placeholder cannot
    meaningfully solve, so the right answer is to refuse it up front
    and surface the misconfiguration.

    Detection heuristic (must be conservative to avoid false-positives
    on small real projects with low merchant rates):
      - active row 7 = 1
      - RC1 toggle = "Generic" (the most common revenue stream slot)
      - RC1 Generic Energy Rate at COD = 0 or None or non-numeric

    Severity: warning (proceed with --strict-preflight to halt). The
    user may have intentionally left these inactive but with row 7 = 1
    (testing), in which case the warning is the right shape.
    """
    findings: list[PreflightFinding] = []
    if "Project Inputs" not in wb.sheetnames:
        return findings
    ws = wb["Project Inputs"]

    placeholders: list[tuple[str, str]] = []  # [(col_letter, project_name)]
    rc1_toggle_row = RC_BLOCK_BASES[0] + RC_OFFSET_TOGGLE      # 159
    rc1_gen_rate_row = RC_BLOCK_BASES[0] + RC_OFFSET_GEN_RATE  # 160
    for col in iter_active_project_cols(ws):
        rc1_toggle = ws.cell(row=rc1_toggle_row, column=col).value
        if rc1_toggle != "Generic":
            continue
        gen_rate = ws.cell(row=rc1_gen_rate_row, column=col).value
        # Zero, None, or non-numeric (formula text in data_only=False is OK;
        # we only flag explicit literal zero / None).
        is_zero = (
            gen_rate is None
            or gen_rate == 0
            or (isinstance(gen_rate, (int, float)) and gen_rate == 0)
        )
        if not is_zero:
            continue
        col_letter = ws.cell(row=PI_ROW_TOGGLE, column=col).coordinate.rstrip("0123456789")
        name = ws.cell(row=4, column=col).value or "(unnamed)"
        placeholders.append((col_letter, str(name)))

    if not placeholders:
        return findings

    detail = ", ".join(f"{lt}={nm}" for lt, nm in placeholders)
    findings.append(PreflightFinding(
        code="E16",
        severity="warning",
        location="Project Inputs row 7 (active toggle) + RC1 sub-block",
        message=(
            f"{len(placeholders)} active project column(s) look like un-"
            f"populated placeholders (RC1 toggle='Generic' with Generic "
            f"Energy Rate at COD = 0): {detail}"
        ),
        impact=(
            "These projects have no revenue stream on RC1 — the primary "
            "energy revenue slot. The macro's chunked GoalSeek hits "
            "#DIV/0! through the cap stack; defensive guards convert the "
            "crash into not-converged, but solving a no-revenue project "
            "produces a nonsensical NPP — there's nothing for Dev Fee / "
            "Equity IRR to balance against. A workbook with 10 such "
            "columns will not ship a single project's results."
        ),
        remediation=(
            "Either (1) populate the Generic Energy Rate at COD on row "
            f"{rc1_gen_rate_row} for each flagged column with the "
            "merchant or contracted rate the project sells into, or "
            "(2) toggle row 7 = 0 (inactive) on the placeholder columns "
            "so the solver skips them. Only solve projects with real "
            "revenue assumptions populated."
        ),
        auto_fixable=False,
    ))
    return findings


def run_preflight(workbook_path: Path | str) -> PreflightResult:
    """Run the full pre-flight pass against `workbook_path`.

    Pure function. Never mutates the input file. Returns a PreflightResult
    with categorized findings; caller decides how to act on them.

    Load budget — one value-pass and one formula-pass openpyxl load,
    shared across all checks:
      value-pass (data_only=True, read_only=True): cell-level error
        scan and C-tier critical-path checks
      formula-pass (data_only=False, keep_vba=True): A/B/E-tier checks
        needing calc properties, structure, or non-cached formula text
      conditional 0-1x extra formula-pass inside scan_workbook_errors
        only when the value-pass already found errors

    On a 13MB xlsm the two loads take ~3-4 min. --auto-import-macro
    runs preflight twice; orchestrator skips the redundant second pass
    by re-checking only D15/D17 directly.
    """
    p = Path(workbook_path)
    findings: list[PreflightFinding] = []

    # D-tier first (cheap zip-level reads of vbaProject.bin and docProps/
    # custom.xml) before paying any openpyxl load cost. Catches outdated-
    # macro workbooks that would also tend to trip downstream checks
    # confusingly. D17 (hash drift) is the strongest signal; D15 (function
    # presence) is its fallback for workbooks predating the stamp convention.
    findings.extend(check_macro_version(p))
    findings.extend(check_macro_hash(p))

    # Open the value-pass workbook ONCE. Shared between scan_workbook_errors
    # (cell-level error scan) and check_critical_path_errors (C-tier).
    # Failure to load here is fatal — bail with X0 + scan_failed validation.
    try:
        wb_v = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
    except Exception as exc:
        # Match the WorkbookValidation shape scan_workbook_errors would
        # have returned, so downstream consumers (orchestrator's strict-
        # validation gate) keep working unchanged.
        error_scan = WorkbookValidation(
            status="scan_failed",
            total_errors=0,
            total_formulas=0,
            error_summary={},
            error=f"Failed to load (values pass): {exc}",
        )
        findings.append(PreflightFinding(
            code="X0",
            severity="error",
            location="workbook",
            message=f"unable to open workbook: {exc}",
            impact="Pre-flight cannot proceed.",
            remediation=(
                "Verify the file exists, is not currently open in another "
                "Excel instance, and is a valid .xlsm/.xlsx."
            ),
            auto_fixable=False,
        ))
        return PreflightResult(
            workbook_path=str(p),
            findings=tuple(findings),
            error_scan=error_scan,
        )

    try:
        # Cell-level error scan — reuses the open value-pass handle.
        error_scan = scan_workbook_errors(p, wb_vals=wb_v)

        # Formula-pass load (single open, all formula-shape checks).
        wb_f = openpyxl.load_workbook(
            str(p), data_only=False, keep_vba=True, read_only=False,
        )
        try:
            findings.extend(check_calc_properties(wb_f))
            findings.extend(check_required_sheets(wb_f))
            findings.extend(check_required_cells(wb_f))
            findings.extend(check_workbook_protection(wb_f))
            findings.extend(check_active_projects(wb_f))
            findings.extend(check_input_bounds(wb_f))
            findings.extend(check_rate_component_config(wb_f))
            findings.extend(check_placeholder_projects(wb_f))
        finally:
            wb_f.close()

        # C-tier reuses the same value-pass handle. Order matters: scan
        # iterated every sheet via iter_rows already, but read_only ws.cell
        # / ws[coord] access re-streams from the part on demand, so the
        # specific-cell lookups here work after the full scan.
        findings.extend(check_critical_path_errors(wb_v, error_scan))
    finally:
        wb_v.close()

    return PreflightResult(
        workbook_path=str(p),
        findings=tuple(findings),
        error_scan=error_scan,
    )


# ---------------------------------------------------------------------------
# Auto-fix
# ---------------------------------------------------------------------------

_CALCPR_RE = re.compile(r"<calcPr[^>]*/?>")


def apply_auto_fixes(
    src: Path | str,
    dst: Path | str,
    findings: tuple[PreflightFinding, ...],
) -> tuple[Path, list[str]]:
    """Patch a copy of `src` to address auto-fixable findings.

    Returns (dst_path, applied_fix_codes). Findings without auto_fixable=True
    are skipped silently — caller is responsible for surfacing them as still-
    blocking errors after the fix pass.

    Bank-grade: never mutates the source file. Always writes a new file.
    Editing happens at the .xlsx/.zip layer rather than via openpyxl save
    to avoid round-tripping every formatted cell on a 13MB workbook (which
    risks unintended cosmetic changes to a heavily styled deliverable).
    """
    src_p = Path(src)
    dst_p = Path(dst)
    if dst_p.exists():
        dst_p.unlink()
    shutil.copy2(src_p, dst_p)

    applied: list[str] = []
    fixable_codes = {f.code for f in findings if f.auto_fixable}

    if "A1" in fixable_codes:
        _patch_iterate_delta_xml(dst_p, ITERATE_DELTA_CEILING)
        applied.append("A1")

    return dst_p, applied


def _patch_iterate_delta_xml(path: Path, value: float) -> None:
    """Add/overwrite iterateDelta="<value>" on the <calcPr/> tag in
    xl/workbook.xml of the .xlsm at `path`. In-place rewrite of the zip.
    """
    # Render value as Excel does: 0.0001 -> "1E-4". Match SMP convention.
    if value == 0.0001:
        rendered = "1E-4"
    elif value == 0.001:
        rendered = "1E-3"
    elif value == 0.00001:
        rendered = "1E-5"
    else:
        rendered = repr(value)

    tmp = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/workbook.xml":
                    text = data.decode("utf-8")

                    def add_or_replace(m: re.Match[str]) -> str:
                        tag = m.group()
                        # Strip any existing iterateDelta="..."
                        tag = re.sub(r' iterateDelta="[^"]*"', "", tag)
                        # Insert before the closing /> or >
                        if tag.endswith("/>"):
                            return tag[:-2] + f' iterateDelta="{rendered}"/>'
                        return tag[:-1] + f' iterateDelta="{rendered}">'

                    text = _CALCPR_RE.sub(add_or_replace, text, count=1)
                    data = text.encode("utf-8")
                zout.writestr(item, data)
    tmp.replace(path)


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def format_preflight_report(result: PreflightResult) -> str:
    """One-screen summary of a PreflightResult, suitable for log output."""
    lines: list[str] = []
    err_n = len(result.errors)
    warn_n = len(result.warnings)
    fix_n = len(result.auto_fixable)

    if result.ok and warn_n == 0:
        lines.append(f"  Pre-flight: PASS (0 findings)")
    else:
        status = "FAIL" if err_n else "WARN"
        lines.append(
            f"  Pre-flight: {status} - {err_n} error(s), {warn_n} warning(s), "
            f"{fix_n} auto-fixable"
        )

    for f in result.findings:
        bullet = "[ERROR]" if f.severity == "error" else "[WARN] "
        fix_tag = " (auto-fixable)" if f.auto_fixable else ""
        lines.append(f"  {bullet} {f.code} @ {f.location}{fix_tag}")
        lines.append(f"          {f.message}")
        lines.append(f"          impact: {f.impact}")
        lines.append(f"          fix:    {f.remediation}")

    return "\n".join(lines)
