"""dn38_solver.merge — Per-worker _SOLVED.xlsm merge strategies.

After parallel mode runs N workers, each writes its own
`<stem>_SOLVED_w{N}.xlsm` containing the converged columns it owned.
This package collapses those into one canonical `<stem>_SOLVED.xlsm`
next to the source workbook.

Two merge paths, in order of preference:

1. `merge_via_openpyxl` — the standard path. Loads the master worker's
   xlsm with `keep_vba=True`, copies each peer worker's per-project
   column cells (rows 32/33/38/39/371 on Project Inputs — Dev Fee, FMV,
   NPP $/W, NPP $, Min Equity DSCR Multiple) into the master via
   openpyxl, and saves. Rows 31 and 37 stay as their sticky-IF formulas
   (Tranche 7.12). Only the project-column cells are
   touched; portfolio aggregates are left as-is on the assumption that
   Excel will recalc them on next interactive open. Fast, no COM.

2. `merge_via_vba_fallback` — used when openpyxl can't round-trip the
   workbook's macro project (rare, but happens on certain VBA project
   structures). Opens the master in a fresh Excel COM session and calls
   `SolveHeadless.bas:StampConvergedValuesHL` for each peer worker's
   converged values. Excel handles the .xlsm save natively so macros
   stay intact.

The orchestration (try openpyxl → fall back to VBA → fall back to copy)
lives in `parallel_runner.run_parallel`; this package only owns the
individual merge implementations. Keeping orchestration there means the
worker-result aggregation, parent_tmp lifecycle, and post-merge
verification stay in one place.

Why this lives outside `com/`:
- `merge_via_openpyxl` doesn't touch COM at all.
- `merge_via_vba_fallback` does, but the contract is "given the master
  xlsm path and peer xlsm paths, produce a merged xlsm" — that's a
  workbook-level operation, not a COM-runner concern.
- Splitting these out drops `parallel_runner.py` from ~1300 lines to
  ~700 and lets the parallel-spawn logic be read in one screen.
"""
from __future__ import annotations

import contextlib
import logging
from pathlib import Path

from openpyxl.utils import column_index_from_string

from dn38_solver.com.direct_runner import (
    _capture_excel_proc,
    _run_macro_with_timeout,
)
from dn38_solver.com.vba_contract import STAMP_CONVERGED_VALUES, vba_call_str
from dn38_solver.types import SolveTask

log = logging.getLogger(__name__)

# The rows VBA hard-stamps at end of each project's solve. These are the
# only cells we copy between workers — everything else in each project's
# column was already correct in the master pre-merge (since each worker
# started from the same source) or is a portfolio aggregate that Excel
# will recalc on next interactive open.
#
# Rows 31 (Live Appraisal IRR) and 37 (Live Levered Pre-Tax IRR) are
# deliberately NOT in this set as of Tranche 7.12 (Caroline's 2026-05-19
# preference). They stay as their sticky-IF circular formulas
# (=IF(<col>2=$F$2, $F$<row>, <col><row>)) so the audit chain is
# preserved when opening the merged file. Row 371 (Min Equity DSCR
# Multiple) is hardcoded instead — locking that upstream input freezes
# debt sizing / equity / IRR all by formula, and a one-cell revert to
# ='PT Returns'!$F$129 puts Min Equity back into fully dynamic solve.
OUTPUT_ROWS: tuple[int, ...] = (32, 33, 38, 39, 371)


def _extract_worker_id(other_path: Path) -> int | None:
    """Parse the worker id off a per-worker SOLVED filename.

    Filenames are produced by `parallel_runner` as
    `<stem>_SOLVED_w{N}.xlsm`. Returns None if the suffix is malformed
    so the caller can skip the file rather than crashing the merge on
    one bad path.
    """
    stem = other_path.stem
    if "_w" not in stem:
        return None
    try:
        return int(stem.rsplit("_w", 1)[1])
    except ValueError:
        return None


def merge_via_openpyxl(
    *,
    master_src: Path,
    others: list[Path],
    final_path: Path,
    partitions: list[list[SolveTask]],
    master_worker_id: int,
) -> None:
    """Copy per-project converged column values from other workers' SOLVED
    workbooks into the master, then save to final_path.

    Only the project-column cells are copied (rows 32, 33, 38, 39, 371
    on Project Inputs — Dev Fee, FMV, NPP $/W, NPP $, Min Equity DSCR
    Multiple). Rows 31/37 stay as their sticky-IF formulas (Tranche
    7.12). The rest of the
    workbook in the master is left as-is; per Caroline's spec, portfolio
    aggregates may be stale and will refresh on next interactive open.

    keep_vba=True is critical — without it, openpyxl strips the macro
    project on save. If the master's vba_archive can't be loaded (rare
    structural issue with the macro project), raise so the caller falls
    back to the VBA-helper merge path; saving without macros would
    produce a silently broken .xlsm.
    """
    import openpyxl

    wb_master = openpyxl.load_workbook(str(master_src), keep_vba=True)
    if getattr(wb_master, "vba_archive", None) is None:
        wb_master.close()
        raise RuntimeError(
            "openpyxl could not load the workbook's VBA project; "
            "saving via openpyxl would strip macros. Falling back."
        )
    ws_pi_master = (
        wb_master["Project Inputs"]
        if "Project Inputs" in wb_master.sheetnames
        else None
    )

    if ws_pi_master is None:
        # Nothing to merge into — just save master to final_path
        wb_master.save(str(final_path))
        wb_master.close()
        return

    for other_path in others:
        if not other_path.exists():
            continue
        try:
            wb_other = openpyxl.load_workbook(
                str(other_path), keep_vba=True, data_only=True,
            )
        except Exception:
            continue
        ws_pi_other = (
            wb_other["Project Inputs"]
            if "Project Inputs" in wb_other.sheetnames
            else None
        )
        if ws_pi_other is None:
            wb_other.close()
            continue

        wid = _extract_worker_id(other_path)
        if wid is None or wid >= len(partitions):
            wb_other.close()
            continue

        # Copy the convergence-output cells for each project this worker owned
        for task in partitions[wid]:
            col_idx = column_index_from_string(task.project_col_letter)
            for row in OUTPUT_ROWS:
                src_val = ws_pi_other.cell(row=row, column=col_idx).value
                ws_pi_master.cell(row=row, column=col_idx).value = src_val

        wb_other.close()

    wb_master.save(str(final_path))
    wb_master.close()


def merge_via_vba_fallback(
    *,
    master_src: Path,
    others: list[Path],
    final_path: Path,
    partitions: list[list[SolveTask]],
    worker_results: dict[int, dict],
    master_worker_id: int,
) -> None:
    """VBA-helper merge path. Used when openpyxl can't round-trip the .xlsm.

    Opens the master in a fresh Excel COM session, reads converged column
    values from each peer worker's SOLVED file via openpyxl (read-only,
    no save), and calls `SolveHeadless.bas:StampConvergedValuesHL` via
    Application.Run to write them into the master. Excel handles the
    .xlsm save natively so the macro project stays intact.

    Requires SolveHeadless.bas's StampConvergedValuesHL to be present in
    the master workbook (it is, since master is a worker's own _SOLVED
    file and workers all imported the module before solving).
    """
    import openpyxl
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        with contextlib.suppress(Exception):
            excel.AutomationSecurity = 1
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False

        wb = excel.Workbooks.Open(str(master_src), ReadOnly=False, UpdateLinks=0)
        excel_proc = _capture_excel_proc(excel)  # for watchdog kill handle

        # Verify StampConvergedValuesHL is callable before the per-project
        # loop. Without this check, every per-project Application.Run
        # would raise, the except-and-continue at the bottom of the loop
        # would swallow each error silently, and we'd SaveAs a workbook
        # that looks right but has only worker-0's converged values.
        # Sourced from vba_contract so a future rename in the .bas surfaces
        # as a Python ImportError instead of a string-mismatch at runtime.
        try:
            vbp = wb.VBProject
            found = False
            for i in range(1, vbp.VBComponents.Count + 1):
                comp = vbp.VBComponents.Item(i)
                try:
                    cm = comp.CodeModule
                    if cm.Find(
                        STAMP_CONVERGED_VALUES.name, 1, 1,
                        cm.CountOfLines, 999, True, False, False,
                    ):
                        found = True
                        break
                except Exception:
                    continue
            if not found:
                raise RuntimeError(
                    f"{STAMP_CONVERGED_VALUES.name} not found in workbook VBA. "
                    "The module is required for the merge fallback path; "
                    "re-import SolveHeadless.bas via import_vba_module.py."
                )
        except Exception as verify_exc:
            log.error("  VBA merge precondition failed: %s", verify_exc)
            raise

        # Track every stamp failure and every project skipped due to missing
        # peer data. ANY of these means the merged file would silently ship
        # with $0 or stale values for those columns — that is an IC-facing
        # data-integrity event, not a "warn and continue" event.
        stamp_failures: list[tuple[int, str, str]] = []
        skipped_missing: list[tuple[int, str, list[str]]] = []

        for other_path in others:
            if not other_path.exists():
                continue
            wid = _extract_worker_id(other_path)
            if wid is None or wid >= len(partitions):
                continue

            # Read peer worker's converged values via openpyxl (read-only).
            try:
                wb_other = openpyxl.load_workbook(
                    str(other_path), data_only=True, read_only=True,
                )
            except Exception:
                continue
            ws_pi_other = (
                wb_other["Project Inputs"]
                if "Project Inputs" in wb_other.sheetnames
                else None
            )
            if ws_pi_other is None:
                wb_other.close()
                continue

            for task in partitions[wid]:
                col_idx = column_index_from_string(task.project_col_letter)
                npp = ws_pi_other.cell(row=38, column=col_idx).value
                dev_fee = ws_pi_other.cell(row=32, column=col_idx).value
                fmv = ws_pi_other.cell(row=33, column=col_idx).value
                npp_total = ws_pi_other.cell(row=39, column=col_idx).value
                # DSCR Multiple from row 371 — the upstream input that
                # locks the per-project debt sizing / equity / IRR
                # cascade. Replaces the prior reads of live_irr (row 37)
                # and appr_live (row 31), which are no longer stamped
                # by the worker (they stay as sticky-IF formulas per
                # Tranche 7.12).
                dscr_mult = ws_pi_other.cell(row=371, column=col_idx).value

                # Skip the entire row if ANY of the five values is missing.
                # Coercing None → 0.0 (prior behavior) silently writes a $0
                # Dev Fee or $0 NPP into the merged file, which downstream
                # IC summaries don't distinguish from a real zero.
                missing = [
                    name for name, val in (
                        ("npp", npp), ("dev_fee", dev_fee), ("fmv", fmv),
                        ("npp_total", npp_total), ("dscr_mult", dscr_mult),
                    ) if val is None
                ]
                if missing:
                    log.error(
                        "  SKIP merge for col %d (%s): peer missing %s",
                        col_idx, task.project_name, missing,
                    )
                    skipped_missing.append((col_idx, task.project_name, missing))
                    continue

                try:
                    _run_macro_with_timeout(
                        excel,
                        (
                            vba_call_str(wb.Name, STAMP_CONVERGED_VALUES),
                            int(col_idx),
                            float(npp), float(dev_fee), float(fmv),
                            float(npp_total), float(dscr_mult),
                        ),
                        timeout_sec=60,  # per-cell stamp is sub-second on healthy state
                        excel_proc=excel_proc,
                        label=f"StampConvergedValues[{task.project_name}]",
                    )
                except Exception as stamp_exc:
                    log.error(
                        "  %s failed for col %d (%s): %s",
                        STAMP_CONVERGED_VALUES.name,
                        col_idx, task.project_name, stamp_exc,
                    )
                    stamp_failures.append(
                        (col_idx, task.project_name, str(stamp_exc))
                    )
            wb_other.close()

        # Any stamp failure or missing-data skip means the workbook in
        # memory is NOT a complete merge. Save under a .MERGE_FAILED.xlsm
        # name so the file is forensically recoverable but unmistakably
        # not authoritative, and raise so the caller falls through to a
        # safer fallback (e.g. copy_master) instead of treating this as
        # a successful merge.
        if stamp_failures or skipped_missing:
            failed_path = final_path.with_suffix(".MERGE_FAILED.xlsm")
            wb.SaveAs(str(failed_path), FileFormat=52)
            raise RuntimeError(
                f"VBA merge incomplete: {len(stamp_failures)} stamp failure(s), "
                f"{len(skipped_missing)} project(s) skipped for missing peer data. "
                f"Forensic copy saved to {failed_path.name}. "
                "Caller should fall through to copy_master to avoid shipping "
                "a partial-merge .xlsm."
            )

        # SaveAs to final canonical path with explicit FileFormat=52
        # (xlOpenXMLWorkbookMacroEnabled) so the VBA project survives.
        wb.SaveAs(str(final_path), FileFormat=52)

    finally:
        if wb is not None:
            with contextlib.suppress(Exception):
                wb.Close(SaveChanges=False)
        if excel is not None:
            with contextlib.suppress(Exception):
                excel.Quit()
        with contextlib.suppress(Exception):
            pythoncom.CoUninitialize()
